"""
DATABASE MODULE - Khởi tạo và quản lý cơ sở dữ liệu SQLite
"""

import sqlite3
import os
import json
import threading
from functools import wraps
from datetime import datetime, date, timedelta
from pathlib import Path

DEFAULT_DB_PATH = 'data/accounting.db'
DB_CONFIG_PATH = 'data/db_path.txt'
DB_PATH = DEFAULT_DB_PATH
_local = threading.local()
BOOTSTRAP_VERSION = '2026-05-19-fast-init-1'


class ConnectionProxy:
    """Đại diện proxy cho kết nối SQLite, tự tái tạo nếu bị đóng."""

    def __init__(self, db_path):
        self._db_path = db_path
        self._conn = None

    def _ensure_connection(self):
        if self._conn is None:
            self._conn = _create_connection(self._db_path)
            return self._conn

        try:
            self._conn.execute('SELECT 1')
        except sqlite3.Error:
            try:
                self._conn.close()
            except sqlite3.Error:
                pass
            self._conn = _create_connection(self._db_path)

        return self._conn

    def close(self):
        if self._conn is not None:
            try:
                self._conn.close()
            except sqlite3.Error:
                pass
            self._conn = None

    def __getattr__(self, name):
        return getattr(self._ensure_connection(), name)


def get_database_path():
    """Lấy đường dẫn database hiện tại, hỗ trợ dùng chung qua LAN/OneDrive."""
    env_path = os.environ.get('ACCOUNTING_DB_PATH')
    if env_path:
        return env_path
    config_path = Path(DB_CONFIG_PATH)
    if config_path.exists():
        configured = config_path.read_text(encoding='utf-8').strip()
        if configured:
            return configured
    return DEFAULT_DB_PATH


def set_database_path(db_path):
    """Lưu đường dẫn database dùng chung cho các lần chạy sau."""
    path = Path(db_path)
    if path.suffix.lower() != '.db':
        raise ValueError('Vui lòng chọn file database có đuôi .db')
    path.parent.mkdir(parents=True, exist_ok=True)
    Path('data').mkdir(exist_ok=True)
    Path(DB_CONFIG_PATH).write_text(str(path), encoding='utf-8')
    close_connection()
    return str(path)


def _create_connection(db_path=None):
    """Tao ket noi SQLite moi voi WAL va busy timeout."""
    db_path = db_path or get_database_path()
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path, timeout=30)
    conn.row_factory = sqlite3.Row
    try:
        conn.execute('PRAGMA busy_timeout = 30000')
        conn.execute('PRAGMA journal_mode = WAL')
    except sqlite3.DatabaseError:
        pass
    return conn


def get_connection():
    """Lay ket noi database thread-local, tu tao lai khi doi DB path."""
    db_path = get_database_path()
    conn = getattr(_local, 'conn', None)
    if conn is not None and getattr(_local, 'db_path', None) == db_path:
        return conn
    conn = ConnectionProxy(db_path)
    _local.conn = conn
    _local.db_path = db_path
    return conn


def close_connection():
    """Dong ket noi thread-local cua thread hien tai."""
    conn = getattr(_local, 'conn', None)
    if conn is not None:
        try:
            conn.close()
        except sqlite3.Error:
            pass
        _local.conn = None
        _local.db_path = None


def _get_setting(cursor, key):
    cursor.execute('SELECT value FROM app_settings WHERE key = ?', (key,))
    row = cursor.fetchone()
    return row['value'] if row else None


def _set_setting(cursor, key, value):
    cursor.execute('''
        INSERT INTO app_settings (key, value, updated_at)
        VALUES (?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(key) DO UPDATE SET
            value = excluded.value,
            updated_at = CURRENT_TIMESTAMP
    ''', (key, value))


class ConnectionPerRequestMixin:
    """Open a short-lived SQLite connection for each public manager method.

    Existing manager code can keep using ``self.conn``. During a public method
    call, ``self.conn`` points at the request-scoped connection; nested public
    calls reuse it. The connection is committed and closed when the outermost
    call finishes, or rolled back on error.
    """

    @property
    def conn(self):
        active = getattr(self, '_request_conn', None)
        if active is not None:
            return active
        # Compatibility path for legacy UI code that still reaches into
        # manager.conn directly. New manager methods should go through the
        # method wrapper above.
        return get_connection()

    def __getattribute__(self, name):
        attr = object.__getattribute__(self, name)
        if name.startswith('_') or name in {'conn'} or not callable(attr):
            return attr
        if getattr(attr, '_connection_per_request_wrapped', False):
            return attr

        @wraps(attr)
        def wrapper(*args, **kwargs):
            depth = getattr(self, '_request_conn_depth', 0)
            if depth:
                object.__setattr__(self, '_request_conn_depth', depth + 1)
                try:
                    return attr(*args, **kwargs)
                finally:
                    object.__setattr__(self, '_request_conn_depth', depth)

            conn = _create_connection()
            object.__setattr__(self, '_request_conn', conn)
            object.__setattr__(self, '_request_conn_depth', 1)
            try:
                result = attr(*args, **kwargs)
                conn.commit()
                return result
            except Exception:
                try:
                    conn.rollback()
                except sqlite3.Error:
                    pass
                raise
            finally:
                object.__setattr__(self, '_request_conn_depth', 0)
                object.__setattr__(self, '_request_conn', None)
                try:
                    conn.close()
                except sqlite3.Error:
                    pass

        wrapper._connection_per_request_wrapped = True
        return wrapper


def init_database():
    """Khởi tạo các bảng database nếu chưa tồn tại."""
    conn = _create_connection()
    cursor = conn.cursor()

    # Bảng Người dùng
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            full_name TEXT,
            email TEXT,
            role TEXT DEFAULT 'employee',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            active INTEGER DEFAULT 1
        )
    ''')

    # Bảng Dự án
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            location TEXT,
            start_date DATE,
            end_date DATE,
            budget REAL,
            status TEXT DEFAULT 'active',
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    # Bảng Danh mục chi phí
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expense_categories (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            parent_id INTEGER,
            description TEXT,
            FOREIGN KEY(parent_id) REFERENCES expense_categories(id)
        )
    ''')

    # Bảng Chi phí
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_date DATE NOT NULL,
            project_id INTEGER,
            category_id INTEGER NOT NULL,
            description TEXT,
            amount REAL NOT NULL,
            paid_by TEXT,
            payment_method TEXT,
            status TEXT DEFAULT 'pending',
            notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(category_id) REFERENCES expense_categories(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    # Bảng Hóa đơn / Chứng từ
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_type TEXT NOT NULL,
            doc_number TEXT,
            doc_date DATE,
            einvoice_provider TEXT,
            einvoice_status TEXT,
            einvoice_tax_authority_code TEXT,
            einvoice_transaction_id TEXT,
            einvoice_payload TEXT,
            qr_token TEXT,
            supplier_id INTEGER,
            supplier_name TEXT,
            description TEXT,
            amount REAL,
            expense_id INTEGER,
            project_id INTEGER,
            category_id INTEGER,
            status TEXT DEFAULT 'draft',
            file_path TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(expense_id) REFERENCES expenses(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(category_id) REFERENCES expense_categories(id),
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS suppliers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_name TEXT UNIQUE NOT NULL,
            tax_code TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Bảng Liên kết file
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER,
            expense_id INTEGER,
            file_path TEXT NOT NULL,
            file_name TEXT,
            file_type TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(expense_id) REFERENCES expenses(id)
        )
    ''')

    # Bảng Vật tư
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL,
            unit TEXT,
            quantity INTEGER DEFAULT 0,
            unit_price REAL,
            average_cost REAL DEFAULT 0,
            min_quantity REAL DEFAULT 0,
            category TEXT,
            supplier TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Bảng Phiếu kho (nhập/xuất vật tư)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS inventory_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            material_id INTEGER NOT NULL,
            transaction_type TEXT NOT NULL,
            quantity INTEGER,
            transaction_date DATE,
            project_id INTEGER,
            notes TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(material_id) REFERENCES materials(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS construction_work_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            item_code TEXT,
            item_name TEXT NOT NULL,
            unit TEXT,
            planned_quantity REAL DEFAULT 0,
            completed_quantity REAL DEFAULT 0,
            percent_complete REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            status TEXT DEFAULT 'planned',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS site_diaries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            diary_date DATE NOT NULL,
            project_id INTEGER,
            weather TEXT,
            manpower TEXT,
            equipment TEXT,
            work_content TEXT,
            issues TEXT,
            reporter TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            milestone_name TEXT NOT NULL,
            planned_date DATE,
            actual_date DATE,
            status TEXT DEFAULT 'planned',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS equipment_usage (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usage_date DATE NOT NULL,
            project_id INTEGER,
            equipment_name TEXT NOT NULL,
            operator TEXT,
            hours REAL DEFAULT 0,
            fuel_cost REAL DEFAULT 0,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS safety_checks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            check_date DATE NOT NULL,
            project_id INTEGER,
            check_item TEXT NOT NULL,
            result TEXT DEFAULT 'Đạt',
            responsible TEXT,
            action_required TEXT,
            deadline DATE,
            status TEXT DEFAULT 'open',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    # Bảng Tài khoản kế toán
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_code TEXT UNIQUE NOT NULL,
            account_name TEXT NOT NULL,
            account_type TEXT,
            account_level INTEGER DEFAULT 1,
            parent_code TEXT,
            legal_basis TEXT,
            active INTEGER DEFAULT 1,
            description TEXT
        )
    ''')

    # Bảng Quy định/quy trình hồ sơ theo nghiệp vụ chi phí
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS compliance_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_code TEXT UNIQUE NOT NULL,
            expense_category_code TEXT,
            transaction_type TEXT NOT NULL,
            rule_name TEXT NOT NULL,
            required_documents TEXT NOT NULL,
            warning_message TEXT,
            legal_basis TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(expense_category_code) REFERENCES expense_categories(code)
        )
    ''')

    # Bảng kho biểu mẫu mặc định, có thể sửa/ẩn/thêm trong phần mềm
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS form_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            form_code TEXT NOT NULL,
            form_name TEXT NOT NULL,
            scope TEXT,
            source_workbook TEXT,
            sheet_name TEXT,
            file_path TEXT,
            used_when TEXT,
            required_signatures TEXT,
            optional_signatures TEXT,
            storage_owner TEXT,
            storage_method TEXT,
            usage_notes TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(form_code, scope, source_workbook, sheet_name)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS form_template_fields (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            form_template_id INTEGER,
            form_code TEXT NOT NULL,
            field_key TEXT NOT NULL,
            field_label TEXT NOT NULL,
            field_type TEXT DEFAULT 'text',
            required INTEGER DEFAULT 0,
            default_value TEXT,
            display_order INTEGER DEFAULT 0,
            notes TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(form_template_id) REFERENCES form_templates(id),
            UNIQUE(form_code, field_key)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS form_field_mappings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            form_template_id INTEGER,
            form_code TEXT NOT NULL,
            sheet_name TEXT,
            field_key TEXT NOT NULL,
            cell_address TEXT NOT NULL,
            row_mode TEXT DEFAULT 'fixed',
            active INTEGER DEFAULT 1,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(form_template_id) REFERENCES form_templates(id),
            UNIQUE(form_code, sheet_name, field_key)
        )
    ''')

    # Bảng yêu cầu hồ sơ/chứng từ theo tình huống/nghiệp vụ
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS document_requirements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ref_code TEXT,
            business_type TEXT NOT NULL,
            cost_group TEXT,
            record_type TEXT,
            scope TEXT,
            required_documents TEXT,
            optional_documents TEXT,
            required_signatures TEXT,
            optional_signatures TEXT,
            approval_authority TEXT,
            deadline TEXT,
            forms TEXT,
            limit_notes TEXT,
            warning_message TEXT,
            source_workbook TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(ref_code, business_type, source_workbook)
        )
    ''')

    # Bảng quy trình vận hành
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS process_steps (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            process_name TEXT NOT NULL,
            step_no TEXT,
            responsible TEXT,
            action TEXT NOT NULL,
            duration TEXT,
            forms TEXT,
            notes TEXT,
            source_workbook TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Bảng định mức/quy chế chi phí
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS policy_limits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            policy_group TEXT,
            item_name TEXT NOT NULL,
            value_a TEXT,
            value_b TEXT,
            value_c TEXT,
            value_d TEXT,
            notes TEXT,
            source_workbook TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(policy_group, item_name, source_workbook)
        )
    ''')

    # Bảng nhắc việc định kỳ
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS recurring_tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_text TEXT NOT NULL,
            task_type TEXT,
            task_content TEXT NOT NULL,
            owner TEXT,
            approver TEXT,
            forms TEXT,
            priority TEXT,
            status TEXT,
            notes TEXT,
            source_workbook TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(schedule_text, task_type, task_content, source_workbook)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS category_account_mappings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category_id INTEGER UNIQUE NOT NULL,
            debit_account TEXT NOT NULL,
            credit_account TEXT NOT NULL,
            notes TEXT,
            active INTEGER DEFAULT 1,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(category_id) REFERENCES expense_categories(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS approval_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expense_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            actor TEXT,
            note TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(expense_id) REFERENCES expenses(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS simple_catalogs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            catalog_type TEXT NOT NULL,
            name TEXT NOT NULL,
            description TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(catalog_type, name)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS schema_migrations (
            name TEXT PRIMARY KEY,
            applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS document_sequences (
            sequence_key TEXT PRIMARY KEY,
            prefix TEXT NOT NULL,
            period TEXT NOT NULL,
            last_number INTEGER DEFAULT 0,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS imported_invoice_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER,
            source_file TEXT NOT NULL,
            tax_code TEXT,
            invoice_number TEXT,
            invoice_date DATE,
            received_date DATE NOT NULL,
            total_amount REAL DEFAULT 0,
            deadline DATE,
            raw_text TEXT,
            status TEXT DEFAULT 'draft',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(document_id) REFERENCES documents(id),
            UNIQUE(source_file, invoice_number)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_code TEXT UNIQUE,
            full_name TEXT NOT NULL,
            department TEXT,
            position TEXT,
            phone TEXT,
            email TEXT,
            telegram_chat_id TEXT,
            active INTEGER DEFAULT 1
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS advance_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            advance_number TEXT UNIQUE NOT NULL,
            requester_id INTEGER,
            project_id INTEGER,
            request_date DATE NOT NULL,
            received_date DATE,
            amount REAL NOT NULL DEFAULT 0,
            purpose TEXT,
            status TEXT DEFAULT 'draft',
            has_requester_signature INTEGER DEFAULT 0,
            has_department_signature INTEGER DEFAULT 0,
            has_accounting_signature INTEGER DEFAULT 0,
            has_director_signature INTEGER DEFAULT 0,
            transferred_to_accounting_at TIMESTAMP,
            deadline DATE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(requester_id) REFERENCES employees(id),
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS advance_settlements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            advance_id INTEGER NOT NULL,
            settlement_number TEXT UNIQUE,
            settlement_date DATE NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            settled_expense_total REAL DEFAULT 0,
            returned_to_fund REAL DEFAULT 0,
            overspend_amount REAL DEFAULT 0,
            settlement_type TEXT DEFAULT 'normal',
            approved_by INTEGER,
            approved_at TIMESTAMP,
            status TEXT DEFAULT 'draft',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(advance_id) REFERENCES advance_requests(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS advance_attachments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            advance_id INTEGER,
            settlement_id INTEGER,
            document_id INTEGER,
            file_path TEXT NOT NULL,
            file_name TEXT,
            document_type TEXT,
            uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(advance_id) REFERENCES advance_requests(id),
            FOREIGN KEY(settlement_id) REFERENCES advance_settlements(id),
            FOREIGN KEY(document_id) REFERENCES documents(id),
            UNIQUE(file_path)
        )
    ''')

    # Bảng Sổ nhật ký
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS journal_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_number TEXT,
            entry_date DATE NOT NULL,
            fiscal_year INTEGER,
            fiscal_period TEXT,
            entry_type TEXT DEFAULT 'manual',
            is_reversed INTEGER DEFAULT 0,
            reversed_by INTEGER,
            reversal_of_entry_id INTEGER,
            posted_by INTEGER,
            posted_at TIMESTAMP,
            description TEXT,
            debit_account TEXT,
            credit_account TEXT,
            amount REAL,
            expense_id INTEGER,
            project_id INTEGER,
            contract_id INTEGER,
            reference_type TEXT,
            reference_id INTEGER,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(expense_id) REFERENCES expenses(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(reversal_of_entry_id) REFERENCES journal_entries(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS journal_entry_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            journal_entry_id INTEGER NOT NULL,
            line_no INTEGER NOT NULL,
            account_code TEXT NOT NULL,
            debit_amount REAL DEFAULT 0,
            credit_amount REAL DEFAULT 0,
            project_id INTEGER,
            expense_id INTEGER,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(journal_entry_id) REFERENCES journal_entries(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(expense_id) REFERENCES expenses(id),
            UNIQUE(journal_entry_id, line_no)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settlement_expense_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            settlement_id INTEGER NOT NULL,
            expense_id INTEGER NOT NULL,
            amount REAL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(settlement_id) REFERENCES advance_settlements(id),
            FOREIGN KEY(expense_id) REFERENCES expenses(id),
            UNIQUE(settlement_id, expense_id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT NOT NULL,
            entity_id INTEGER,
            action TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            actor_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS approval_thresholds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            role TEXT UNIQUE NOT NULL,
            max_amount REAL NOT NULL,
            can_final_approve INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS fixed_assets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_code TEXT UNIQUE NOT NULL,
            asset_name TEXT NOT NULL,
            asset_type TEXT,
            acquisition_date DATE,
            acquisition_cost REAL DEFAULT 0,
            useful_life_months INTEGER DEFAULT 0,
            depreciation_method TEXT DEFAULT 'straight_line',
            salvage_value REAL DEFAULT 0,
            accumulated_depreciation REAL DEFAULT 0,
            project_id INTEGER,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS asset_depreciation_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            asset_id INTEGER NOT NULL,
            period TEXT NOT NULL,
            depreciation_amount REAL DEFAULT 0,
            project_id INTEGER,
            journal_entry_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(asset_id) REFERENCES fixed_assets(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(journal_entry_id) REFERENCES journal_entries(id),
            UNIQUE(asset_id, period)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS overhead_allocations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period TEXT NOT NULL,
            basis TEXT NOT NULL,
            source_project_id INTEGER,
            target_project_id INTEGER NOT NULL,
            amount REAL DEFAULT 0,
            journal_entry_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(source_project_id) REFERENCES projects(id),
            FOREIGN KEY(target_project_id) REFERENCES projects(id),
            FOREIGN KEY(journal_entry_id) REFERENCES journal_entries(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS expiring_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            item_type TEXT NOT NULL,
            item_name TEXT NOT NULL,
            reference_no TEXT,
            project_id INTEGER,
            expiry_date DATE NOT NULL,
            owner TEXT,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ar_ap_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_type TEXT NOT NULL,
            partner_name TEXT NOT NULL,
            project_id INTEGER,
            doc_id INTEGER,
            due_date DATE NOT NULL,
            amount REAL NOT NULL,
            paid_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'open',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(doc_id) REFERENCES documents(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS qs_reconciliation_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            work_item_id INTEGER,
            original_budget REAL DEFAULT 0,
            revised_budget REAL DEFAULT 0,
            actual_cost REAL DEFAULT 0,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(work_item_id) REFERENCES construction_work_items(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS template_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            template_name TEXT NOT NULL,
            version_no INTEGER NOT NULL,
            file_path TEXT NOT NULL,
            checksum TEXT,
            effective_from DATE,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(template_name, version_no)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS site_diary_expense_suggestions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            diary_id INTEGER NOT NULL,
            project_id INTEGER,
            suggestion_type TEXT,
            description TEXT,
            amount REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            expense_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(diary_id) REFERENCES site_diaries(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(expense_id) REFERENCES expenses(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS check_register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            check_no TEXT UNIQUE NOT NULL,
            bank_account TEXT,
            issue_date DATE,
            due_date DATE,
            payee TEXT,
            amount REAL DEFAULT 0,
            status TEXT DEFAULT 'issued',
            expense_id INTEGER,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(expense_id) REFERENCES expenses(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS poc_revenue_recognitions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            period TEXT NOT NULL,
            contract_value REAL DEFAULT 0,
            previous_percent REAL DEFAULT 0,
            current_percent REAL DEFAULT 0,
            revenue_amount REAL DEFAULT 0,
            journal_entry_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(journal_entry_id) REFERENCES journal_entries(id),
            UNIQUE(project_id, period)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS currency_rates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            currency TEXT NOT NULL,
            rate_date DATE NOT NULL,
            exchange_rate REAL NOT NULL,
            source TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(currency, rate_date)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS foreign_currency_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transaction_type TEXT NOT NULL,
            reference_id INTEGER,
            currency TEXT NOT NULL,
            foreign_amount REAL NOT NULL,
            exchange_rate REAL NOT NULL,
            local_amount REAL NOT NULL,
            transaction_date DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS vendor_scorecards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            supplier_id INTEGER,
            supplier_name TEXT NOT NULL,
            period TEXT,
            price_score REAL DEFAULT 0,
            quality_score REAL DEFAULT 0,
            delivery_score REAL DEFAULT 0,
            document_score REAL DEFAULT 0,
            violation_notes TEXT,
            status TEXT DEFAULT 'preferred',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS fiscal_calendar (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fiscal_year INTEGER NOT NULL,
            fiscal_period TEXT NOT NULL UNIQUE,
            period_start DATE NOT NULL,
            period_end DATE NOT NULL,
            is_closed INTEGER DEFAULT 0,
            is_locked INTEGER DEFAULT 0,
            locked_at TIMESTAMP,
            locked_by INTEGER
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS powerbi_sync_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dataset_name TEXT NOT NULL,
            synced_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'success',
            row_count INTEGER DEFAULT 0,
            message TEXT
        )
    ''')

    # Hợp đồng dự án (thi công / thầu phụ / cung cấp)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_contracts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            contract_type TEXT NOT NULL,
            contract_no TEXT,
            partner_name TEXT,
            signed_date DATE,
            contract_value REAL DEFAULT 0,
            vat_rate REAL DEFAULT 10,
            retention_rate REAL DEFAULT 5,
            advance_received REAL DEFAULT 0,
            advance_paid REAL DEFAULT 0,
            start_date DATE,
            end_date DATE,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    # Nghiệm thu / đợt thanh toán hợp đồng
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS contract_billings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER NOT NULL,
            billing_date DATE NOT NULL,
            milestone_name TEXT,
            quantity_or_percent REAL,
            amount_before_vat REAL DEFAULT 0,
            vat_amount REAL DEFAULT 0,
            retention_amount REAL DEFAULT 0,
            net_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            document_id INTEGER,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(contract_id) REFERENCES project_contracts(id),
            FOREIGN KEY(document_id) REFERENCES documents(id)
        )
    ''')

    # Dự toán chi phí theo loại
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_cost_plans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            category_id INTEGER NOT NULL,
            planned_amount REAL DEFAULT 0,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(project_id, category_id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(category_id) REFERENCES expense_categories(id)
        )
    ''')

    # Doanh thu dự án
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_revenues (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            contract_id INTEGER,
            billing_id INTEGER,
            revenue_date DATE NOT NULL,
            amount REAL DEFAULT 0,
            vat_amount REAL DEFAULT 0,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(contract_id) REFERENCES project_contracts(id),
            FOREIGN KEY(billing_id) REFERENCES contract_billings(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS material_standards (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_item_id INTEGER,
            material_id INTEGER NOT NULL,
            basis_unit TEXT DEFAULT 'm2',
            standard_qty_per_unit REAL NOT NULL DEFAULT 0,
            tolerance_percent REAL DEFAULT 15,
            notes TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(work_item_id) REFERENCES construction_work_items(id),
            FOREIGN KEY(material_id) REFERENCES materials(id),
            UNIQUE(work_item_id, material_id, basis_unit)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS timesheets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            project_id INTEGER NOT NULL,
            work_item_id INTEGER,
            work_date DATE NOT NULL,
            work_days REAL DEFAULT 1,
            quantity_completed REAL DEFAULT 0,
            daily_rate REAL DEFAULT 0,
            piece_rate REAL DEFAULT 0,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(employee_id) REFERENCES employees(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(work_item_id) REFERENCES construction_work_items(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payroll_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_period TEXT NOT NULL UNIQUE,
            run_date DATE NOT NULL,
            gross_amount REAL DEFAULT 0,
            pit_amount REAL DEFAULT 0,
            net_amount REAL DEFAULT 0,
            journal_entry_id INTEGER,
            status TEXT DEFAULT 'draft',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(journal_entry_id) REFERENCES journal_entries(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payroll_run_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_run_id INTEGER NOT NULL,
            employee_id INTEGER,
            project_id INTEGER,
            work_days REAL DEFAULT 0,
            quantity_completed REAL DEFAULT 0,
            gross_amount REAL DEFAULT 0,
            pit_amount REAL DEFAULT 0,
            net_amount REAL DEFAULT 0,
            FOREIGN KEY(payroll_run_id) REFERENCES payroll_runs(id),
            FOREIGN KEY(employee_id) REFERENCES employees(id),
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS contract_payment_milestones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER NOT NULL,
            milestone_name TEXT NOT NULL,
            planned_date DATE,
            planned_amount REAL DEFAULT 0,
            actual_billing_id INTEGER,
            status TEXT DEFAULT 'planned',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(contract_id) REFERENCES project_contracts(id),
            FOREIGN KEY(actual_billing_id) REFERENCES contract_billings(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS guarantee_bonds (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER NOT NULL,
            milestone_id INTEGER,
            bond_type TEXT NOT NULL,
            bond_number TEXT,
            issuer TEXT,
            amount REAL DEFAULT 0,
            issue_date DATE,
            expiry_date DATE,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(contract_id) REFERENCES project_contracts(id),
            FOREIGN KEY(milestone_id) REFERENCES contract_payment_milestones(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS warranty_periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            contract_id INTEGER NOT NULL,
            milestone_id INTEGER,
            warranty_scope TEXT,
            start_date DATE,
            end_date DATE,
            retention_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'active',
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(contract_id) REFERENCES project_contracts(id),
            FOREIGN KEY(milestone_id) REFERENCES contract_payment_milestones(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS user_project_access (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            project_id INTEGER NOT NULL,
            access_level TEXT DEFAULT 'view',
            can_view INTEGER DEFAULT 1,
            can_edit INTEGER DEFAULT 0,
            can_approve INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            UNIQUE(user_id, project_id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bank_statement_rows (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_account TEXT,
            transaction_date DATE NOT NULL,
            description TEXT,
            amount REAL NOT NULL,
            reference_no TEXT,
            source_name TEXT,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(bank_account, transaction_date, amount, reference_no, description)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bank_reconciliation_matches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_row_id INTEGER NOT NULL,
            system_record_type TEXT NOT NULL,
            system_record_id INTEGER NOT NULL,
            match_status TEXT DEFAULT 'auto',
            confidence INTEGER DEFAULT 0,
            matched_by TEXT,
            matched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY(bank_row_id) REFERENCES bank_statement_rows(id),
            UNIQUE(bank_row_id),
            UNIQUE(system_record_type, system_record_id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS purchase_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            po_number TEXT UNIQUE NOT NULL,
            supplier_id INTEGER,
            supplier_name TEXT,
            order_date DATE NOT NULL,
            expected_date DATE,
            status TEXT DEFAULT 'draft',
            total_amount REAL DEFAULT 0,
            approved_by INTEGER,
            approved_at TIMESTAMP,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY(approved_by) REFERENCES users(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS purchase_order_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_order_id INTEGER NOT NULL,
            material_id INTEGER,
            description TEXT,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            received_quantity REAL DEFAULT 0,
            project_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(purchase_order_id) REFERENCES purchase_orders(id),
            FOREIGN KEY(material_id) REFERENCES materials(id),
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bank_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            account_number TEXT UNIQUE NOT NULL,
            bank_name TEXT,
            account_name TEXT,
            opening_balance REAL DEFAULT 0,
            current_balance REAL DEFAULT 0,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bank_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_account_id INTEGER,
            transaction_date DATE NOT NULL,
            description TEXT,
            debit_amount REAL DEFAULT 0,
            credit_amount REAL DEFAULT 0,
            amount REAL DEFAULT 0,
            reference_no TEXT,
            matched_record_type TEXT,
            matched_record_id INTEGER,
            reconciled INTEGER DEFAULT 0,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(bank_account_id) REFERENCES bank_accounts(id),
            UNIQUE(bank_account_id, transaction_date, amount, reference_no, description)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS bank_reconciliations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bank_transaction_id INTEGER NOT NULL,
            system_record_type TEXT NOT NULL,
            system_record_id INTEGER NOT NULL,
            confidence INTEGER DEFAULT 0,
            status TEXT DEFAULT 'matched',
            reconciled_by INTEGER,
            reconciled_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT,
            FOREIGN KEY(bank_transaction_id) REFERENCES bank_transactions(id),
            UNIQUE(bank_transaction_id),
            UNIQUE(system_record_type, system_record_id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payroll_periods (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period TEXT NOT NULL UNIQUE,
            start_date DATE NOT NULL,
            end_date DATE NOT NULL,
            gross_amount REAL DEFAULT 0,
            insurance_amount REAL DEFAULT 0,
            pit_amount REAL DEFAULT 0,
            net_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            approved_by INTEGER,
            approved_at TIMESTAMP,
            payroll_run_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(approved_by) REFERENCES users(id),
            FOREIGN KEY(payroll_run_id) REFERENCES payroll_runs(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS payroll_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            payroll_period_id INTEGER NOT NULL,
            employee_id INTEGER,
            project_id INTEGER,
            work_days REAL DEFAULT 0,
            quantity_completed REAL DEFAULT 0,
            gross_amount REAL DEFAULT 0,
            bhxh_employee REAL DEFAULT 0,
            bhyt_employee REAL DEFAULT 0,
            bhtn_employee REAL DEFAULT 0,
            pit_amount REAL DEFAULT 0,
            net_amount REAL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(payroll_period_id) REFERENCES payroll_periods(id),
            FOREIGN KEY(employee_id) REFERENCES employees(id),
            FOREIGN KEY(project_id) REFERENCES projects(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS budget_versions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            version_no TEXT NOT NULL,
            name TEXT,
            status TEXT DEFAULT 'draft',
            approved_by INTEGER,
            approved_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(project_id) REFERENCES projects(id),
            UNIQUE(project_id, version_no)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS budget_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            budget_version_id INTEGER NOT NULL,
            work_item_id INTEGER,
            cost_category TEXT,
            description TEXT,
            quantity REAL DEFAULT 0,
            unit_price REAL DEFAULT 0,
            budget_amount REAL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(budget_version_id) REFERENCES budget_versions(id),
            FOREIGN KEY(work_item_id) REFERENCES construction_work_items(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subcontractors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            tax_code TEXT,
            phone TEXT,
            email TEXT,
            address TEXT,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS subcontract_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subcontractor_id INTEGER NOT NULL,
            project_id INTEGER,
            contract_id INTEGER,
            payment_date DATE NOT NULL,
            amount REAL DEFAULT 0,
            description TEXT,
            status TEXT DEFAULT 'draft',
            journal_entry_id INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(subcontractor_id) REFERENCES subcontractors(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(contract_id) REFERENCES project_contracts(id),
            FOREIGN KEY(journal_entry_id) REFERENCES journal_entries(id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS tax_declarations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tax_type TEXT NOT NULL,
            period TEXT NOT NULL,
            start_date DATE,
            end_date DATE,
            output_tax REAL DEFAULT 0,
            input_tax REAL DEFAULT 0,
            payable_amount REAL DEFAULT 0,
            status TEXT DEFAULT 'draft',
            file_path TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(tax_type, period)
        )
    ''')

    _migrate_database(cursor)
    _relax_document_number_unique(cursor)
    _prepare_duplicate_invoice_controls(cursor)
    _populate_fiscal_calendar(cursor)

    bootstrap_current = _get_setting(cursor, 'bootstrap_version') == BOOTSTRAP_VERSION
    force_bootstrap = os.environ.get('FASTRACK_FORCE_BOOTSTRAP', '').lower() in ('1', 'true', 'yes', 'on')
    conn.commit()
    conn.close()

    if bootstrap_current and not force_bootstrap:
        print("Database initialized successfully")
        return

    add_default_projects()
    add_default_expense_categories()
    add_default_accounts()
    add_default_account_mappings()
    add_default_simple_catalogs()
    add_default_company_settings()
    add_default_admin_templates()
    create_default_document_folders()
    add_default_compliance_rules()
    add_default_project_accounting_data()
    import_default_knowledge_from_templates()
    try:
        from modules.template_library import (
            extract_workbook_sheets_to_templates,
            create_construction_template_workbooks,
            infer_default_field_mappings,
        )
        extract_workbook_sheets_to_templates()
        create_construction_template_workbooks()
        infer_default_field_mappings()
    except Exception as exc:
        print(f"Skip template file generation: {exc}")

    conn = _create_connection()
    cursor = conn.cursor()
    _set_setting(cursor, 'bootstrap_version', BOOTSTRAP_VERSION)
    conn.commit()
    conn.close()

    print("Database initialized successfully")


def _migrate_database(cursor):
    """Bổ sung cột cho database cũ mà không làm mất dữ liệu."""
    table_columns = {}

    def has_column(table_name, column_name):
        if table_name not in table_columns:
            cursor.execute(f'PRAGMA table_info({table_name})')
            table_columns[table_name] = {row[1] for row in cursor.fetchall()}
        return column_name in table_columns[table_name]

    migrations = [
        ('documents', 'expense_id', 'ALTER TABLE documents ADD COLUMN expense_id INTEGER'),
        ('documents', 'updated_at', 'ALTER TABLE documents ADD COLUMN updated_at TIMESTAMP'),
        ('documents', 'supplier_id', 'ALTER TABLE documents ADD COLUMN supplier_id INTEGER'),
        ('documents', 'einvoice_provider', 'ALTER TABLE documents ADD COLUMN einvoice_provider TEXT'),
        ('documents', 'einvoice_status', 'ALTER TABLE documents ADD COLUMN einvoice_status TEXT'),
        ('documents', 'einvoice_tax_authority_code', 'ALTER TABLE documents ADD COLUMN einvoice_tax_authority_code TEXT'),
        ('documents', 'einvoice_transaction_id', 'ALTER TABLE documents ADD COLUMN einvoice_transaction_id TEXT'),
        ('documents', 'einvoice_payload', 'ALTER TABLE documents ADD COLUMN einvoice_payload TEXT'),
        ('documents', 'qr_token', 'ALTER TABLE documents ADD COLUMN qr_token TEXT'),
        ('documents', 'vat_rate', 'ALTER TABLE documents ADD COLUMN vat_rate REAL DEFAULT 10'),
        ('attachments', 'expense_id', 'ALTER TABLE attachments ADD COLUMN expense_id INTEGER'),
        ('accounts', 'account_level', 'ALTER TABLE accounts ADD COLUMN account_level INTEGER DEFAULT 1'),
        ('accounts', 'parent_code', 'ALTER TABLE accounts ADD COLUMN parent_code TEXT'),
        ('accounts', 'legal_basis', 'ALTER TABLE accounts ADD COLUMN legal_basis TEXT'),
        ('accounts', 'active', 'ALTER TABLE accounts ADD COLUMN active INTEGER DEFAULT 1'),
        ('accounts', 'account_class', 'ALTER TABLE accounts ADD COLUMN account_class TEXT'),
        ('accounts', 'normal_balance', 'ALTER TABLE accounts ADD COLUMN normal_balance TEXT'),
        ('accounts', 'is_cost_account', 'ALTER TABLE accounts ADD COLUMN is_cost_account INTEGER DEFAULT 0'),
        ('expenses', 'department', 'ALTER TABLE expenses ADD COLUMN department TEXT'),
        ('expenses', 'purpose', 'ALTER TABLE expenses ADD COLUMN purpose TEXT'),
        ('expenses', 'item_list', 'ALTER TABLE expenses ADD COLUMN item_list TEXT'),
        ('expenses', 'accounting_staff', 'ALTER TABLE expenses ADD COLUMN accounting_staff TEXT'),
        ('expenses', 'department_head', 'ALTER TABLE expenses ADD COLUMN department_head TEXT'),
        ('expenses', 'prepared_by', 'ALTER TABLE expenses ADD COLUMN prepared_by TEXT'),
        ('expenses', 'attachments', 'ALTER TABLE expenses ADD COLUMN attachments TEXT'),
        ('expenses', 'work_item_id', 'ALTER TABLE expenses ADD COLUMN work_item_id INTEGER'),
        ('expenses', 'contract_id', 'ALTER TABLE expenses ADD COLUMN contract_id INTEGER'),
        ('expenses', 'debit_account_id', 'ALTER TABLE expenses ADD COLUMN debit_account_id INTEGER'),
        ('expenses', 'credit_account_id', 'ALTER TABLE expenses ADD COLUMN credit_account_id INTEGER'),
        ('expenses', 'fiscal_period', 'ALTER TABLE expenses ADD COLUMN fiscal_period TEXT'),
        ('expenses', 'advance_request_id', 'ALTER TABLE expenses ADD COLUMN advance_request_id INTEGER'),
        ('expenses', 'invoice_id', 'ALTER TABLE expenses ADD COLUMN invoice_id INTEGER'),
        ('expenses', 'is_ocr_imported', 'ALTER TABLE expenses ADD COLUMN is_ocr_imported INTEGER DEFAULT 0'),
        ('materials', 'average_cost', 'ALTER TABLE materials ADD COLUMN average_cost REAL DEFAULT 0'),
        ('materials', 'min_quantity', 'ALTER TABLE materials ADD COLUMN min_quantity REAL DEFAULT 0'),
        ('employees', 'dependents', 'ALTER TABLE employees ADD COLUMN dependents INTEGER DEFAULT 0'),
        ('employees', 'base_salary', 'ALTER TABLE employees ADD COLUMN base_salary REAL DEFAULT 0'),
        ('construction_work_items', 'percent_complete', 'ALTER TABLE construction_work_items ADD COLUMN percent_complete REAL DEFAULT 0'),
        ('projects', 'contract_value', 'ALTER TABLE projects ADD COLUMN contract_value REAL DEFAULT 0'),
        ('projects', 'owner_name', 'ALTER TABLE projects ADD COLUMN owner_name TEXT'),
        ('projects', 'project_type', 'ALTER TABLE projects ADD COLUMN project_type TEXT DEFAULT "construction"'),
        ('projects', 'project_manager_id', 'ALTER TABLE projects ADD COLUMN project_manager_id INTEGER'),
        ('projects', 'cost_center_code', 'ALTER TABLE projects ADD COLUMN cost_center_code TEXT'),
        ('projects', 'phase', 'ALTER TABLE projects ADD COLUMN phase TEXT'),
        ('projects', 'actual_end_date', 'ALTER TABLE projects ADD COLUMN actual_end_date DATE'),
        ('advance_settlements', 'settled_expense_total', 'ALTER TABLE advance_settlements ADD COLUMN settled_expense_total REAL DEFAULT 0'),
        ('advance_settlements', 'returned_to_fund', 'ALTER TABLE advance_settlements ADD COLUMN returned_to_fund REAL DEFAULT 0'),
        ('advance_settlements', 'overspend_amount', 'ALTER TABLE advance_settlements ADD COLUMN overspend_amount REAL DEFAULT 0'),
        ('advance_settlements', 'settlement_type', 'ALTER TABLE advance_settlements ADD COLUMN settlement_type TEXT DEFAULT "normal"'),
        ('advance_settlements', 'approved_by', 'ALTER TABLE advance_settlements ADD COLUMN approved_by INTEGER'),
        ('advance_settlements', 'approved_at', 'ALTER TABLE advance_settlements ADD COLUMN approved_at TIMESTAMP'),
        ('users', 'failed_login_count', 'ALTER TABLE users ADD COLUMN failed_login_count INTEGER DEFAULT 0'),
        ('users', 'locked_until', 'ALTER TABLE users ADD COLUMN locked_until TIMESTAMP'),
        ('users', 'password_changed_at', 'ALTER TABLE users ADD COLUMN password_changed_at TIMESTAMP'),
        ('users', 'must_change_password', 'ALTER TABLE users ADD COLUMN must_change_password INTEGER DEFAULT 0'),
        ('journal_entries', 'entry_number', 'ALTER TABLE journal_entries ADD COLUMN entry_number TEXT'),
        ('journal_entries', 'fiscal_year', 'ALTER TABLE journal_entries ADD COLUMN fiscal_year INTEGER'),
        ('journal_entries', 'fiscal_period', 'ALTER TABLE journal_entries ADD COLUMN fiscal_period TEXT'),
        ('journal_entries', 'entry_type', 'ALTER TABLE journal_entries ADD COLUMN entry_type TEXT DEFAULT "manual"'),
        ('journal_entries', 'is_reversed', 'ALTER TABLE journal_entries ADD COLUMN is_reversed INTEGER DEFAULT 0'),
        ('journal_entries', 'reversed_by', 'ALTER TABLE journal_entries ADD COLUMN reversed_by INTEGER'),
        ('journal_entries', 'reversal_of_entry_id', 'ALTER TABLE journal_entries ADD COLUMN reversal_of_entry_id INTEGER'),
        ('journal_entries', 'posted_by', 'ALTER TABLE journal_entries ADD COLUMN posted_by INTEGER'),
        ('journal_entries', 'posted_at', 'ALTER TABLE journal_entries ADD COLUMN posted_at TIMESTAMP'),
        ('journal_entries', 'project_id', 'ALTER TABLE journal_entries ADD COLUMN project_id INTEGER'),
        ('journal_entries', 'contract_id', 'ALTER TABLE journal_entries ADD COLUMN contract_id INTEGER'),
        ('journal_entries', 'reference_type', 'ALTER TABLE journal_entries ADD COLUMN reference_type TEXT'),
        ('journal_entries', 'reference_id', 'ALTER TABLE journal_entries ADD COLUMN reference_id INTEGER'),
        ('fiscal_calendar', 'is_locked', 'ALTER TABLE fiscal_calendar ADD COLUMN is_locked INTEGER DEFAULT 0'),
        ('fiscal_calendar', 'locked_at', 'ALTER TABLE fiscal_calendar ADD COLUMN locked_at TIMESTAMP'),
        ('fiscal_calendar', 'locked_by', 'ALTER TABLE fiscal_calendar ADD COLUMN locked_by INTEGER'),
        ('ar_ap_items', 'source_type', 'ALTER TABLE ar_ap_items ADD COLUMN source_type TEXT'),
        ('ar_ap_items', 'source_id', 'ALTER TABLE ar_ap_items ADD COLUMN source_id INTEGER'),
    ]

    for table_name, column_name, sql in migrations:
        if not has_column(table_name, column_name):
            cursor.execute(sql)
            cursor.execute('''
                INSERT OR IGNORE INTO schema_migrations (name, applied_at)
                VALUES (?, CURRENT_TIMESTAMP)
            ''', (f'{table_name}.{column_name}',))
        else:
            cursor.execute('''
                INSERT OR IGNORE INTO schema_migrations (name, applied_at)
                VALUES (?, CURRENT_TIMESTAMP)
            ''', (f'{table_name}.{column_name}',))

    _ensure_fiscal_lock_triggers(cursor)
    _ensure_performance_indexes(cursor)


def _ensure_performance_indexes(cursor):
    """Indexes for common filters used by dashboards and accounting screens."""
    indexes = [
        'CREATE INDEX IF NOT EXISTS idx_expenses_date ON expenses(expense_date)',
        'CREATE INDEX IF NOT EXISTS idx_expenses_project ON expenses(project_id)',
        'CREATE INDEX IF NOT EXISTS idx_expenses_status ON expenses(status)',
        'CREATE INDEX IF NOT EXISTS idx_expenses_project_date ON expenses(project_id, expense_date)',
        'CREATE INDEX IF NOT EXISTS idx_documents_date ON documents(doc_date)',
        'CREATE INDEX IF NOT EXISTS idx_documents_expense ON documents(expense_id)',
        'CREATE INDEX IF NOT EXISTS idx_documents_supplier ON documents(supplier_id)',
        'CREATE INDEX IF NOT EXISTS idx_journal_entries_date ON journal_entries(entry_date)',
        'CREATE INDEX IF NOT EXISTS idx_journal_entries_debit ON journal_entries(debit_account)',
        'CREATE INDEX IF NOT EXISTS idx_journal_entries_credit ON journal_entries(credit_account)',
        'CREATE INDEX IF NOT EXISTS idx_journal_entries_project ON journal_entries(project_id)',
        'CREATE INDEX IF NOT EXISTS idx_materials_code ON materials(code)',
        'CREATE INDEX IF NOT EXISTS idx_inventory_material_date ON inventory_transactions(material_id, transaction_date)',
        'CREATE INDEX IF NOT EXISTS idx_contracts_project ON project_contracts(project_id)',
        'CREATE INDEX IF NOT EXISTS idx_contracts_type_status ON project_contracts(contract_type, status)',
        'CREATE INDEX IF NOT EXISTS idx_billings_contract_date ON contract_billings(contract_id, billing_date)',
        'CREATE INDEX IF NOT EXISTS idx_revenues_project_date ON project_revenues(project_id, revenue_date)',
        'CREATE INDEX IF NOT EXISTS idx_timesheets_employee_date ON timesheets(employee_id, work_date)',
        'CREATE INDEX IF NOT EXISTS idx_timesheets_project_date ON timesheets(project_id, work_date)',
        'CREATE INDEX IF NOT EXISTS idx_ar_ap_status_due ON ar_ap_items(status, due_date)',
        'CREATE INDEX IF NOT EXISTS idx_ar_ap_partner ON ar_ap_items(partner_type, partner_name)',
        'CREATE UNIQUE INDEX IF NOT EXISTS idx_ar_ap_source ON ar_ap_items(source_type, source_id, partner_type) WHERE source_type IS NOT NULL AND source_id IS NOT NULL',
        'CREATE INDEX IF NOT EXISTS idx_fixed_assets_status ON fixed_assets(status)',
        'CREATE INDEX IF NOT EXISTS idx_asset_depr_period ON asset_depreciation_runs(period)',
        'CREATE INDEX IF NOT EXISTS idx_expiring_items_status_date ON expiring_items(status, expiry_date)',
    ]
    for sql in indexes:
        cursor.execute(sql)


def _ensure_fiscal_lock_triggers(cursor):
    cursor.execute('DROP TRIGGER IF EXISTS trg_journal_entries_period_lock_insert')
    cursor.execute('DROP TRIGGER IF EXISTS trg_journal_entries_period_lock_update')
    cursor.execute('DROP TRIGGER IF EXISTS trg_journal_entries_period_lock_delete')
    cursor.execute('DROP TRIGGER IF EXISTS trg_journal_entry_lines_period_lock_insert')
    cursor.execute('DROP TRIGGER IF EXISTS trg_journal_entry_lines_period_lock_update')
    cursor.execute('DROP TRIGGER IF EXISTS trg_journal_entry_lines_period_lock_delete')
    cursor.execute('''
        CREATE TRIGGER trg_journal_entries_period_lock_insert
        BEFORE INSERT ON journal_entries
        WHEN EXISTS (
            SELECT 1 FROM fiscal_calendar
            WHERE NEW.entry_date BETWEEN period_start AND period_end
              AND (COALESCE(is_locked, 0) = 1 OR COALESCE(is_closed, 0) = 1)
        )
        BEGIN
            SELECT RAISE(ABORT, 'Ky ke toan da khoa, khong duoc them but toan.');
        END
    ''')
    cursor.execute('''
        CREATE TRIGGER trg_journal_entries_period_lock_update
        BEFORE UPDATE ON journal_entries
        WHEN EXISTS (
            SELECT 1 FROM fiscal_calendar
            WHERE OLD.entry_date BETWEEN period_start AND period_end
              AND (COALESCE(is_locked, 0) = 1 OR COALESCE(is_closed, 0) = 1)
        ) OR EXISTS (
            SELECT 1 FROM fiscal_calendar
            WHERE NEW.entry_date BETWEEN period_start AND period_end
              AND (COALESCE(is_locked, 0) = 1 OR COALESCE(is_closed, 0) = 1)
        )
        BEGIN
            SELECT RAISE(ABORT, 'Ky ke toan da khoa, khong duoc sua but toan.');
        END
    ''')
    cursor.execute('''
        CREATE TRIGGER trg_journal_entries_period_lock_delete
        BEFORE DELETE ON journal_entries
        WHEN EXISTS (
            SELECT 1 FROM fiscal_calendar
            WHERE OLD.entry_date BETWEEN period_start AND period_end
              AND (COALESCE(is_locked, 0) = 1 OR COALESCE(is_closed, 0) = 1)
        )
        BEGIN
            SELECT RAISE(ABORT, 'Ky ke toan da khoa, khong duoc xoa but toan.');
        END
    ''')
    cursor.execute('''
        CREATE TRIGGER trg_journal_entry_lines_period_lock_insert
        BEFORE INSERT ON journal_entry_lines
        WHEN EXISTS (
            SELECT 1
            FROM journal_entries j
            JOIN fiscal_calendar f ON j.entry_date BETWEEN f.period_start AND f.period_end
            WHERE j.id = NEW.journal_entry_id
              AND (COALESCE(f.is_locked, 0) = 1 OR COALESCE(f.is_closed, 0) = 1)
        )
        BEGIN
            SELECT RAISE(ABORT, 'Ky ke toan da khoa, khong duoc them dong but toan.');
        END
    ''')
    cursor.execute('''
        CREATE TRIGGER trg_journal_entry_lines_period_lock_update
        BEFORE UPDATE ON journal_entry_lines
        WHEN EXISTS (
            SELECT 1
            FROM journal_entries j
            JOIN fiscal_calendar f ON j.entry_date BETWEEN f.period_start AND f.period_end
            WHERE j.id IN (OLD.journal_entry_id, NEW.journal_entry_id)
              AND (COALESCE(f.is_locked, 0) = 1 OR COALESCE(f.is_closed, 0) = 1)
        )
        BEGIN
            SELECT RAISE(ABORT, 'Ky ke toan da khoa, khong duoc sua dong but toan.');
        END
    ''')
    cursor.execute('''
        CREATE TRIGGER trg_journal_entry_lines_period_lock_delete
        BEFORE DELETE ON journal_entry_lines
        WHEN EXISTS (
            SELECT 1
            FROM journal_entries j
            JOIN fiscal_calendar f ON j.entry_date BETWEEN f.period_start AND f.period_end
            WHERE j.id = OLD.journal_entry_id
              AND (COALESCE(f.is_locked, 0) = 1 OR COALESCE(f.is_closed, 0) = 1)
        )
        BEGIN
            SELECT RAISE(ABORT, 'Ky ke toan da khoa, khong duoc xoa dong but toan.');
        END
    ''')


def _prepare_duplicate_invoice_controls(cursor):
    """Backfill suppliers and add duplicate invoice indexes where data allows."""
    cursor.execute('''
        INSERT OR IGNORE INTO suppliers (supplier_name)
        SELECT DISTINCT TRIM(supplier_name)
        FROM documents
        WHERE COALESCE(TRIM(supplier_name), '') <> ''
    ''')
    cursor.execute('''
        UPDATE documents
        SET supplier_id = (
            SELECT s.id FROM suppliers s WHERE s.supplier_name = TRIM(documents.supplier_name)
        )
        WHERE supplier_id IS NULL AND COALESCE(TRIM(supplier_name), '') <> ''
    ''')
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_documents_invoice_lookup
        ON documents(doc_number, supplier_id, amount)
    ''')
    try:
        cursor.execute('''
            CREATE UNIQUE INDEX IF NOT EXISTS uq_documents_invoice_supplier_amount
            ON documents(doc_number, supplier_id, amount)
            WHERE COALESCE(TRIM(doc_number), '') <> ''
              AND supplier_id IS NOT NULL
              AND amount IS NOT NULL
        ''')
    except sqlite3.IntegrityError:
        # Existing data already has duplicates. Runtime validation still blocks
        # new duplicates; the index can be created after historical cleanup.
        cursor.execute('''
            INSERT INTO audit_log (entity_type, entity_id, action, new_value)
            VALUES ('database', NULL, 'duplicate_invoice_index_skipped',
                    'Historical duplicate documents must be cleaned before UNIQUE invoice index can be created.')
        ''')


def _relax_document_number_unique(cursor):
    """Older databases made doc_number globally unique; invoice control is now scoped by supplier and amount."""
    cursor.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'documents'")
    row = cursor.fetchone()
    create_sql = row[0] if row else ''
    if 'doc_number TEXT UNIQUE' not in create_sql:
        return

    cursor.execute('ALTER TABLE documents RENAME TO documents_legacy_unique')
    cursor.execute('''
        CREATE TABLE documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            doc_type TEXT NOT NULL,
            doc_number TEXT,
            doc_date DATE,
            einvoice_provider TEXT,
            einvoice_status TEXT,
            einvoice_tax_authority_code TEXT,
            einvoice_transaction_id TEXT,
            einvoice_payload TEXT,
            qr_token TEXT,
            supplier_id INTEGER,
            supplier_name TEXT,
            description TEXT,
            amount REAL,
            expense_id INTEGER,
            project_id INTEGER,
            category_id INTEGER,
            status TEXT DEFAULT 'draft',
            file_path TEXT,
            created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(expense_id) REFERENCES expenses(id),
            FOREIGN KEY(project_id) REFERENCES projects(id),
            FOREIGN KEY(category_id) REFERENCES expense_categories(id),
            FOREIGN KEY(supplier_id) REFERENCES suppliers(id),
            FOREIGN KEY(created_by) REFERENCES users(id)
        )
    ''')
    cursor.execute('PRAGMA table_info(documents_legacy_unique)')
    old_columns = {row[1] for row in cursor.fetchall()}
    columns = [
        'id', 'doc_type', 'doc_number', 'doc_date',
        'einvoice_provider', 'einvoice_status', 'einvoice_tax_authority_code',
        'einvoice_transaction_id', 'einvoice_payload', 'qr_token',
        'supplier_id', 'supplier_name',
        'description', 'amount', 'expense_id', 'project_id', 'category_id', 'status',
        'file_path', 'created_by', 'created_at', 'updated_at'
    ]
    copy_columns = [column for column in columns if column in old_columns]
    column_sql = ', '.join(copy_columns)
    cursor.execute(f'''
        INSERT INTO documents ({column_sql})
        SELECT {column_sql}
        FROM documents_legacy_unique
    ''')
    cursor.execute('DROP TABLE documents_legacy_unique')


def _populate_fiscal_calendar(cursor):
    """Tạo lịch kỳ kế toán theo tháng để báo cáo/Power BI dùng chung một chuẩn."""
    current_year = date.today().year
    for year in range(current_year - 2, current_year + 3):
        for month in range(1, 13):
            start = date(year, month, 1)
            if month == 12:
                end = date(year + 1, 1, 1)
            else:
                end = date(year, month + 1, 1)
            period = f'{year}-{month:02d}'
            cursor.execute('''
                INSERT OR IGNORE INTO fiscal_calendar
                (fiscal_year, fiscal_period, period_start, period_end, is_closed)
                VALUES (?, ?, ?, ?, 0)
            ''', (year, period, start.isoformat(), (end - timedelta(days=1)).isoformat()))


def add_default_projects():
    """Thêm dự án mặc định để form chi phí luôn có lựa chọn ban đầu."""
    conn = get_connection()
    cursor = conn.cursor()

    projects = [
        ('CHUNG', 'Chi phí chung công ty', 'Văn phòng', None, None, 0),
        ('CT001', 'Công trình mẫu 01', 'Đang cập nhật', None, None, 0),
    ]

    for code, name, location, start_date, end_date, budget in projects:
        try:
            cursor.execute('''
                INSERT INTO projects (code, name, location, start_date, end_date, budget, created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (code, name, location, start_date, end_date, budget, 1))
        except sqlite3.IntegrityError:
            pass

    conn.commit()
    conn.close()


def add_default_project_accounting_data():
    """Seed hợp đồng, dự toán chi phí mẫu cho dự án CT001."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM projects WHERE code = 'CT001'")
    row = cursor.fetchone()
    if not row:
        conn.close()
        return
    project_id = row[0]
    cursor.execute('''
        UPDATE projects SET contract_value = 2000000000, owner_name = 'Chủ đầu tư mẫu',
               project_type = 'construction', budget = 1500000000
        WHERE id = ?
    ''', (project_id,))

    contracts = [
        ('customer', 'HD-TC-CT001', 'Chủ đầu tư mẫu', '2025-01-15', 2000000000, 10, 5, 200000000, 0),
        ('subcontract', 'HD-TP-CT001', 'Nhà thầu phụ mẫu', '2025-02-01', 500000000, 10, 5, 0, 50000000),
    ]
    for ctype, cno, partner, signed, value, vat, ret, adv_recv, adv_paid in contracts:
        cursor.execute('''
            SELECT id FROM project_contracts WHERE project_id = ? AND contract_no = ?
        ''', (project_id, cno))
        if cursor.fetchone():
            continue
        cursor.execute('''
            INSERT INTO project_contracts
            (project_id, contract_type, contract_no, partner_name, signed_date,
             contract_value, vat_rate, retention_rate, advance_received, advance_paid, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active')
        ''', (project_id, ctype, cno, partner, signed, value, vat, ret, adv_recv, adv_paid))

    cursor.execute('SELECT id, name FROM expense_categories ORDER BY id')
    categories = cursor.fetchall()
    plan_amounts = {
        'Vật liệu xây dựng': 400000000,
        'Vật tư phụ': 80000000,
        'Nhân công': 300000000,
        'Thầu phụ': 500000000,
        'Máy thi công': 120000000,
        'Quản lý công trình': 100000000,
    }
    for cat in categories:
        amount = plan_amounts.get(cat['name'], 0)
        if amount <= 0:
            continue
        try:
            cursor.execute('''
                INSERT INTO project_cost_plans (project_id, category_id, planned_amount, notes)
                VALUES (?, ?, ?, 'Dự toán mẫu')
            ''', (project_id, cat['id'], amount))
        except sqlite3.IntegrityError:
            pass

    cursor.execute('''
        SELECT id FROM project_contracts WHERE project_id = ? AND contract_type = 'customer'
    ''', (project_id,))
    cust = cursor.fetchone()
    if cust:
        cursor.execute('''
            SELECT id FROM contract_billings WHERE contract_id = ? AND milestone_name = 'Đợt 1'
        ''', (cust[0],))
        if not cursor.fetchone():
            amt = 400000000
            vat = amt * 0.1
            ret = amt * 0.05
            net = amt + vat - ret
            cursor.execute('''
                INSERT INTO contract_billings
                (contract_id, billing_date, milestone_name, quantity_or_percent,
                 amount_before_vat, vat_amount, retention_amount, net_amount, status)
                VALUES (?, '2025-03-15', 'Đợt 1', 20, ?, ?, ?, ?, 'approved')
            ''', (cust[0], amt, vat, ret, net))
            billing_id = cursor.lastrowid
            cursor.execute('''
                INSERT INTO project_revenues
                (project_id, contract_id, billing_id, revenue_date, amount, vat_amount, description)
                VALUES (?, ?, ?, '2025-03-15', ?, ?, 'Nghiệm thu đợt 1')
            ''', (project_id, cust[0], billing_id, amt, vat))

    conn.commit()
    conn.close()


def add_default_accounts():
    """Thêm hệ thống tài khoản kế toán theo Thông tư 99/2025/TT-BTC."""
    conn = get_connection()
    cursor = conn.cursor()

    accounts = [
        ('111', 'Tiền mặt', 'Tài sản', 1, None, 'Tiền mặt tại quỹ'),
        ('112', 'Tiền gửi ngân hàng', 'Tài sản', 1, None, 'Tiền gửi không kỳ hạn và có kỳ hạn tại ngân hàng'),
        ('113', 'Tiền đang chuyển', 'Tài sản', 1, None, 'Các khoản tiền đang chuyển'),
        ('121', 'Chứng khoán kinh doanh', 'Tài sản', 1, None, 'Chứng khoán nắm giữ cho mục đích kinh doanh'),
        ('128', 'Đầu tư nắm giữ đến ngày đáo hạn', 'Tài sản', 1, None, 'Tiền gửi, trái phiếu, khoản cho vay nắm giữ đến đáo hạn'),
        ('131', 'Phải thu của khách hàng', 'Tài sản', 1, None, 'Theo dõi công nợ phải thu khách hàng'),
        ('133', 'Thuế GTGT được khấu trừ', 'Tài sản', 1, None, 'Thuế giá trị gia tăng đầu vào được khấu trừ'),
        ('136', 'Phải thu nội bộ', 'Tài sản', 1, None, 'Các khoản phải thu nội bộ'),
        ('138', 'Phải thu khác', 'Tài sản', 1, None, 'Các khoản phải thu ngoài khách hàng và nội bộ'),
        ('141', 'Tạm ứng', 'Tài sản', 1, None, 'Theo dõi tiền tạm ứng cho nhân viên, đội thi công'),
        ('151', 'Hàng mua đang đi đường', 'Tài sản', 1, None, 'Vật tư, hàng hóa đã mua đang đi đường'),
        ('152', 'Nguyên liệu, vật liệu', 'Tài sản', 1, None, 'Vật liệu xây dựng, vật tư phụ, nhiên liệu'),
        ('153', 'Công cụ, dụng cụ', 'Tài sản', 1, None, 'Công cụ, dụng cụ phục vụ thi công và quản lý'),
        ('154', 'Chi phí sản xuất, kinh doanh dở dang', 'Tài sản', 1, None, 'Tập hợp chi phí công trình, hạng mục dở dang'),
        ('155', 'Thành phẩm', 'Tài sản', 1, None, 'Thành phẩm hoàn thành'),
        ('156', 'Hàng hóa', 'Tài sản', 1, None, 'Hàng hóa mua vào để bán'),
        ('211', 'Tài sản cố định hữu hình', 'Tài sản', 1, None, 'Máy móc, thiết bị, phương tiện vận tải, nhà xưởng'),
        ('213', 'Tài sản cố định vô hình', 'Tài sản', 1, None, 'Quyền sử dụng đất, phần mềm, tài sản vô hình khác'),
        ('214', 'Hao mòn tài sản cố định', 'Điều chỉnh tài sản', 1, None, 'Giá trị hao mòn lũy kế của tài sản cố định'),
        ('241', 'Xây dựng cơ bản dở dang', 'Tài sản', 1, None, 'Chi phí đầu tư xây dựng cơ bản dở dang'),
        ('242', 'Chi phí trả trước', 'Tài sản', 1, None, 'Chi phí phân bổ nhiều kỳ'),
        ('331', 'Phải trả cho người bán', 'Nợ phải trả', 1, None, 'Công nợ phải trả nhà cung cấp, thầu phụ'),
        ('333', 'Thuế và các khoản phải nộp Nhà nước', 'Nợ phải trả', 1, None, 'Thuế GTGT, TNDN, TNCN và khoản phải nộp khác'),
        ('334', 'Phải trả người lao động', 'Nợ phải trả', 1, None, 'Tiền lương, phụ cấp và khoản phải trả người lao động'),
        ('335', 'Chi phí phải trả', 'Nợ phải trả', 1, None, 'Chi phí đã phát sinh chưa đủ hồ sơ thanh toán'),
        ('338', 'Phải trả, phải nộp khác', 'Nợ phải trả', 1, None, 'Bảo hiểm, kinh phí công đoàn và khoản phải trả khác'),
        ('341', 'Vay và nợ thuê tài chính', 'Nợ phải trả', 1, None, 'Vay ngắn hạn, dài hạn, nợ thuê tài chính'),
        ('411', 'Vốn đầu tư của chủ sở hữu', 'Vốn chủ sở hữu', 1, None, 'Vốn góp của chủ sở hữu'),
        ('421', 'Lợi nhuận sau thuế chưa phân phối', 'Vốn chủ sở hữu', 1, None, 'Lợi nhuận sau thuế chưa phân phối'),
        ('511', 'Doanh thu bán hàng và cung cấp dịch vụ', 'Doanh thu', 1, None, 'Doanh thu bán hàng, cung cấp dịch vụ xây lắp'),
        ('515', 'Doanh thu hoạt động tài chính', 'Doanh thu', 1, None, 'Lãi tiền gửi, lãi chênh lệch tỷ giá, doanh thu tài chính khác'),
        ('521', 'Các khoản giảm trừ doanh thu', 'Điều chỉnh doanh thu', 1, None, 'Chiết khấu thương mại, giảm giá, hàng bán bị trả lại'),
        ('621', 'Chi phí nguyên liệu, vật liệu trực tiếp', 'Chi phí sản xuất', 1, None, 'Chi phí vật liệu trực tiếp cho công trình'),
        ('622', 'Chi phí nhân công trực tiếp', 'Chi phí sản xuất', 1, None, 'Chi phí nhân công trực tiếp thi công'),
        ('623', 'Chi phí sử dụng máy thi công', 'Chi phí sản xuất', 1, None, 'Nhiên liệu, khấu hao, thuê máy thi công'),
        ('627', 'Chi phí sản xuất chung', 'Chi phí sản xuất', 1, None, 'Chi phí chung tại đội/công trường'),
        ('632', 'Giá vốn hàng bán', 'Chi phí', 1, None, 'Giá vốn của sản phẩm, dịch vụ đã bán'),
        ('635', 'Chi phí tài chính', 'Chi phí', 1, None, 'Chi phí lãi vay, lỗ tỷ giá, chi phí tài chính khác'),
        ('641', 'Chi phí bán hàng', 'Chi phí', 1, None, 'Chi phí phục vụ bán hàng'),
        ('642', 'Chi phí quản lý doanh nghiệp', 'Chi phí', 1, None, 'Chi phí quản lý văn phòng, hành chính'),
        ('711', 'Thu nhập khác', 'Thu nhập khác', 1, None, 'Khoản thu nhập ngoài hoạt động thông thường'),
        ('811', 'Chi phí khác', 'Chi phí khác', 1, None, 'Khoản chi phí ngoài hoạt động thông thường'),
        ('821', 'Chi phí thuế thu nhập doanh nghiệp', 'Chi phí thuế', 1, None, 'Chi phí thuế TNDN hiện hành và hoãn lại'),
        ('911', 'Xác định kết quả kinh doanh', 'Xác định kết quả', 1, None, 'Kết chuyển doanh thu, chi phí để xác định kết quả'),
    ]

    legal_basis = 'Thông tư 99/2025/TT-BTC, Phụ lục II - Hệ thống tài khoản kế toán doanh nghiệp'

    for code, name, acc_type, level, parent_code, desc in accounts:
        try:
            cursor.execute(
                '''INSERT INTO accounts
                   (account_code, account_name, account_type, account_level, parent_code, legal_basis, description)
                   VALUES (?, ?, ?, ?, ?, ?, ?)''',
                (code, name, acc_type, level, parent_code, legal_basis, desc)
            )
        except sqlite3.IntegrityError:
            cursor.execute('''
                UPDATE accounts
                SET account_name = ?, account_type = ?, account_level = ?, parent_code = ?,
                    legal_basis = ?, description = ?, active = 1
                WHERE account_code = ?
            ''', (name, acc_type, level, parent_code, legal_basis, desc, code))

    conn.commit()
    conn.close()


def add_default_expense_categories():
    """Thêm danh mục chi phí mặc định cho công ty xây dựng."""
    conn = get_connection()
    cursor = conn.cursor()

    categories = [
        ('VLXD', 'Vật liệu xây dựng', 'Xi măng, thép, cát, đá, gạch, bê tông thương phẩm'),
        ('VTP', 'Vật tư phụ', 'Đinh, dây buộc, phụ kiện, vật tư tiêu hao'),
        ('NC', 'Nhân công', 'Lương công nhân, khoán nhân công, đội thi công'),
        ('THAU_PHU', 'Thầu phụ', 'Chi phí thuê thầu phụ, khoán hạng mục'),
        ('MAY_TC', 'Máy thi công', 'Thuê máy, nhiên liệu, sửa chữa, bảo dưỡng máy'),
        ('VAN_CHUYEN', 'Vận chuyển', 'Vận chuyển vật liệu, bốc xếp, cẩu hàng'),
        ('QL_CT', 'Quản lý công trình', 'Chi phí lán trại, điện nước, bảo vệ, an toàn lao động'),
        ('VAN_PHONG', 'Văn phòng', 'Chi phí quản lý doanh nghiệp, văn phòng phẩm, hành chính'),
        ('TAM_UNG', 'Tạm ứng', 'Tạm ứng cho nhân viên/đội thi công'),
        ('KHAC', 'Chi phí khác', 'Các chi phí khác cần bổ sung hồ sơ rõ ràng'),
    ]

    for code, name, description in categories:
        try:
            cursor.execute(
                'INSERT INTO expense_categories (code, name, parent_id, description) VALUES (?, ?, ?, ?)',
                (code, name, None, description)
            )
        except sqlite3.IntegrityError:
            cursor.execute(
                'UPDATE expense_categories SET name = ?, description = ? WHERE code = ?',
                (name, description, code)
            )

    conn.commit()
    conn.close()


def add_default_account_mappings():
    """Thêm mapping gợi ý tài khoản theo loại chi phí phổ biến."""
    conn = get_connection()
    cursor = conn.cursor()

    mappings = {
        'VLXD': ('621', '111', 'Vật liệu xây dựng: tập hợp chi phí vật liệu trực tiếp'),
        'VTP': ('621', '111', 'Vật tư phụ phục vụ thi công'),
        'NC': ('622', '334', 'Nhân công trực tiếp'),
        'THAU_PHU': ('627', '331', 'Chi phí thầu phụ/công nợ nhà cung cấp'),
        'MAY_TC': ('623', '111', 'Chi phí sử dụng máy thi công'),
        'VAN_CHUYEN': ('627', '111', 'Chi phí vận chuyển/bốc xếp'),
        'QL_CT': ('627', '111', 'Chi phí chung công trình'),
        'VAN_PHONG': ('642', '111', 'Chi phí quản lý doanh nghiệp'),
        'TAM_UNG': ('141', '111', 'Tạm ứng cho nhân viên/đội thi công'),
        'KHAC': ('642', '111', 'Chi phí khác'),
    }

    for category_code, (debit, credit, notes) in mappings.items():
        cursor.execute('SELECT id FROM expense_categories WHERE code = ?', (category_code,))
        row = cursor.fetchone()
        if not row:
            continue
        cursor.execute('''
            INSERT INTO category_account_mappings
            (category_id, debit_account, credit_account, notes, active)
            VALUES (?, ?, ?, ?, 1)
            ON CONFLICT(category_id) DO UPDATE SET
                debit_account = excluded.debit_account,
                credit_account = excluded.credit_account,
                notes = excluded.notes,
                active = 1
        ''', (row['id'], debit, credit, notes))

    conn.commit()
    conn.close()


def add_default_simple_catalogs():
    """Thêm danh mục đơn giản để người dùng chọn/sửa trong app."""
    conn = get_connection()
    cursor = conn.cursor()

    catalogs = [
        ('supplier', 'Nhà cung cấp mặc định', 'Có thể sửa thành nhà cung cấp thực tế'),
        ('employee', 'Kế toán', 'Người chi/người phụ trách'),
        ('employee', 'Chỉ huy trưởng', 'Người phụ trách công trình'),
        ('payment_method', 'Tiền mặt', ''),
        ('payment_method', 'Chuyển khoản', ''),
        ('payment_method', 'Sec', ''),
        ('payment_method', 'Khac', ''),
        ('customer', 'Khách hàng/Chủ đầu tư mặc định', 'Dùng cho hợp đồng, nghiệm thu, thanh toán'),
        ('department', 'Phòng VT-TB', 'Phòng vật tư thiết bị'),
        ('department', 'Phòng Kế toán', ''),
        ('department', 'Phòng Kế hoạch', ''),
        ('cost_subitem', 'Vật tư chính', 'Tiểu mục chi phí'),
        ('cost_subitem', 'Vận chuyển', 'Tiểu mục chi phí'),
        ('cost_subitem', 'Nhân công', 'Tiểu mục chi phí'),
        ('cost_group', 'Chi phí dự án', 'Nhóm chi phí dưới/ngang 10 triệu ưu tiên mẫu dự án'),
        ('cost_group', 'Chi phí văn phòng', 'Nhóm chi phí trên 10 triệu ưu tiên mẫu văn phòng'),
        ('opening_balance', 'Số dư đầu kỳ tiền mặt', 'Khai báo số dư ban đầu'),
        ('item_type', 'Vật tư xây dựng', 'Chủng loại vật tư hàng hóa'),
        ('item_type', 'Công cụ dụng cụ', 'Chủng loại vật tư hàng hóa'),
    ]

    for catalog_type, name, description in catalogs:
        cursor.execute('''
            INSERT INTO simple_catalogs (catalog_type, name, description, active)
            VALUES (?, ?, ?, 1)
            ON CONFLICT(catalog_type, name) DO UPDATE SET
                description = excluded.description,
                active = 1
        ''', (catalog_type, name, description))

    conn.commit()
    conn.close()


def add_default_company_settings():
    """Luu thong tin cong ty mac dinh de dung chung trong giao dien va bieu mau."""
    conn = get_connection()
    cursor = conn.cursor()
    settings = {
        'company_name': 'CÔNG TY CỔ PHẦN XÂY DỰNG VÀ ĐẦU TƯ TRUNG HẢI',
        'company_tax_code': '0312019045',
        'company_representative': '',
        'company_short_name': 'TRUNG HẢI',
    }
    for key, value in settings.items():
        cursor.execute('''
            INSERT INTO app_settings (key, value, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = CURRENT_TIMESTAMP
        ''', (key, value))
    conn.commit()
    conn.close()


def add_default_admin_templates():
    """Bo sung mau/hop dong/ho so hanh chinh mac dinh cho cong ty xay dung."""
    conn = get_connection()
    cursor = conn.cursor()

    templates = [
        ('HD-TC', 'Hợp đồng thi công xây dựng', 'Hợp đồng', 'Ký hợp đồng thi công với chủ đầu tư/khách hàng', 'Giám đốc; Kế toán trưởng; Phụ trách kế hoạch', 'Phòng Kế hoạch', 'documents/01_Hop_dong/Thi_cong', 'Kèm phụ lục tiến độ, dự toán, điều khoản thanh toán'),
        ('HD-TP', 'Hợp đồng thầu phụ', 'Hợp đồng', 'Thuê thầu phụ thực hiện hạng mục công trình', 'Giám đốc; Chỉ huy trưởng; Kế toán', 'Phòng Kế hoạch', 'documents/01_Hop_dong/Thau_phu', 'Cần gắn nghiệm thu, khối lượng, hóa đơn và thanh lý'),
        ('HD-VT', 'Hợp đồng mua vật tư', 'Hợp đồng', 'Mua vật tư/vật liệu xây dựng', 'Giám đốc; Kế toán; Thủ kho/Chỉ huy trưởng', 'Phòng Kế toán', 'documents/01_Hop_dong/Vat_tu', 'Cần kèm báo giá, đơn hàng, phiếu nhập, biên bản giao nhận'),
        ('BB-NT', 'Biên bản nghiệm thu khối lượng', 'Dự án công trường', 'Nghiệm thu khối lượng với chủ đầu tư/thầu phụ/tổ đội', 'Chỉ huy trưởng; Kỹ thuật; Đại diện các bên', 'Ban chỉ huy công trường', 'documents/03_Du_an_cong_truong/Nghiem_thu', 'Cần đối chiếu hợp đồng và bản vẽ/dự toán'),
        ('BB-BG', 'Biên bản bàn giao vật tư/thiết bị', 'Dự án công trường', 'Bàn giao vật tư, thiết bị, công cụ tại công trường', 'Người giao; Người nhận; Thủ kho; Chỉ huy trưởng', 'Kho/Công trường', 'documents/03_Du_an_cong_truong/Ban_giao', 'Dùng để đối chiếu tồn kho và trách nhiệm bảo quản'),
        ('KH-NS', 'Kế hoạch nhân sự công trường', 'Kế hoạch nhân sự', 'Lập kế hoạch điều động nhân sự theo dự án/ca thi công', 'Phòng Nhân sự; Chỉ huy trưởng; Giám đốc', 'Phòng Nhân sự', 'documents/04_Nhan_su/Ke_hoach_nhan_su', 'Cần cập nhật khi thay đổi đội thi công'),
        ('QD-DCNS', 'Quyết định điều chuyển nhân sự', 'Nhân sự', 'Điều chuyển nhân sự đến/rời công trường', 'Giám đốc; Phòng Nhân sự; Người lao động', 'Phòng Nhân sự', 'documents/04_Nhan_su/Dieu_chuyen', 'Liên kết với chấm công, tạm ứng, lương công trình'),
        ('CC-CT', 'Bảng chấm công công trường', 'Nhân sự', 'Theo dõi ngày công/ca công theo dự án', 'Người chấm công; Chỉ huy trưởng; Nhân sự', 'Công trường/Nhân sự', 'documents/04_Nhan_su/Cham_cong', 'Là căn cứ tính lương, khoán nhân công và hồ sơ chi phí'),
        ('DN-TU', 'Đề nghị tạm ứng', 'Kế toán', 'Tạm ứng tiền cho nhân viên/đội thi công/nhà cung cấp', 'Người đề nghị; Trưởng bộ phận; Kế toán; Giám đốc', 'Phòng Kế toán', 'documents/02_Ke_toan/Tam_ung', 'Cần theo dõi hoàn ứng và chứng từ gốc'),
        ('BK-HU', 'Bảng kê hoàn ứng', 'Kế toán', 'Quyết toán các khoản tạm ứng bằng chứng từ gốc', 'Người hoàn ứng; Kế toán; Trưởng bộ phận', 'Phòng Kế toán', 'documents/02_Ke_toan/Hoan_ung', 'Đối chiếu với chi phí và file hóa đơn đính kèm'),
        ('DN-TT', 'Đề nghị thanh toán', 'Kế toán', 'Thanh toán nhà cung cấp, thầu phụ, nhân công, chi phí khác', 'Người đề nghị; Bộ phận đề nghị; Kế toán; Giám đốc', 'Phòng Kế toán', 'documents/02_Ke_toan/Thanh_toan', 'Cần kiểm tra đủ hồ sơ trước khi thanh toán'),
        ('TL-HD', 'Biên bản thanh lý hợp đồng', 'Hợp đồng', 'Thanh lý hợp đồng sau khi hoàn thành/nghiệm thu/quyết toán', 'Đại diện các bên; Kế toán; Giám đốc', 'Phòng Kế hoạch/Kế toán', 'documents/01_Hop_dong/Thanh_ly', 'Khóa công nợ và hồ sơ hợp đồng'),
    ]

    fields = [
        ('company_name', 'Tên công ty', 'text', 1),
        ('company_tax_code', 'Mã số thuế', 'text', 0),
        ('company_representative', 'Người đại diện', 'text', 0),
        ('document_number', 'Số văn bản/chứng từ', 'text', 1),
        ('document_date', 'Ngày lập', 'date', 1),
        ('project_name', 'Tên dự án/công trình', 'text', 0),
        ('partner_name', 'Đối tác/Nhà cung cấp/Người nhận', 'text', 0),
        ('amount', 'Số tiền/Giá trị', 'number', 0),
        ('content', 'Nội dung', 'long_text', 1),
        ('attachments', 'Hồ sơ đính kèm', 'long_text', 0),
        ('prepared_by', 'Người lập', 'text', 0),
        ('approved_by', 'Người phê duyệt', 'text', 0),
    ]

    for form_code, form_name, scope, used_when, signatures, owner, folder, notes in templates:
        try:
            cursor.execute('''
                INSERT INTO form_templates
                (form_code, form_name, scope, source_workbook, sheet_name, file_path,
                 used_when, required_signatures, storage_owner, storage_method, usage_notes, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (form_code, form_name, scope, 'He thong mac dinh', form_code, folder, used_when,
                  signatures, owner, folder, notes))
        except sqlite3.IntegrityError:
            cursor.execute('''
                UPDATE form_templates
                SET form_name = ?, scope = ?, file_path = ?, used_when = ?,
                    required_signatures = ?, storage_owner = ?, storage_method = ?,
                    usage_notes = ?, active = 1
                WHERE form_code = ? AND scope = ? AND source_workbook = ? AND sheet_name = ?
            ''', (form_name, scope, folder, used_when, signatures, owner, folder, notes,
                  form_code, scope, 'He thong mac dinh', form_code))

        cursor.execute('''
            SELECT id FROM form_templates
            WHERE form_code = ? AND source_workbook = ? AND sheet_name = ?
            ORDER BY id LIMIT 1
        ''', (form_code, 'He thong mac dinh', form_code))
        template_row = cursor.fetchone()
        template_id = template_row['id'] if template_row else None

        for order, (field_key, field_label, field_type, required) in enumerate(fields, 1):
            try:
                cursor.execute('''
                    INSERT INTO form_template_fields
                    (form_template_id, form_code, field_key, field_label, field_type, required, display_order)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (template_id, form_code, field_key, field_label, field_type, required, order))
            except sqlite3.IntegrityError:
                cursor.execute('''
                    UPDATE form_template_fields
                    SET form_template_id = ?, field_label = ?, field_type = ?,
                        required = ?, display_order = ?, active = 1
                    WHERE form_code = ? AND field_key = ?
                ''', (template_id, field_label, field_type, required, order, form_code, field_key))

    requirements = [
        ('HD-TC', 'Hợp đồng thi công', 'Hợp đồng', 'Hợp đồng; Phụ lục hợp đồng; Dự toán; Tiến độ; Bảo lãnh nếu có; Biên bản nghiệm thu; Thanh lý', 'Thiếu hợp đồng/phụ lục sẽ làm yếu căn cứ thanh toán và quyết toán công trình'),
        ('HD-TP', 'Thầu phụ', 'Hợp đồng', 'Hợp đồng thầu phụ; Hồ sơ năng lực; Bảng khối lượng; Biên bản nghiệm thu; Hóa đơn; Thanh lý; Chứng từ thanh toán', 'Cần đủ nghiệm thu và hóa đơn trước khi thanh toán thầu phụ'),
        ('HD-VT', 'Mua vật tư', 'Vật tư', 'Báo giá; Đề nghị mua hàng; Hợp đồng/đơn hàng; Phiếu nhập; Biên bản giao nhận; Hóa đơn; Chứng từ thanh toán', 'Vật tư phải đối chiếu kho, công trình và hóa đơn'),
        ('KH-NS', 'Kế hoạch nhân sự công trường', 'Nhân sự', 'Kế hoạch nhân sự; Quyết định điều động; Bảng chấm công; Bảng lương/khoán; Đề nghị tạm ứng nếu có', 'Nhân sự công trường cần liên kết chấm công và chi phí nhân công'),
        ('DN-TT', 'Thanh toán', 'Kế toán', 'Đề nghị thanh toán; Hợp đồng/đơn hàng; Nghiệm thu/giao nhận; Hóa đơn; Bảng kê; Phê duyệt; Chứng từ chi tiền', 'Phải kiểm tra đủ hồ sơ trước khi thanh toán'),
    ]
    for ref_code, business_type, cost_group, required_docs, warning in requirements:
        try:
            cursor.execute('''
                INSERT INTO document_requirements
                (ref_code, business_type, cost_group, record_type, scope,
                 required_documents, warning_message, source_workbook, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (ref_code, business_type, cost_group, 'Hồ sơ hành chính/xây dựng',
                  'Toan cong ty', required_docs, warning, 'He thong mac dinh'))
        except sqlite3.IntegrityError:
            cursor.execute('''
                UPDATE document_requirements
                SET cost_group = ?, required_documents = ?, warning_message = ?, active = 1
                WHERE ref_code = ? AND business_type = ? AND source_workbook = ?
            ''', (cost_group, required_docs, warning, ref_code, business_type, 'He thong mac dinh'))

    conn.commit()
    conn.close()


def create_default_document_folders():
    """Tao cau truc thu muc luu tru ho so mac dinh."""
    folders = [
        'documents/00_Van_ban_quy_dinh',
        'documents/01_Hop_dong/Thi_cong',
        'documents/01_Hop_dong/Thau_phu',
        'documents/01_Hop_dong/Vat_tu',
        'documents/01_Hop_dong/Thanh_ly',
        'documents/02_Ke_toan/Tam_ung',
        'documents/02_Ke_toan/Hoan_ung',
        'documents/02_Ke_toan/Thanh_toan',
        'documents/02_Ke_toan/Hoa_don',
        'documents/02_Ke_toan/Phieu_thu_chi',
        'documents/03_Du_an_cong_truong/Nghiem_thu',
        'documents/03_Du_an_cong_truong/Ban_giao',
        'documents/03_Du_an_cong_truong/Nhat_ky',
        'documents/03_Du_an_cong_truong/Khoi_luong',
        'documents/04_Nhan_su/Ke_hoach_nhan_su',
        'documents/04_Nhan_su/Dieu_chuyen',
        'documents/04_Nhan_su/Cham_cong',
        'documents/04_Nhan_su/Luong_khoan',
        'documents/05_Ke_hoach/Ke_hoach_thi_cong',
        'documents/05_Ke_hoach/Tien_do',
        'documents/05_Ke_hoach/De_xuat',
        'documents/06_Phap_ly_du_an/Giay_phep',
        'documents/06_Phap_ly_du_an/Ban_ve',
        'documents/06_Phap_ly_du_an/Phe_duyet',
        'documents/99_Luu_tru_khac',
    ]
    for folder in folders:
        Path(folder).mkdir(parents=True, exist_ok=True)


def add_default_compliance_rules():
    """Thêm quy định hồ sơ mặc định để cảnh báo theo nghiệp vụ chi phí."""
    conn = get_connection()
    cursor = conn.cursor()

    rules = [
        ('RULE-VLXD', 'VLXD', 'Mua vật liệu xây dựng', 'Hồ sơ mua vật liệu',
         'Hóa đơn GTGT/hóa đơn bán hàng; Hợp đồng/đơn đặt hàng; Phiếu nhập kho hoặc biên bản giao nhận; Biên bản nghiệm thu khối lượng nếu giao thẳng công trình; Chứng từ thanh toán không dùng tiền mặt nếu hóa đơn từ 20 triệu đồng',
         'Chi phí vật liệu cần đủ hóa đơn, giao nhận/nhập kho và chứng từ thanh toán để tránh rủi ro loại chi phí.',
         'Luật Kế toán; Luật Thuế GTGT; Luật Thuế TNDN; Thông tư 99/2025/TT-BTC'),
        ('RULE-NC', 'NC', 'Chi trả nhân công', 'Hồ sơ nhân công',
         'Hợp đồng lao động/hợp đồng khoán; Bảng chấm công; Bảng lương hoặc bảng thanh toán khoán; Căn cước/Thông tin người nhận; Chứng từ chi tiền; Hồ sơ khấu trừ/khai thuế TNCN nếu phát sinh',
         'Chi phí nhân công thiếu bảng chấm công, bảng thanh toán hoặc hồ sơ thuế TNCN thường có rủi ro khi quyết toán.',
         'Bộ luật Lao động; Luật Thuế TNCN; Luật Kế toán; Thông tư 99/2025/TT-BTC'),
        ('RULE-THAU-PHU', 'THAU_PHU', 'Thuê thầu phụ', 'Hồ sơ thầu phụ',
         'Hợp đồng thầu phụ; Hồ sơ năng lực nếu cần; Biên bản nghiệm thu; Bảng xác nhận khối lượng; Hóa đơn; Biên bản thanh lý; Chứng từ thanh toán ngân hàng nếu từ 20 triệu đồng',
         'Nên gắn đủ nghiệm thu, hóa đơn và thanh toán theo từng lần thanh toán thầu phụ.',
         'Luật Xây dựng; Luật Kế toán; Luật Thuế TNDN; Thông tư 99/2025/TT-BTC'),
        ('RULE-MAY-TC', 'MAY_TC', 'Thuê/sử dụng máy thi công', 'Hồ sơ máy thi công',
         'Hợp đồng thuê máy hoặc lệnh điều động; Nhật trình máy; Biên bản xác nhận ca máy/giờ máy; Hóa đơn; Chứng từ nhiên liệu/sửa chữa liên quan; Chứng từ thanh toán',
         'Chi phí máy thi công nên có nhật trình hoặc xác nhận giờ máy để chứng minh phục vụ công trình.',
         'Luật Kế toán; quy định thuế TNDN; Thông tư 99/2025/TT-BTC'),
        ('RULE-VC', 'VAN_CHUYEN', 'Vận chuyển/bốc xếp', 'Hồ sơ vận chuyển',
         'Hợp đồng/đơn thuê vận chuyển; Lệnh điều xe hoặc phiếu giao hàng; Biên bản giao nhận; Hóa đơn vận chuyển; Chứng từ thanh toán',
         'Cần đối chiếu tuyến vận chuyển, vật tư và công trình nhận hàng.',
         'Luật Kế toán; quy định hóa đơn chứng từ; Thông tư 99/2025/TT-BTC'),
        ('RULE-TAM-UNG', 'TAM_UNG', 'Tạm ứng và hoàn ứng', 'Hồ sơ tạm ứng/hoàn ứng',
         'Đề nghị tạm ứng; Phiếu chi/ủy nhiệm chi; Bảng kê hoàn ứng; Hóa đơn/chứng từ gốc; Phê duyệt quyết toán tạm ứng',
         'Khoản tạm ứng cần được hoàn ứng bằng chứng từ gốc theo từng nghiệp vụ phát sinh.',
         'Luật Kế toán; quy chế tài chính nội bộ; Thông tư 99/2025/TT-BTC'),
        ('RULE-KHAC', 'KHAC', 'Chi phí khác', 'Hồ sơ chi phí khác',
         'Đề nghị thanh toán; Hóa đơn/chứng từ hợp pháp; Giải trình mục đích chi; Phê duyệt của người có thẩm quyền; Chứng từ thanh toán',
         'Chi phí khác cần mô tả rõ mục đích và người phê duyệt để dễ tra cứu sau này.',
         'Luật Kế toán; quy định thuế liên quan; Thông tư 99/2025/TT-BTC'),
    ]

    for code, category_code, transaction_type, rule_name, docs, warning, legal_basis in rules:
        try:
            cursor.execute('''
                INSERT INTO compliance_rules
                (rule_code, expense_category_code, transaction_type, rule_name,
                 required_documents, warning_message, legal_basis)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (code, category_code, transaction_type, rule_name, docs, warning, legal_basis))
        except sqlite3.IntegrityError:
            cursor.execute('''
                UPDATE compliance_rules
                SET expense_category_code = ?, transaction_type = ?, rule_name = ?,
                    required_documents = ?, warning_message = ?, legal_basis = ?, active = 1
                WHERE rule_code = ?
            ''', (category_code, transaction_type, rule_name, docs, warning, legal_basis, code))

    conn.commit()
    conn.close()


def import_default_knowledge_from_templates():
    """Đọc các workbook mẫu trong templates và đưa vào database tra cứu."""
    try:
        import openpyxl
    except ImportError:
        return

    template_dir = Path('templates')
    if not template_dir.exists():
        return

    workbook_paths = {
        'QT06_Ke_Toan_Van_Phong.xlsx': 'Văn phòng',
        'QT06_Ke_Toan_Du_An.xlsx': 'Dự án',
        'Bieu_mau_QT06_Ke_Toan_Van_Phong.xlsx': 'Văn phòng',
        'Bieu_mau_QT06_Ke_Toan_Du_An.xlsx': 'Dự án',
        'He_Thong_Ho_So_Chung_Tu_Thanh_Phat_Trung_Hai.xlsx': 'Hệ thống hồ sơ',
        'He_Thong_Quan_Ly_Chi_Phi_Trung_Hai.xlsx': 'Hệ thống chi phí',
    }

    conn = get_connection()
    cursor = conn.cursor()

    for filename, default_scope in workbook_paths.items():
        path = template_dir / filename
        if not path.exists():
            continue

        try:
            cursor.execute('DELETE FROM process_steps WHERE source_workbook = ?', (filename,))
            cursor.execute('DELETE FROM recurring_tasks WHERE source_workbook = ?', (filename,))
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            _import_workbook_forms(cursor, wb, filename, str(path), default_scope)
            _import_workbook_requirements(cursor, wb, filename)
            _import_workbook_processes(cursor, wb, filename)
            _import_workbook_limits(cursor, wb, filename)
            _import_workbook_tasks(cursor, wb, filename)
        except Exception as exc:
            print(f"Skip template import {filename}: {exc}")

    conn.commit()
    conn.close()


def _clean_cell(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return str(value).replace('\n', ' | ').strip()


def _row_values(row):
    return [_clean_cell(cell) for cell in row]


def _import_workbook_forms(cursor, wb, filename, file_path, default_scope):
    for ws in wb.worksheets:
        title = ws.title.strip()

        if title.startswith('BM') or title in ('Phiếu Chi', 'Phiếu thu'):
            form_code = title.replace(' ', '')
            form_name = _detect_sheet_title(ws) or title
            _upsert_form_template(cursor, form_code, form_name, default_scope, filename, title, file_path)
            continue

        if 'BIỂU MẪU' in title.upper() or 'Tổng hợp BM' in title or 'Tong hop' in title:
            rows = list(ws.iter_rows(values_only=True))
            header_index = _find_header_row(rows, ['MÃ BIỂU MẪU', 'Mã BM'])
            if header_index is None:
                continue
            headers = _row_values(rows[header_index])
            for raw_row in rows[header_index + 1:]:
                data = dict(zip(headers, _row_values(raw_row)))
                form_code = data.get('MÃ BIỂU MẪU') or data.get('Mã BM') or data.get('Mã biểu mẫu')
                form_name = data.get('TÊN BIỂU MẪU / MÔ TẢ') or data.get('Tên hồ sơ') or data.get('Tên biểu mẫu / Hồ sơ')
                if not form_code or not form_name:
                    continue
                _upsert_form_template(
                    cursor, form_code, form_name, data.get('ÁP DỤNG') or default_scope,
                    filename, title, file_path,
                    used_when=data.get('Dùng khi nào'),
                    required_signatures=data.get('★ CHỮ KÝ BẮT BUỘC'),
                    optional_signatures=data.get('◎ CHỮ KÝ TÙY CHỌN | (Có hoặc không)') or data.get('◎ CHỮ KÝ TÙY CHỌN\n(Có hoặc không)'),
                    storage_owner=data.get('Người lưu/ \nnơi lưu') or data.get('Người lưu'),
                    storage_method=data.get('Phương pháp lưu'),
                    usage_notes=data.get('GHI CHÚ SỬ DỤNG') or data.get('Ghi chú') or data.get('Thời gian lưu'),
                )


def _detect_sheet_title(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10), values_only=True):
        for value in row:
            text = _clean_cell(value)
            if text and any(key in text.upper() for key in ['ĐỀ NGHỊ', 'BẢNG KÊ', 'PHIẾU', 'GIẤY ĐỀ XUẤT']):
                return text
    return ''


def _upsert_form_template(cursor, form_code, form_name, scope, filename, sheet_name, file_path,
                          used_when='', required_signatures='', optional_signatures='',
                          storage_owner='', storage_method='', usage_notes=''):
    try:
        cursor.execute('''
            INSERT INTO form_templates
            (form_code, form_name, scope, source_workbook, sheet_name, file_path,
             used_when, required_signatures, optional_signatures, storage_owner,
             storage_method, usage_notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (form_code, form_name, scope, filename, sheet_name, file_path, used_when,
              required_signatures, optional_signatures, storage_owner, storage_method, usage_notes))
    except sqlite3.IntegrityError:
        cursor.execute('''
            UPDATE form_templates
            SET form_name = ?, scope = ?, file_path = ?, used_when = ?,
                required_signatures = ?, optional_signatures = ?, storage_owner = ?,
                storage_method = ?, usage_notes = ?, active = 1
            WHERE form_code = ? AND scope = ? AND source_workbook = ? AND sheet_name = ?
        ''', (form_name, scope, file_path, used_when, required_signatures, optional_signatures,
              storage_owner, storage_method, usage_notes, form_code, scope, filename, sheet_name))

    cursor.execute('''
        SELECT id FROM form_templates
        WHERE form_code = ? AND scope = ? AND source_workbook = ? AND sheet_name = ?
        ORDER BY id LIMIT 1
    ''', (form_code, scope, filename, sheet_name))
    row = cursor.fetchone()
    _ensure_default_template_fields(cursor, row['id'] if row else None, form_code)


def _ensure_default_template_fields(cursor, template_id, form_code):
    fields = [
        ('company_name', 'Tên công ty', 'text', 1),
        ('company_tax_code', 'Mã số thuế', 'text', 0),
        ('company_representative', 'Người đại diện', 'text', 0),
        ('department', 'Phòng ban/bộ phận', 'text', 0),
        ('document_number', 'Số văn bản/chứng từ', 'text', 1),
        ('document_date', 'Ngày lập', 'date', 1),
        ('project_name', 'Tên dự án/công trình', 'text', 0),
        ('requester', 'Người đề nghị/người lập', 'text', 0),
        ('partner_name', 'Đối tác/Nhà cung cấp/Người nhận', 'text', 0),
        ('amount', 'Số tiền/Giá trị', 'number', 0),
        ('content', 'Nội dung', 'long_text', 1),
        ('attachments', 'Hồ sơ đính kèm', 'long_text', 0),
        ('approved_by', 'Người phê duyệt', 'text', 0),
    ]

    for order, (field_key, field_label, field_type, required) in enumerate(fields, 1):
        try:
            cursor.execute('''
                INSERT INTO form_template_fields
                (form_template_id, form_code, field_key, field_label, field_type, required, display_order)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (template_id, form_code, field_key, field_label, field_type, required, order))
        except sqlite3.IntegrityError:
            cursor.execute('''
                UPDATE form_template_fields
                SET form_template_id = COALESCE(?, form_template_id), active = 1
                WHERE form_code = ? AND field_key = ?
            ''', (template_id, form_code, field_key))


def _import_workbook_requirements(cursor, wb, filename):
    sheet_specs = [
        ('📋 DANH MỤC HỒ SƠ', ['LOẠI NGHIỆP VỤ / CHI PHÍ']),
        ('📋 DANH MỤC CP', ['LOẠI CHI PHÍ']),
        ('🔍 TRA CỨU NHANH', ['TÌNH HUỐNG', 'Loại chi phí']),
    ]

    for sheet_name, markers in sheet_specs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_index = _find_header_row(rows, markers)
        if header_index is None:
            continue
        headers = _row_values(rows[header_index])

        for raw_row in rows[header_index + 1:]:
            values = _row_values(raw_row)
            if not any(values):
                continue
            data = dict(zip(headers, values))
            ref_code = data.get('STT') or data.get('#') or values[0]
            business_type = (
                data.get('LOẠI NGHIỆP VỤ / CHI PHÍ')
                or data.get('LOẠI CHI PHÍ')
                or data.get('Loại chi phí')
                or data.get('TÌNH HUỐNG')
            )
            if not business_type or business_type.startswith('═══') or business_type.startswith('──'):
                continue
            required_docs = (
                data.get('CHỨNG TỪ BẮT BUỘC (★)')
                or data.get('CHỨNG TỪ BẮT BUỘC')
                or data.get('Chứng từ bắt buộc')
                or data.get('HỒ SƠ CẦN CHUẨN BỊ (theo thứ tự)')
            )
            optional_docs = data.get('CHỨNG TỪ BỔ SUNG (◎) | (Có hoặc không)') or data.get('CHỨNG TỪ BỔ SUNG') or data.get('Chứng từ bổ sung')
            required_signatures = data.get('CHỮ KÝ BẮT BUỘC (★)') or data.get('CHỮ KÝ CẦN LẤY')
            forms = data.get('BIỂU MẪU | SỬ DỤNG') or data.get('BIỂU MẪU') or data.get('Biểu mẫu')
            warning = data.get('Cảnh báo đặc biệt') or data.get('CẢNH BÁO ĐẶC BIỆT') or ''

            try:
                cursor.execute('''
                    INSERT INTO document_requirements
                    (ref_code, business_type, cost_group, record_type, scope,
                     required_documents, optional_documents, required_signatures,
                     optional_signatures, approval_authority, deadline, forms,
                     limit_notes, warning_message, source_workbook)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    ref_code, business_type, data.get('NHÓM') or data.get('NHÓM CHI PHÍ'),
                    data.get('LOẠI | HỒ SƠ') or data.get('LOẠI HỒ SƠ'),
                    data.get('MỨC ÁP | DỤNG') or data.get('Áp dụng') or '',
                    required_docs, optional_docs, required_signatures,
                    data.get('CHỮ KÝ BỔ SUNG (◎) | (Có hoặc không)'),
                    data.get('THẨM QUYỀN DUYỆT') or data.get('Thẩm quyền duyệt'),
                    data.get('DEADLINE HỒ SƠ') or data.get('Deadline') or data.get('MỨC TIỀN / LƯU Ý'),
                    forms,
                    data.get('ĐỊNH MỨC / GHI CHÚ') or data.get('Định mức / Lưu ý') or data.get('MỨC TIỀN / LƯU Ý'),
                    warning,
                    filename,
                ))
            except sqlite3.IntegrityError:
                cursor.execute('''
                    UPDATE document_requirements
                    SET cost_group = ?, record_type = ?, scope = ?, required_documents = ?,
                        optional_documents = ?, required_signatures = ?, optional_signatures = ?,
                        approval_authority = ?, deadline = ?, forms = ?, limit_notes = ?,
                        warning_message = ?, active = 1
                    WHERE ref_code = ? AND business_type = ? AND source_workbook = ?
                ''', (
                    data.get('NHÓM') or data.get('NHÓM CHI PHÍ'),
                    data.get('LOẠI | HỒ SƠ') or data.get('LOẠI HỒ SƠ'),
                    data.get('MỨC ÁP | DỤNG') or data.get('Áp dụng') or '',
                    required_docs, optional_docs, required_signatures,
                    data.get('CHỮ KÝ BỔ SUNG (◎) | (Có hoặc không)'),
                    data.get('THẨM QUYỀN DUYỆT') or data.get('Thẩm quyền duyệt'),
                    data.get('DEADLINE HỒ SƠ') or data.get('Deadline') or data.get('MỨC TIỀN / LƯU Ý'),
                    forms,
                    data.get('ĐỊNH MỨC / GHI CHÚ') or data.get('Định mức / Lưu ý') or data.get('MỨC TIỀN / LƯU Ý'),
                    warning,
                    ref_code, business_type, filename,
                ))


def _import_workbook_processes(cursor, wb, filename):
    for ws in wb.worksheets:
        if 'QUY TRÌNH' not in ws.title.upper():
            continue
        process_name = ''
        for row in ws.iter_rows(values_only=True):
            values = _row_values(row)
            text = next((v for v in values if v), '')
            if not text:
                continue
            if 'QUY TRÌNH' in text.upper() or text[:2] in ('A.', 'B.', 'C.'):
                process_name = text
                continue
            if values[0].lower() in ('bước', 'step'):
                continue
            if values[0] and values[0][0].isdigit():
                cursor.execute('''
                    INSERT INTO process_steps
                    (process_name, step_no, responsible, action, duration, forms, notes, source_workbook)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    process_name or ws.title, values[0], values[1] if len(values) > 1 else '',
                    values[2] if len(values) > 2 else text,
                    values[3] if len(values) > 3 else '',
                    values[4] if len(values) > 4 else '',
                    values[5] if len(values) > 5 else '',
                    filename,
                ))
            elif text.startswith('→') or text.startswith('⚠') or text.startswith('BƯỚC'):
                cursor.execute('''
                    INSERT INTO process_steps
                    (process_name, step_no, responsible, action, duration, forms, notes, source_workbook)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (process_name or ws.title, '', '', text, '', '', '', filename))


def _import_workbook_limits(cursor, wb, filename):
    if '💰 ĐỊNH MỨC' not in wb.sheetnames:
        return
    ws = wb['💰 ĐỊNH MỨC']
    policy_group = ''
    for row in ws.iter_rows(values_only=True):
        values = _row_values(row)
        if not any(values):
            continue
        if values[0] and not values[1] and any(ch.isdigit() for ch in values[0]):
            policy_group = values[0]
            continue
        if values[0] in ('Khu vực', 'Nhóm', 'Mã'):
            continue
        if values[0] and values[1]:
            try:
                cursor.execute('''
                    INSERT INTO policy_limits
                    (policy_group, item_name, value_a, value_b, value_c, value_d, notes, source_workbook)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (policy_group, values[0], values[1], values[2], values[3], values[4], ' | '.join(values[5:8]), filename))
            except sqlite3.IntegrityError:
                cursor.execute('''
                    UPDATE policy_limits
                    SET value_a = ?, value_b = ?, value_c = ?, value_d = ?, notes = ?, active = 1
                    WHERE policy_group = ? AND item_name = ? AND source_workbook = ?
                ''', (values[1], values[2], values[3], values[4], ' | '.join(values[5:8]), policy_group, values[0], filename))


def _import_workbook_tasks(cursor, wb, filename):
    if '⏰ NHẮC NHỞ' not in wb.sheetnames:
        return
    ws = wb['⏰ NHẮC NHỞ']
    rows = list(ws.iter_rows(values_only=True))
    header_index = _find_header_row(rows, ['Ngày/Thời gian'])
    if header_index is None:
        return
    headers = _row_values(rows[header_index])
    for raw_row in rows[header_index + 1:]:
        data = dict(zip(headers, _row_values(raw_row)))
        if not data.get('Ngày/Thời gian') or not data.get('Nội dung'):
            continue
        try:
            cursor.execute('''
                INSERT INTO recurring_tasks
                (schedule_text, task_type, task_content, owner, approver, forms,
                 priority, status, notes, source_workbook)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data.get('Ngày/Thời gian'), data.get('Loại CV'), data.get('Nội dung'),
                data.get('Người thực hiện'), data.get('Người duyệt'), data.get('Biểu mẫu'),
                data.get('Mức độ ưu tiên'), data.get('Trạng thái'), data.get('Ghi chú'), filename,
            ))
        except sqlite3.IntegrityError:
            pass


def _find_header_row(rows, markers):
    for index, row in enumerate(rows):
        values = _row_values(row)
        joined = ' | '.join(values)
        if all(any(marker.lower() in value.lower() for value in values) or marker.lower() in joined.lower()
               for marker in markers):
            return index
    return None

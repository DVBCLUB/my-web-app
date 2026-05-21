"""
MODULE PROJECT ACCOUNTING - Kế toán dự án / công trình xây dựng
"""

from database import ConnectionPerRequestMixin
import config


class ProjectAccountingManager(ConnectionPerRequestMixin):
    """Quản lý hợp đồng, dự toán, doanh thu, WIP 154 và báo cáo dự án."""

    def __init__(self):
        pass

    # ── Dashboard ─────────────────────────────────────────
    def get_global_dashboard(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM projects WHERE status = 'active' AND code != 'CHUNG'")
        active_projects = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(planned_amount), 0) FROM project_cost_plans')
        total_planned = cursor.fetchone()[0]
        cursor.execute('''
            SELECT COALESCE(SUM(e.amount), 0) FROM expenses e
            JOIN projects p ON e.project_id = p.id WHERE p.code != 'CHUNG'
        ''')
        total_spent = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(amount), 0) FROM project_revenues')
        total_revenue = cursor.fetchone()[0]
        cursor.execute('''
            SELECT COUNT(*) FROM projects p
            WHERE p.code != 'CHUNG' AND COALESCE(p.budget, 0) > 0
            AND (SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE project_id = p.id)
                >= COALESCE(p.budget, 0) * 0.9
        ''')
        over_budget = cursor.fetchone()[0]
        cursor.execute('''
            SELECT COUNT(*) FROM project_contracts
            WHERE status = 'active' AND end_date IS NOT NULL
            AND date(end_date) <= date('now', '+30 days')
        ''')
        expiring_contracts = cursor.fetchone()[0]
        return {
            'active_projects': active_projects,
            'total_planned': total_planned,
            'total_spent': total_spent,
            'total_revenue': total_revenue,
            'profit': total_revenue - total_spent,
            'over_budget': over_budget,
            'expiring_contracts': expiring_contracts,
        }

    def get_project_dashboard(self, project_id):
        cursor = self.conn.cursor()
        cursor.execute('SELECT * FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project:
            return None
        cursor.execute('''
            SELECT COALESCE(SUM(planned_amount), 0) FROM project_cost_plans WHERE project_id = ?
        ''', (project_id,))
        planned = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(amount), 0) FROM expenses WHERE project_id = ?', (project_id,))
        spent = cursor.fetchone()[0]
        cursor.execute('SELECT COALESCE(SUM(amount), 0) FROM project_revenues WHERE project_id = ?', (project_id,))
        revenue = cursor.fetchone()[0]
        cursor.execute('''
            SELECT COALESCE(SUM(contract_value), 0) FROM project_contracts
            WHERE project_id = ? AND contract_type = 'customer'
        ''', (project_id,))
        contract_value = cursor.fetchone()[0]
        cursor.execute('''
            SELECT COALESCE(SUM(net_amount), 0) FROM contract_billings b
            JOIN project_contracts c ON b.contract_id = c.id
            WHERE c.project_id = ? AND c.contract_type = 'customer'
        ''', (project_id,))
        billed = cursor.fetchone()[0]
        wip = self.get_wip_summary(project_id)
        budget = project['budget'] or planned or 0
        return {
            'project': dict(project),
            'planned': planned,
            'spent': spent,
            'revenue': revenue,
            'profit': revenue - spent,
            'contract_value': contract_value,
            'billed': billed,
            'budget': budget,
            'remaining': budget - spent if budget else 0,
            'usage_percent': (spent / budget * 100) if budget else 0,
            'wip_total': wip['total_wip'],
        }

    def list_projects_active(self):
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT id, code, name FROM projects
            WHERE status = 'active' ORDER BY code
        ''')
        return cursor.fetchall()

    # ── Hợp đồng ──────────────────────────────────────────
    def get_contracts(self, project_id=None, contract_type=None, keyword=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT c.id, p.code, p.name, c.contract_type, c.contract_no, c.partner_name,
                   c.signed_date, c.contract_value, c.vat_rate, c.retention_rate,
                   c.advance_received, c.advance_paid, c.start_date, c.end_date,
                   c.status, COALESCE(c.notes, ''),
                   COALESCE((SELECT SUM(net_amount) FROM contract_billings WHERE contract_id = c.id), 0) AS billed
            FROM project_contracts c
            JOIN projects p ON c.project_id = p.id
            WHERE 1=1
        '''
        if project_id:
            query += ' AND c.project_id = ?'
            params.append(project_id)
        if contract_type:
            query += ' AND c.contract_type = ?'
            params.append(contract_type)
        if keyword:
            query += ' AND (c.contract_no LIKE ? OR c.partner_name LIKE ? OR p.name LIKE ?)'
            s = f'%{keyword}%'
            params.extend([s, s, s])
        query += ' ORDER BY c.signed_date DESC, c.id DESC'
        cursor.execute(query, params)
        return cursor.fetchall()

    def save_contract(self, data):
        cursor = self.conn.cursor()
        if data.get('id'):
            cursor.execute('''
                UPDATE project_contracts SET
                    project_id=?, contract_type=?, contract_no=?, partner_name=?,
                    signed_date=?, contract_value=?, vat_rate=?, retention_rate=?,
                    advance_received=?, advance_paid=?, start_date=?, end_date=?,
                    status=?, notes=?
                WHERE id=?
            ''', (
                data['project_id'], data['contract_type'], data['contract_no'],
                data['partner_name'], data.get('signed_date'), data.get('contract_value', 0),
                data.get('vat_rate', 10), data.get('retention_rate', 5),
                data.get('advance_received', 0), data.get('advance_paid', 0),
                data.get('start_date'), data.get('end_date'), data.get('status', 'active'),
                data.get('notes', ''), data['id'],
            ))
            self.conn.commit()
            return data['id']
        cursor.execute('''
            INSERT INTO project_contracts
            (project_id, contract_type, contract_no, partner_name, signed_date,
             contract_value, vat_rate, retention_rate, advance_received, advance_paid,
             start_date, end_date, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['project_id'], data['contract_type'], data['contract_no'],
            data['partner_name'], data.get('signed_date'), data.get('contract_value', 0),
            data.get('vat_rate', 10), data.get('retention_rate', 5),
            data.get('advance_received', 0), data.get('advance_paid', 0),
            data.get('start_date'), data.get('end_date'), data.get('status', 'active'),
            data.get('notes', ''),
        ))
        self.conn.commit()
        return cursor.lastrowid

    def get_contracts_by_project(self, project_id):
        return self.get_contracts(project_id=project_id)

    # ── Nghiệm thu / billing ──────────────────────────────
    def get_billings(self, contract_id=None, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT b.id, c.contract_no, c.partner_name, c.contract_type,
                   b.billing_date, b.milestone_name, b.quantity_or_percent,
                   b.amount_before_vat, b.vat_amount, b.retention_amount, b.net_amount,
                   b.status, COALESCE(b.notes, ''), p.code, p.name
            FROM contract_billings b
            JOIN project_contracts c ON b.contract_id = c.id
            JOIN projects p ON c.project_id = p.id
            WHERE 1=1
        '''
        if contract_id:
            query += ' AND b.contract_id = ?'
            params.append(contract_id)
        if project_id:
            query += ' AND c.project_id = ?'
            params.append(project_id)
        query += ' ORDER BY b.billing_date DESC'
        cursor.execute(query, params)
        return cursor.fetchall()

    def save_billing(self, data, create_revenue=False):
        cursor = self.conn.cursor()
        amt = float(data.get('amount_before_vat', 0) or 0)
        vat_rate = float(data.get('vat_rate', 10) or 10) / 100
        ret_rate = float(data.get('retention_rate', 5) or 5) / 100
        vat = amt * vat_rate
        ret = amt * ret_rate
        net = amt + vat - ret
        if data.get('id'):
            cursor.execute('''
                UPDATE contract_billings SET
                    contract_id=?, billing_date=?, milestone_name=?, quantity_or_percent=?,
                    amount_before_vat=?, vat_amount=?, retention_amount=?, net_amount=?,
                    status=?, notes=?
                WHERE id=?
            ''', (
                data['contract_id'], data['billing_date'], data.get('milestone_name', ''),
                data.get('quantity_or_percent'), amt, vat, ret, net,
                data.get('status', 'draft'), data.get('notes', ''), data['id'],
            ))
            billing_id = data['id']
        else:
            cursor.execute('''
                INSERT INTO contract_billings
                (contract_id, billing_date, milestone_name, quantity_or_percent,
                 amount_before_vat, vat_amount, retention_amount, net_amount, status, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['contract_id'], data['billing_date'], data.get('milestone_name', ''),
                data.get('quantity_or_percent'), amt, vat, ret, net,
                data.get('status', 'draft'), data.get('notes', ''),
            ))
            billing_id = cursor.lastrowid
        self.conn.commit()
        self._sync_billing_ar_ap(billing_id)
        if create_revenue and data.get('status') in ('approved', 'paid'):
            cursor.execute('SELECT project_id FROM project_contracts WHERE id = ?', (data['contract_id'],))
            row = cursor.fetchone()
            if row:
                self.save_revenue({
                    'project_id': row[0],
                    'contract_id': data['contract_id'],
                    'billing_id': billing_id,
                    'revenue_date': data['billing_date'],
                    'amount': amt,
                    'vat_amount': vat,
                    'description': data.get('milestone_name', 'Nghiệm thu'),
                })
        return billing_id

    # ── Dự toán chi phí ───────────────────────────────────
    def _sync_billing_ar_ap(self, billing_id):
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT b.id, b.contract_id, b.billing_date, b.net_amount, b.status,
                   b.milestone_name, c.contract_type, c.partner_name, c.project_id, c.contract_no
            FROM contract_billings b
            JOIN project_contracts c ON c.id = b.contract_id
            WHERE b.id = ?
        ''', (billing_id,))
        row = cursor.fetchone()
        if not row or float(row['net_amount'] or 0) <= 0:
            return
        partner_type = 'customer' if row['contract_type'] == 'customer' else 'supplier'
        paid = float(row['net_amount'] or 0) if row['status'] == 'paid' else 0
        status = 'closed' if row['status'] == 'paid' else 'open'
        cursor.execute('''
            INSERT INTO ar_ap_items
            (partner_type, partner_name, project_id, doc_id, due_date, amount, paid_amount,
             status, notes, source_type, source_id)
            VALUES (?, ?, ?, NULL, ?, ?, ?, ?, ?, 'billing', ?)
            ON CONFLICT(source_type, source_id, partner_type)
            WHERE source_type IS NOT NULL AND source_id IS NOT NULL
            DO UPDATE SET
                partner_name = excluded.partner_name,
                project_id = excluded.project_id,
                due_date = excluded.due_date,
                amount = excluded.amount,
                paid_amount = excluded.paid_amount,
                status = excluded.status,
                notes = excluded.notes
        ''', (
            partner_type, row['partner_name'], row['project_id'], row['billing_date'],
            float(row['net_amount'] or 0), paid, status,
            f"Tự sinh từ nghiệm thu {row['contract_no']} - {row['milestone_name'] or ''}".strip(),
            billing_id,
        ))
        self.conn.commit()

    def get_cost_plans(self, project_id):
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT cp.id, ec.name, cp.planned_amount,
                   COALESCE((SELECT SUM(amount) FROM expenses e
                             WHERE e.project_id = cp.project_id AND e.category_id = cp.category_id), 0) AS actual,
                   COALESCE(cp.notes, '')
            FROM project_cost_plans cp
            JOIN expense_categories ec ON cp.category_id = ec.id
            WHERE cp.project_id = ?
            ORDER BY ec.name
        ''', (project_id,))
        rows = cursor.fetchall()
        result = []
        for row in rows:
            planned = row['planned_amount'] or 0
            actual = row['actual'] or 0
            pct = (actual / planned * 100) if planned else 0
            result.append({
                'id': row['id'],
                'category_name': row['name'],
                'planned_amount': planned,
                'actual_amount': actual,
                'variance': planned - actual,
                'percent': pct,
                'notes': row['notes'] if 'notes' in row.keys() else '',
            })
        return result

    def save_cost_plan(self, project_id, category_id, planned_amount, notes=''):
        cursor = self.conn.cursor()
        cursor.execute('''
            INSERT INTO project_cost_plans (project_id, category_id, planned_amount, notes)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(project_id, category_id) DO UPDATE SET
                planned_amount = excluded.planned_amount,
                notes = excluded.notes
        ''', (project_id, category_id, planned_amount, notes))
        self.conn.commit()
        return cursor.lastrowid

    def get_cost_plan_vs_actual_report(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.code, p.name, ec.name AS category,
                   COALESCE(cp.planned_amount, 0) AS planned,
                   COALESCE(SUM(e.amount), 0) AS actual
            FROM projects p
            CROSS JOIN expense_categories ec
            LEFT JOIN project_cost_plans cp ON cp.project_id = p.id AND cp.category_id = ec.id
            LEFT JOIN expenses e ON e.project_id = p.id AND e.category_id = ec.id
            WHERE p.code != 'CHUNG'
        '''
        if project_id:
            query += ' AND p.id = ?'
            params.append(project_id)
        query += '''
            GROUP BY p.id, ec.id
            HAVING planned > 0 OR actual > 0
            ORDER BY p.code, ec.name
        '''
        cursor.execute(query, params)
        return cursor.fetchall()

    # ── Doanh thu ──────────────────────────────────────────
    def get_revenues(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT r.id, p.code, p.name, COALESCE(c.contract_no, ''),
                   r.revenue_date, r.amount, r.vat_amount, COALESCE(r.description, '')
            FROM project_revenues r
            JOIN projects p ON r.project_id = p.id
            LEFT JOIN project_contracts c ON r.contract_id = c.id
            WHERE 1=1
        '''
        if project_id:
            query += ' AND r.project_id = ?'
            params.append(project_id)
        query += ' ORDER BY r.revenue_date DESC'
        cursor.execute(query, params)
        return cursor.fetchall()

    def save_revenue(self, data):
        cursor = self.conn.cursor()
        if data.get('id'):
            cursor.execute('''
                UPDATE project_revenues SET
                    project_id=?, contract_id=?, billing_id=?, revenue_date=?,
                    amount=?, vat_amount=?, description=?
                WHERE id=?
            ''', (
                data['project_id'], data.get('contract_id'), data.get('billing_id'),
                data['revenue_date'], data.get('amount', 0), data.get('vat_amount', 0),
                data.get('description', ''), data['id'],
            ))
            self.conn.commit()
            return data['id']
        cursor.execute('''
            INSERT INTO project_revenues
            (project_id, contract_id, billing_id, revenue_date, amount, vat_amount, description)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['project_id'], data.get('contract_id'), data.get('billing_id'),
            data['revenue_date'], data.get('amount', 0), data.get('vat_amount', 0),
            data.get('description', ''),
        ))
        self.conn.commit()
        return cursor.lastrowid

    # ── WIP / TK 154 ──────────────────────────────────────
    def get_wip_summary(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT COALESCE(p.code, ''), COALESCE(p.name, ''),
                   COALESCE(SUM(j.amount), 0)
            FROM journal_entries j
            LEFT JOIN projects p ON j.project_id = p.id
            WHERE j.debit_account = '154'
        '''
        if project_id:
            query += ' AND j.project_id = ?'
            params.append(project_id)
        query += ' GROUP BY j.project_id ORDER BY p.code'
        cursor.execute(query, params)
        rows = cursor.fetchall()
        total = sum(r[2] for r in rows) if rows else 0
        if project_id and not rows:
            cursor.execute('''
                SELECT COALESCE(SUM(j.amount), 0) FROM journal_entries j
                WHERE j.project_id = ? AND j.debit_account IN ('621','622','623','627')
            ''', (project_id,))
            alt = cursor.fetchone()[0] or 0
            return {'items': [], 'total_wip': alt}
        return {'items': rows, 'total_wip': total}

    def post_expense_to_wip(self, expense_id, created_by=1):
        """Tạo bút toán Nợ 154 / Có 621-627 từ chi phí đã hạch toán."""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT e.*, ec.name AS category_name
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE e.id = ?
        ''', (expense_id,))
        expense = cursor.fetchone()
        if not expense:
            raise ValueError('Không tìm thấy chi phí')
        if not expense['project_id']:
            raise ValueError('Chi phí chưa gắn dự án')
        cursor.execute('''
            SELECT debit_account FROM journal_entries
            WHERE expense_id = ? ORDER BY id DESC LIMIT 1
        ''', (expense_id,))
        je = cursor.fetchone()
        cost_account = je['debit_account'] if je else '621'
        if cost_account == '154':
            raise ValueError('Chi phí đã tập hợp vào 154')
        cursor.execute('''
            INSERT INTO journal_entries
            (entry_date, description, debit_account, credit_account, amount,
             expense_id, project_id, contract_id, reference_type, reference_id, created_by)
            VALUES (?, ?, '154', ?, ?, ?, ?, ?, 'expense_wip', ?, ?)
        ''', (
            expense['expense_date'],
            f"Tập hợp CPSXDD: {expense['description']}",
            cost_account, expense['amount'], expense_id,
            expense['project_id'], expense['contract_id'],
            expense_id, created_by,
        ))
        self.conn.commit()
        return cursor.lastrowid

    # ── Bút toán dự án ────────────────────────────────────
    def list_journal_by_project(self, project_id=None, keyword=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT j.id, j.entry_date, p.code, p.name, j.description,
                   j.debit_account, j.credit_account, j.amount, j.reference_type
            FROM journal_entries j
            LEFT JOIN projects p ON j.project_id = p.id
            WHERE j.project_id IS NOT NULL
        '''
        if project_id:
            query += ' AND j.project_id = ?'
            params.append(project_id)
        if keyword:
            query += ' AND (j.description LIKE ? OR j.debit_account LIKE ?)'
            s = f'%{keyword}%'
            params.extend([s, s])
        query += ' ORDER BY j.entry_date DESC, j.id DESC'
        cursor.execute(query, params)
        return cursor.fetchall()

    def create_revenue_journal(self, revenue_id, created_by=1):
        cursor = self.conn.cursor()
        cursor.execute('SELECT * FROM project_revenues WHERE id = ?', (revenue_id,))
        rev = cursor.fetchone()
        if not rev:
            raise ValueError('Không tìm thấy doanh thu')
        total = (rev['amount'] or 0) + (rev['vat_amount'] or 0)
        cursor.execute('''
            INSERT INTO journal_entries
            (entry_date, description, debit_account, credit_account, amount,
             project_id, contract_id, reference_type, reference_id, created_by)
            VALUES (?, ?, '131', '511', ?, ?, ?, 'revenue', ?, ?)
        ''', (
            rev['revenue_date'], rev['description'] or 'Ghi nhận doanh thu',
            rev['amount'], rev['project_id'], rev['contract_id'], revenue_id, created_by,
        ))
        if rev['vat_amount']:
            cursor.execute('''
                INSERT INTO journal_entries
                (entry_date, description, debit_account, credit_account, amount,
                 project_id, contract_id, reference_type, reference_id, created_by)
                VALUES (?, ?, '131', '333', ?, ?, ?, 'revenue_vat', ?, ?)
            ''', (
                rev['revenue_date'], 'VAT doanh thu', rev['vat_amount'],
                rev['project_id'], rev['contract_id'], revenue_id, created_by,
            ))
        self.conn.commit()
        return cursor.lastrowid

    # ── Báo cáo ───────────────────────────────────────────
    def get_cost_collection_report(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.code, p.name, ec.name,
                   COALESCE(SUM(e.amount), 0) AS total
            FROM expenses e
            JOIN projects p ON e.project_id = p.id
            JOIN expense_categories ec ON e.category_id = ec.id
            WHERE p.code != 'CHUNG'
        '''
        if project_id:
            query += ' AND p.id = ?'
            params.append(project_id)
        query += ' GROUP BY p.id, ec.id ORDER BY p.code, total DESC'
        cursor.execute(query, params)
        return cursor.fetchall()

    def get_project_pl_report(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.id, p.code, p.name,
                   COALESCE((SELECT SUM(amount) FROM project_revenues WHERE project_id = p.id), 0),
                   COALESCE((SELECT SUM(amount) FROM expenses WHERE project_id = p.id), 0)
            FROM projects p WHERE p.code != 'CHUNG'
        '''
        if project_id:
            query += ' AND p.id = ?'
            params.append(project_id)
        cursor.execute(query, params)
        rows = cursor.fetchall()
        return [{
            'id': r[0], 'code': r[1], 'name': r[2],
            'revenue': r[3], 'cost': r[4], 'profit': r[3] - r[4],
        } for r in rows]

    def get_contract_progress_report(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.code, p.name, c.contract_no, c.partner_name,
                   c.contract_type, c.contract_value,
                   COALESCE(SUM(b.net_amount), 0) AS billed,
                   c.contract_value - COALESCE(SUM(b.net_amount), 0) AS remaining
            FROM project_contracts c
            JOIN projects p ON c.project_id = p.id
            LEFT JOIN contract_billings b ON b.contract_id = c.id
            WHERE 1=1
        '''
        if project_id:
            query += ' AND p.id = ?'
            params.append(project_id)
        query += ' GROUP BY c.id ORDER BY p.code, c.contract_no'
        cursor.execute(query, params)
        return cursor.fetchall()

    def get_subcontract_control_report(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.code, p.name, c.id, c.contract_no, c.partner_name,
                   c.contract_value,
                   COALESCE(SUM(b.net_amount), 0) AS performed_value,
                   c.contract_value - COALESCE(SUM(b.net_amount), 0) AS remaining,
                   COALESCE((SELECT SUM(amount) FROM guarantee_bonds g
                             WHERE g.contract_id = c.id AND g.status = 'active'), 0) AS active_bonds,
                   COALESCE((SELECT COUNT(*) FROM warranty_periods w
                             WHERE w.contract_id = c.id AND w.status = 'active'), 0) AS active_warranties
            FROM project_contracts c
            JOIN projects p ON p.id = c.project_id
            LEFT JOIN contract_billings b ON b.contract_id = c.id
            WHERE c.contract_type = 'subcontract'
        '''
        if project_id:
            query += ' AND c.project_id = ?'
            params.append(project_id)
        query += ' GROUP BY c.id ORDER BY p.code, c.contract_no'
        cursor.execute(query, params)
        return cursor.fetchall()

    def save_guarantee_bond(self, data):
        cursor = self.conn.cursor()
        if data.get('id'):
            cursor.execute('''
                UPDATE guarantee_bonds SET
                    contract_id=?, milestone_id=?, bond_type=?, bond_number=?,
                    issuer=?, amount=?, issue_date=?, expiry_date=?, status=?, notes=?
                WHERE id=?
            ''', (
                data['contract_id'], data.get('milestone_id'), data['bond_type'],
                data.get('bond_number', ''), data.get('issuer', ''), data.get('amount', 0),
                data.get('issue_date'), data.get('expiry_date'), data.get('status', 'active'),
                data.get('notes', ''), data['id'],
            ))
            self.conn.commit()
            return data['id']
        cursor.execute('''
            INSERT INTO guarantee_bonds
            (contract_id, milestone_id, bond_type, bond_number, issuer, amount,
             issue_date, expiry_date, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['contract_id'], data.get('milestone_id'), data['bond_type'],
            data.get('bond_number', ''), data.get('issuer', ''), data.get('amount', 0),
            data.get('issue_date'), data.get('expiry_date'), data.get('status', 'active'),
            data.get('notes', ''),
        ))
        self.conn.commit()
        return cursor.lastrowid

    def save_warranty_period(self, data):
        cursor = self.conn.cursor()
        if data.get('id'):
            cursor.execute('''
                UPDATE warranty_periods SET
                    contract_id=?, milestone_id=?, warranty_scope=?, start_date=?,
                    end_date=?, retention_amount=?, status=?, notes=?
                WHERE id=?
            ''', (
                data['contract_id'], data.get('milestone_id'), data.get('warranty_scope', ''),
                data.get('start_date'), data.get('end_date'), data.get('retention_amount', 0),
                data.get('status', 'active'), data.get('notes', ''), data['id'],
            ))
            self.conn.commit()
            return data['id']
        cursor.execute('''
            INSERT INTO warranty_periods
            (contract_id, milestone_id, warranty_scope, start_date, end_date,
             retention_amount, status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['contract_id'], data.get('milestone_id'), data.get('warranty_scope', ''),
            data.get('start_date'), data.get('end_date'), data.get('retention_amount', 0),
            data.get('status', 'active'), data.get('notes', ''),
        ))
        self.conn.commit()
        return cursor.lastrowid

    def get_expense_categories_for_plan(self):
        cursor = self.conn.cursor()
        cursor.execute('SELECT id, name FROM expense_categories ORDER BY name')
        return cursor.fetchall()

    def export_qt06_data(self, project_id):
        """Chuẩn bị dữ liệu xuất Excel QT06 kế toán dự án."""
        dash = self.get_project_dashboard(project_id)
        cost_plans = self.get_cost_plans(project_id)
        revenues = self.get_revenues(project_id)
        contracts = self.get_contracts(project_id=project_id)
        expenses = []
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT e.expense_date, ec.name, e.description, e.amount, e.status
            FROM expenses e
            JOIN expense_categories ec ON e.category_id = ec.id
            WHERE e.project_id = ?
            ORDER BY e.expense_date
        ''', (project_id,))
        for row in cursor.fetchall():
            expenses.append(dict(row))
        return {
            'project': dash['project'] if dash else {},
            'dashboard': dash or {},
            'cost_plans': cost_plans,
            'revenues': [dict(r) for r in revenues],
            'contracts': [dict(r) for r in contracts],
            'expenses': expenses,
        }


def export_project_qt06_excel(manager, project_id):
    """Xuất báo cáo QT06 kế toán dự án ra Excel."""
    from pathlib import Path
    from datetime import datetime
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError('Cần cài openpyxl: pip install openpyxl')

    data = manager.export_qt06_data(project_id)
    project = data.get('project', {})
    dash = data.get('dashboard', {})
    reports_dir = Path(__file__).resolve().parent.parent / 'reports'
    reports_dir.mkdir(exist_ok=True)
    code = project.get('code', 'DA')
    filename = reports_dir / f"QT06_Ke_Toan_Du_An_{code}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    template_path = Path(__file__).resolve().parent.parent / 'templates' / 'QT06_Ke_Toan_Du_An.xlsx'
    if template_path.exists():
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            def _safe_write(cell_ref, value):
                cell = ws[cell_ref]
                if getattr(cell, '__class__', None).__name__ == 'MergedCell':
                    return False
                cell.value = value
                return True

            _safe_write('B3', project.get('name', ''))
            _safe_write('B4', project.get('code', ''))
            _safe_write('B5', project.get('owner_name', ''))
            start_row = 10
            for idx, plan in enumerate(data.get('cost_plans', [])):
                r = start_row + idx
                _safe_write(f'A{r}', plan.get('category_name', ''))
                _safe_write(f'B{r}', plan.get('planned_amount', 0))
                _safe_write(f'C{r}', plan.get('actual_amount', 0))
                _safe_write(f'D{r}', plan.get('variance', 0))
            wb.save(filename)
            return str(filename)
        except Exception:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QT06 Du an'
    ws['A1'] = 'BÁO CÁO KẾ TOÁN DỰ ÁN - QT06'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A3'] = 'Dự án:'
    ws['B3'] = project.get('name', '')
    ws['A4'] = 'Mã:'
    ws['B4'] = project.get('code', '')
    ws['A5'] = 'Chủ đầu tư:'
    ws['B5'] = project.get('owner_name', '')
    ws['A7'] = 'Dự toán:'
    ws['B7'] = dash.get('planned', 0)
    ws['A8'] = 'Đã chi:'
    ws['B8'] = dash.get('spent', 0)
    ws['A9'] = 'Doanh thu:'
    ws['B9'] = dash.get('revenue', 0)
    ws['A10'] = 'Lãi/lỗ:'
    ws['B10'] = dash.get('profit', 0)
    headers = ['Loại chi phí', 'Dự toán', 'Thực tế', 'Chênh lệch', '% TH']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=12, column=col, value=h)
        cell.font = Font(bold=True)
    row = 13
    for plan in data.get('cost_plans', []):
        ws.cell(row=row, column=1, value=plan.get('category_name', ''))
        ws.cell(row=row, column=2, value=plan.get('planned_amount', 0))
        ws.cell(row=row, column=3, value=plan.get('actual_amount', 0))
        ws.cell(row=row, column=4, value=plan.get('variance', 0))
        ws.cell(row=row, column=5, value=round(plan.get('percent', 0), 1))
        row += 1
    row += 2
    ws.cell(row=row, column=1, value='CHI PHÍ PHÁT SINH').font = Font(bold=True)
    row += 1
    for col, h in enumerate(['Ngày', 'Loại', 'Mô tả', 'Số tiền'], 1):
        ws.cell(row=row, column=col, value=h).font = Font(bold=True)
    row += 1
    for exp in data.get('expenses', []):
        ws.cell(row=row, column=1, value=exp.get('expense_date', ''))
        ws.cell(row=row, column=2, value=exp.get('name', ''))
        ws.cell(row=row, column=3, value=exp.get('description', ''))
        ws.cell(row=row, column=4, value=exp.get('amount', 0))
        row += 1
    wb.save(filename)
    return str(filename)

"""
Bank reconciliation helpers.

This module keeps statement import, matching and exception reporting out of
the UI so the reconciliation workflow can be reused by reports or jobs later.
"""

import csv
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

from database import ConnectionPerRequestMixin


class BankReconciliationManager(ConnectionPerRequestMixin):
    """Import bank statement rows and match them to system transactions."""

    DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y")

    def __init__(self):
        pass

    def import_csv(self, file_path, bank_account="", source_name=None):
        """Import a CSV bank statement.

        Expected columns are flexible. The importer accepts date/amount/
        description variants commonly exported by Vietnamese banks.
        """
        source_name = source_name or file_path
        imported = 0
        skipped = 0
        with open(file_path, newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                normalized = {self._clean_key(k): (v or "").strip() for k, v in row.items()}
                txn_date = self._pick(normalized, "date", "transaction_date", "ngay", "ngay_gd")
                description = self._pick(normalized, "description", "memo", "dien_giai", "noi_dung")
                amount = self._amount_from_row(normalized)
                ref_no = self._pick(normalized, "reference", "ref", "so_gd", "ma_gd")
                if not txn_date or amount == 0:
                    skipped += 1
                    continue
                txn_date = self._parse_date(txn_date)
                if not txn_date:
                    skipped += 1
                    continue
                self.add_statement_row(txn_date, amount, description, ref_no, bank_account, source_name)
                imported += 1
        return {"imported": imported, "skipped": skipped}

    def import_bank_statement(self, bank_account_id, file_path):
        """Import CSV/XLSX statement into bank_transactions."""
        path = Path(file_path)
        rows = self._read_statement_rows(path)
        imported = 0
        skipped = 0
        for row in rows:
            normalized = {self._clean_key(k): (v or "") for k, v in row.items()}
            txn_date = self._pick(normalized, "date", "transaction_date", "ngay", "ngay_gd")
            description = self._pick(normalized, "description", "memo", "dien_giai", "noi_dung")
            ref_no = self._pick(normalized, "reference", "ref", "so_gd", "ma_gd")
            amount = self._amount_from_row(normalized)
            if not txn_date or amount == 0:
                skipped += 1
                continue
            txn_date = self._parse_date(txn_date)
            if not txn_date:
                skipped += 1
                continue
            self.add_bank_transaction(bank_account_id, txn_date, amount, description, ref_no)
            imported += 1
        return {"imported": imported, "skipped": skipped}

    def add_statement_row(self, transaction_date, amount, description="", reference_no="",
                          bank_account="", source_name="manual"):
        cursor = self.conn.cursor()
        cursor.execute("""
            INSERT OR IGNORE INTO bank_statement_rows
            (bank_account, transaction_date, description, amount, reference_no, source_name)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (bank_account, transaction_date, description, float(amount or 0), reference_no, source_name))
        self.conn.commit()
        return cursor.lastrowid

    def add_bank_transaction(self, bank_account_id, transaction_date, amount,
                             description="", reference_no=""):
        cursor = self.conn.cursor()
        amount = float(amount or 0)
        debit = abs(amount) if amount < 0 else 0
        credit = amount if amount > 0 else 0
        cursor.execute("""
            INSERT OR IGNORE INTO bank_transactions
            (bank_account_id, transaction_date, description, debit_amount,
             credit_amount, amount, reference_no)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (bank_account_id, transaction_date, description, debit, credit, amount, reference_no))
        self.conn.commit()
        return cursor.lastrowid

    def auto_match(self, date_tolerance_days=3):
        """Auto-match unmatched bank rows to expenses by amount and date window."""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT *
            FROM bank_statement_rows
            WHERE id NOT IN (SELECT bank_row_id FROM bank_reconciliation_matches)
            ORDER BY transaction_date, id
        """)
        matched = 0
        for bank_row in cursor.fetchall():
            start_date, end_date = self._date_window(bank_row["transaction_date"], date_tolerance_days)
            cursor.execute("""
                SELECT e.*
                FROM expenses e
                WHERE ABS(COALESCE(e.amount, 0) - ABS(?)) < 0.01
                  AND e.expense_date BETWEEN ? AND ?
                  AND e.id NOT IN (
                      SELECT system_record_id
                      FROM bank_reconciliation_matches
                      WHERE system_record_type = 'expense'
                  )
                ORDER BY ABS(julianday(e.expense_date) - julianday(?)), e.id
            """, (bank_row["amount"], start_date, end_date, bank_row["transaction_date"]))
            candidates = cursor.fetchall()
            expense = self._best_expense_candidate(bank_row, candidates)
            if not expense:
                continue
            confidence = self._confidence(bank_row, expense)
            cursor.execute("""
                INSERT INTO bank_reconciliation_matches
                (bank_row_id, system_record_type, system_record_id, match_status, confidence, matched_by)
                VALUES (?, 'expense', ?, 'auto', ?, 'system')
            """, (bank_row["id"], expense["id"], confidence))
            matched += 1
        self.conn.commit()
        return matched

    def get_match_suggestions(self, limit=100, date_tolerance_days=5):
        """Return unmatched bank rows with best candidate expenses, without writing matches."""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT *
            FROM bank_statement_rows
            WHERE id NOT IN (SELECT bank_row_id FROM bank_reconciliation_matches)
            ORDER BY transaction_date DESC, id DESC
            LIMIT ?
        """, (limit,))
        suggestions = []
        for bank_row in cursor.fetchall():
            start_date, end_date = self._date_window(bank_row["transaction_date"], date_tolerance_days)
            cursor.execute("""
                SELECT e.id, e.expense_date, COALESCE(p.name, '') AS project_name,
                       e.description, e.amount, e.payment_method, e.status
                FROM expenses e
                LEFT JOIN projects p ON p.id = e.project_id
                WHERE ABS(COALESCE(e.amount, 0) - ABS(?)) < 0.01
                  AND e.expense_date BETWEEN ? AND ?
                  AND e.id NOT IN (
                      SELECT system_record_id
                      FROM bank_reconciliation_matches
                      WHERE system_record_type = 'expense'
                  )
                ORDER BY ABS(julianday(e.expense_date) - julianday(?)), e.id
            """, (bank_row["amount"], start_date, end_date, bank_row["transaction_date"]))
            candidates = cursor.fetchall()
            best = self._best_expense_candidate(bank_row, candidates)
            if best:
                suggestions.append({
                    'bank_row': bank_row,
                    'expense': best,
                    'confidence': self._confidence(bank_row, best),
                })
        return suggestions

    def auto_match_bank_transactions(self, date_tolerance_days=3):
        """Auto-match bank_transactions to expenses by amount/date."""
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT *
            FROM bank_transactions
            WHERE COALESCE(reconciled, 0) = 0
            ORDER BY transaction_date, id
        """)
        matched = 0
        for bank_row in cursor.fetchall():
            start_date, end_date = self._date_window(bank_row["transaction_date"], date_tolerance_days)
            cursor.execute("""
                SELECT e.*
                FROM expenses e
                WHERE ABS(COALESCE(e.amount, 0) - ABS(?)) < 0.01
                  AND e.expense_date BETWEEN ? AND ?
                  AND e.id NOT IN (
                      SELECT system_record_id
                      FROM bank_reconciliations
                      WHERE system_record_type = 'expense'
                  )
                ORDER BY ABS(julianday(e.expense_date) - julianday(?)), e.id
                LIMIT 1
            """, (bank_row["amount"], start_date, end_date, bank_row["transaction_date"]))
            expense = cursor.fetchone()
            if not expense:
                continue
            confidence = self._confidence(bank_row, expense)
            self.reconcile(bank_row["id"], "expense", expense["id"], confidence, "system")
            matched += 1
        return matched

    def reconcile(self, bank_transaction_id, system_record_type, system_record_id,
                  confidence=100, reconciled_by=None, notes=""):
        cursor = self.conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO bank_reconciliations
            (bank_transaction_id, system_record_type, system_record_id, confidence,
             status, reconciled_by, notes)
            VALUES (?, ?, ?, ?, 'matched', ?, ?)
        """, (bank_transaction_id, system_record_type, system_record_id,
              confidence, reconciled_by, notes))
        cursor.execute("""
            UPDATE bank_transactions
            SET reconciled = 1, matched_record_type = ?, matched_record_id = ?
            WHERE id = ?
        """, (system_record_type, system_record_id, bank_transaction_id))
        self.conn.commit()
        return cursor.lastrowid

    def get_unreconciled(self):
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT * FROM bank_transactions
            WHERE COALESCE(reconciled, 0) = 0
            ORDER BY transaction_date DESC, id DESC
        """)
        return cursor.fetchall()

    def get_bank_summary(self, bank_account_id=None):
        cursor = self.conn.cursor()
        params = []
        where = ""
        if bank_account_id:
            where = "WHERE bank_account_id = ?"
            params.append(bank_account_id)
        cursor.execute(f"""
            SELECT COUNT(*) AS transaction_count,
                   SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END) AS total_in,
                   SUM(CASE WHEN amount < 0 THEN ABS(amount) ELSE 0 END) AS total_out,
                   SUM(CASE WHEN COALESCE(reconciled, 0) = 1 THEN 1 ELSE 0 END) AS reconciled_count,
                   SUM(CASE WHEN COALESCE(reconciled, 0) = 0 THEN 1 ELSE 0 END) AS unreconciled_count
            FROM bank_transactions
            {where}
        """, params)
        row = cursor.fetchone()
        return dict(row) if row else {}

    def get_unmatched_report(self):
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT b.*
            FROM bank_statement_rows b
            LEFT JOIN bank_reconciliation_matches m ON m.bank_row_id = b.id
            WHERE m.id IS NULL
            ORDER BY b.transaction_date DESC, b.id DESC
        """)
        bank_unmatched = cursor.fetchall()

        cursor.execute("""
            SELECT e.id, e.expense_date, COALESCE(p.name, '') AS project_name,
                   e.description, e.amount, e.payment_method, e.status
            FROM expenses e
            LEFT JOIN projects p ON p.id = e.project_id
            WHERE LOWER(COALESCE(e.payment_method, '')) LIKE '%chuyen%'
              AND e.id NOT IN (
                  SELECT system_record_id
                  FROM bank_reconciliation_matches
                  WHERE system_record_type = 'expense'
              )
            ORDER BY e.expense_date DESC, e.id DESC
        """)
        system_unmatched = cursor.fetchall()
        return {"bank_unmatched": bank_unmatched, "system_unmatched": system_unmatched}

    def get_matches(self):
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT m.id, b.transaction_date, b.amount AS bank_amount, b.description AS bank_description,
                   e.id AS expense_id, e.expense_date, e.amount AS expense_amount,
                   e.description AS expense_description, m.match_status, m.confidence
            FROM bank_reconciliation_matches m
            JOIN bank_statement_rows b ON b.id = m.bank_row_id
            LEFT JOIN expenses e ON m.system_record_type = 'expense' AND e.id = m.system_record_id
            ORDER BY b.transaction_date DESC, m.id DESC
        """)
        return cursor.fetchall()

    def _amount_from_row(self, row):
        amount = self._pick(row, "amount", "so_tien", "gia_tri")
        debit = self._pick(row, "debit", "ghi_no", "rut_ra")
        credit = self._pick(row, "credit", "ghi_co", "nap_vao")
        if amount:
            return self._parse_amount(amount)
        if debit:
            return -abs(self._parse_amount(debit))
        if credit:
            return abs(self._parse_amount(credit))
        return 0.0

    def _confidence(self, bank_row, expense):
        score = 70
        bank_text = self._normalize_text(bank_row["description"] or "")
        expense_text = self._normalize_text(expense["description"] or "")
        if bank_text and expense_text:
            bank_tokens = set(bank_text.split())
            expense_tokens = set(expense_text.split())
            overlap = len(bank_tokens & expense_tokens)
            if bank_text in expense_text or expense_text in bank_text:
                score += 20
            elif overlap:
                score += min(20, overlap * 5)
        if str(bank_row["transaction_date"]) == str(expense["expense_date"]):
            score += 10
        return min(score, 100)

    def _best_expense_candidate(self, bank_row, candidates):
        if not candidates:
            return None
        return max(candidates, key=lambda expense: self._confidence(bank_row, expense))

    def _normalize_text(self, value):
        text = unicodedata.normalize('NFKD', str(value or '').lower())
        text = ''.join(ch for ch in text if not unicodedata.combining(ch))
        text = re.sub(r'[^a-z0-9]+', ' ', text)
        return re.sub(r'\s+', ' ', text).strip()

    def _date_window(self, value, days):
        dt = datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
        return (dt - timedelta(days=days)).isoformat(), (dt + timedelta(days=days)).isoformat()

    def _parse_date(self, value):
        for fmt in self.DATE_FORMATS:
            try:
                return datetime.strptime(str(value).strip()[:10], fmt).date().isoformat()
            except ValueError:
                continue
        return ""

    def _parse_amount(self, value):
        text = str(value or "").replace(",", "").replace(" ", "")
        if text.count(".") > 1:
            text = text.replace(".", "")
        if "," in str(value) and "." not in str(value):
            text = str(value).replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0

    def _clean_key(self, key):
        return str(key or "").strip().lower().replace(" ", "_").replace("-", "_")

    def _pick(self, row, *keys):
        for key in keys:
            if row.get(key):
                return row[key]
        return ""

    def _read_statement_rows(self, path):
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("Can cai openpyxl de nhap sao ke Excel") from exc
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(value or "").strip() for value in rows[0]]
            return [
                {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
                for row in rows[1:]
            ]
        with open(path, newline="", encoding="utf-8-sig") as handle:
            return list(csv.DictReader(handle))

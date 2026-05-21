"""
MODULE REPORTS - Tạo báo cáo & biểu đồ thống kê
"""

import sqlite3
import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from database import get_connection
try:
    import matplotlib
    matplotlib.use('TkAgg')
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except Exception:
    FigureCanvasTkAgg = None
    Figure = None
    MATPLOTLIB_AVAILABLE = False


class ReportGenerator:
    """Tạo báo cáo tài chính."""

    def __init__(self):
        self.conn = get_connection()

    def _charts_available(self, parent):
        if MATPLOTLIB_AVAILABLE:
            return True
        tk.Label(
            parent,
            text="Chưa cài matplotlib nên chưa hiển thị được biểu đồ. Các bảng báo cáo vẫn dùng bình thường.",
            font=('Arial', 10),
            bg='#f0f4f8',
            fg='#e74c3c',
            wraplength=520,
            justify='left',
        ).pack(pady=20)
        return False

    def get_expense_summary(self, start_date=None, end_date=None):
        """Lấy tóm tắt chi phí."""
        cursor = self.conn.cursor()

        query = '''
            SELECT ec.name, SUM(e.amount) as total, COUNT(e.id) as count
            FROM expenses e
            JOIN expense_categories ec ON e.category_id = ec.id
        '''

        params = []
        if start_date and end_date:
            query += ' WHERE e.expense_date BETWEEN ? AND ?'
            params = [start_date, end_date]

        query += ' GROUP BY e.category_id ORDER BY total DESC'

        cursor.execute(query, params)
        return cursor.fetchall()

    def get_project_expense_summary(self):
        """Lấy tóm tắt chi phí theo dự án."""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT p.name, SUM(e.amount) as total, COUNT(e.id) as count
            FROM expenses e
            LEFT JOIN projects p ON e.project_id = p.id
            GROUP BY e.project_id
            ORDER BY total DESC
        ''')
        return cursor.fetchall()

    def generate_income_statement(self, start_date, end_date):
        """Tạo báo cáo kết quả hoạt động."""
        cursor = self.conn.cursor()

        # Lấy tổng chi phí theo các nhóm
        cursor.execute('''
            SELECT ec.name, SUM(e.amount) as total
            FROM expenses e
            JOIN expense_categories ec ON e.category_id = ec.id
            WHERE e.expense_date BETWEEN ? AND ?
            GROUP BY e.category_id
        ''', (start_date, end_date))

        expenses = cursor.fetchall()

        report = {
            'period': f"{start_date} to {end_date}",
            'expenses': expenses,
            'total_expenses': sum([exp[1] for exp in expenses])
        }

        return report

    def get_cash_flow_statement(self, start_date, end_date):
        """Tạo báo cáo lưu chuyển tiền mặt."""
        cursor = self.conn.cursor()

        cursor.execute('''
            SELECT 
                DATE(expense_date) as date,
                payment_method,
                SUM(amount) as total
            FROM expenses
            WHERE expense_date BETWEEN ? AND ?
            GROUP BY DATE(expense_date), payment_method
            ORDER BY date
        ''', (start_date, end_date))

        return cursor.fetchall()

    def get_monthly_expense_summary(self):
        """Lấy chi phí theo tháng."""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT strftime('%Y-%m', expense_date) AS month, SUM(amount) AS total
            FROM expenses
            GROUP BY strftime('%Y-%m', expense_date)
            ORDER BY month
        ''')
        return cursor.fetchall()

    def get_project_cost_collection(self, project_id=None):
        """Bảng tập hợp chi phí công trình theo loại."""
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.code, p.name, ec.name, COALESCE(SUM(e.amount), 0)
            FROM expenses e
            JOIN projects p ON e.project_id = p.id
            JOIN expense_categories ec ON e.category_id = ec.id
            WHERE p.code != 'CHUNG'
        '''
        if project_id:
            query += ' AND p.id = ?'
            params.append(project_id)
        query += ' GROUP BY p.id, ec.id ORDER BY p.code, SUM(e.amount) DESC'
        cursor.execute(query, params)
        return cursor.fetchall()

    def get_project_pl_table(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT p.code, p.name,
                   COALESCE((SELECT SUM(amount) FROM project_revenues WHERE project_id = p.id), 0),
                   COALESCE((SELECT SUM(amount) FROM expenses WHERE project_id = p.id), 0)
            FROM projects p WHERE p.code != 'CHUNG'
        '''
        if project_id:
            query += ' AND p.id = ?'
            params.append(project_id)
        cursor.execute(query, params)
        return cursor.fetchall()

    def get_project_wip_report(self, project_id=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT COALESCE(p.code, ''), COALESCE(p.name, ''),
                   COALESCE(SUM(j.amount), 0)
            FROM journal_entries j
            LEFT JOIN projects p ON j.project_id = p.id
            WHERE j.debit_account = '154'
        '''
        if project_id:
            query += ' AND j.project_id = ?'
            params.append(project_id)
        query += ' GROUP BY j.project_id'
        cursor.execute(query, params)
        return cursor.fetchall()

    def get_balance_sheet(self, as_of_date=None, previous_as_of_date=None):
        """Build a balance sheet grouped into Assets and Capital/Liabilities."""
        as_of_date = as_of_date or datetime.now().date().isoformat()
        previous_as_of_date = previous_as_of_date or self._previous_period_end(as_of_date)
        current = self._account_balances_as_of(as_of_date)
        previous = self._account_balances_as_of(previous_as_of_date) if previous_as_of_date else {}
        rows = []
        for code in sorted(set(current) | set(previous)):
            account = current.get(code, previous.get(code))
            group = self._balance_sheet_group(code, account.get('account_type') or account.get('account_class'))
            if not group:
                continue
            current_amount = self._presentation_balance(code, current.get(code, {}).get('balance', 0))
            previous_amount = self._presentation_balance(code, previous.get(code, {}).get('balance', 0))
            rows.append({
                'group': group,
                'account_code': code,
                'account_name': account.get('account_name') or '',
                'current': current_amount,
                'previous': previous_amount,
                'change': current_amount - previous_amount,
            })
        totals = {
            'assets': sum(row['current'] for row in rows if row['group'] == 'Tai san'),
            'capital': sum(row['current'] for row in rows if row['group'] == 'Nguon von'),
            'previous_assets': sum(row['previous'] for row in rows if row['group'] == 'Tai san'),
            'previous_capital': sum(row['previous'] for row in rows if row['group'] == 'Nguon von'),
        }
        totals['difference'] = totals['assets'] - totals['capital']
        totals['previous_difference'] = totals['previous_assets'] - totals['previous_capital']
        return {
            'as_of_date': as_of_date,
            'previous_as_of_date': previous_as_of_date,
            'rows': rows,
            'totals': totals,
        }

    def display_balance_sheet(self, parent):
        report = self.get_balance_sheet()
        summary = tk.Frame(parent, bg='#FFFFFF')
        summary.pack(fill='x', padx=10, pady=(6, 8))
        totals = report['totals']
        text = (
            f"Ngay bao cao: {report['as_of_date']}    "
            f"Tai san: {totals['assets']:,.0f}    "
            f"Nguon von: {totals['capital']:,.0f}    "
            f"Lech: {totals['difference']:,.0f}"
        )
        tk.Label(summary, text=text, bg='#FFFFFF', fg='#17324D',
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w')

        cols = ('Nhom', 'TK', 'Ten tai khoan', 'Ky nay', 'Ky truoc', 'Chenh lech')
        tree = ttk.Treeview(parent, columns=cols, show='headings', height=16)
        for col, width in zip(cols, (110, 90, 260, 130, 130, 130)):
            tree.heading(col, text=col)
            tree.column(col, width=width, anchor='e' if col in cols[3:] else 'w')
        for row in report['rows']:
            tree.insert('', 'end', values=(
                row['group'], row['account_code'], row['account_name'],
                f"{row['current']:,.0f}", f"{row['previous']:,.0f}", f"{row['change']:,.0f}",
            ))
        tree.pack(fill='both', expand=True, padx=10, pady=10)

    def _account_balances_as_of(self, as_of_date):
        cursor = self.conn.cursor()
        balances = {}

        cursor.execute('''
            SELECT account_code, SUM(debit_amount) AS debit, SUM(credit_amount) AS credit
            FROM (
                SELECT debit_account AS account_code, amount AS debit_amount, 0 AS credit_amount
                FROM journal_entries
                WHERE COALESCE(debit_account, '') <> ''
                  AND entry_date <= ?
                  AND COALESCE(is_reversed, 0) = 0
                UNION ALL
                SELECT credit_account AS account_code, 0 AS debit_amount, amount AS credit_amount
                FROM journal_entries
                WHERE COALESCE(credit_account, '') <> ''
                  AND entry_date <= ?
                  AND COALESCE(is_reversed, 0) = 0
                UNION ALL
                SELECT l.account_code, l.debit_amount, l.credit_amount
                FROM journal_entry_lines l
                JOIN journal_entries j ON j.id = l.journal_entry_id
                WHERE j.entry_date <= ? AND COALESCE(j.is_reversed, 0) = 0
            )
            GROUP BY account_code
        ''', (as_of_date, as_of_date, as_of_date))
        for row in cursor.fetchall():
            code = row['account_code']
            balances[code] = {
                'account_code': code,
                'account_name': '',
                'account_type': '',
                'account_class': '',
                'debit': row['debit'] or 0,
                'credit': row['credit'] or 0,
                'balance': (row['debit'] or 0) - (row['credit'] or 0),
            }

        if balances:
            placeholders = ','.join('?' for _ in balances)
            cursor.execute(f'''
                SELECT account_code, account_name, account_type,
                       COALESCE(account_class, '') AS account_class
                FROM accounts
                WHERE account_code IN ({placeholders})
            ''', tuple(balances.keys()))
            for account in cursor.fetchall():
                balances[account['account_code']].update({
                    'account_name': account['account_name'],
                    'account_type': account['account_type'] or '',
                    'account_class': account['account_class'] or '',
                })
        return balances

    def _balance_sheet_group(self, account_code, account_type=''):
        first = str(account_code or '')[:1]
        text = (account_type or '').lower()
        if first in ('1', '2') or 'asset' in text or 'tai san' in text:
            return 'Tai san'
        if first in ('3', '4') or 'liabil' in text or 'equity' in text or 'nguon' in text:
            return 'Nguon von'
        return ''

    def _presentation_balance(self, account_code, balance):
        return -balance if str(account_code or '')[:1] in ('3', '4') else balance

    def _previous_period_end(self, as_of_date):
        try:
            dt = datetime.strptime(str(as_of_date)[:10], '%Y-%m-%d').date()
        except ValueError:
            return None
        first_day = dt.replace(day=1)
        return (first_day - timedelta(days=1)).isoformat()

    def display_project_pl_table(self, parent, project_id=None):
        rows = self.get_project_pl_table(project_id)
        columns = ('Mã DA', 'Dự án', 'Doanh thu', 'Chi phí', 'Lãi/lỗ')
        tree = ttk.Treeview(parent, columns=columns, show='headings', height=12)
        for col, w in zip(columns, (80, 240, 120, 120, 120)):
            tree.heading(col, text=col)
            tree.column(col, width=w)
        for row in rows:
            profit = row[2] - row[3]
            tree.insert('', 'end', values=(row[0], row[1], f"{row[2]:,.0f}", f"{row[3]:,.0f}", f"{profit:,.0f}"))
        tree.pack(fill='both', expand=True, padx=10, pady=10)

    def get_material_stock_summary(self):
        """Lấy tồn kho vật tư theo giá trị."""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT name, quantity, unit_price, quantity * COALESCE(unit_price, 0) AS total_value
            FROM materials
            WHERE status = 'active'
            ORDER BY total_value DESC
            LIMIT 12
        ''')
        return cursor.fetchall()

    def export_to_excel(self, filename, data):
        """Xuất báo cáo ra Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ImportError("Cần cài openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Báo cáo"

        # Header
        ws['A1'] = "CÔNG TY TRUNG HẢI - BÁO CÁO KẾ TOÁN"
        ws['A1'].font = Font(bold=True, size=14)

        # Dữ liệu
        row = 3
        for item in data:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            row += 1

        # Định dạng
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        wb.save(filename)
        return filename

    def display_expense_chart(self, parent):
        """Hiển thị biểu đồ chi phí trong tkinter."""
        try:
            if not self._charts_available(parent):
                return
            # Lấy dữ liệu
            expenses = self.get_expense_summary()

            if not expenses:
                tk.Label(parent, text="Chưa có dữ liệu chi phí",
                        font=('Arial', 12), bg='#f0f4f8', fg='#999').pack(pady=50)
                return

            # Chuẩn bị dữ liệu
            categories = [exp[0] for exp in expenses]
            amounts = [exp[1] for exp in expenses]

            # Tạo biểu đồ
            fig = Figure(figsize=(10, 4), dpi=100, facecolor='#FFFFFF')
            ax = fig.add_subplot(111)
            ax.set_facecolor('#FFFFFF')

            # Biểu đồ cột
            palette = ['#0F4C81', '#2E7D32', '#D68910', '#8E44AD', '#16A085', '#C0392B']
            colors = [palette[i % len(palette)] for i in range(len(categories))]
            bars = ax.bar(categories, amounts, color=colors, edgecolor='#17324D', linewidth=0.8)

            # Định dạng
            ax.set_ylabel('Số tiền (₫)', fontsize=10)
            ax.set_xlabel('Loại chi phí', fontsize=10)
            ax.set_title('Thống kê chi phí theo loại', fontsize=12, fontweight='bold')
            ax.grid(axis='y', color='#E8EEF5', linewidth=0.8)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

            # Xoay nhãn X
            ax.tick_params(axis='x', rotation=45)

            # Thêm giá trị trên cột
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'₫{int(height):,}',
                       ha='center', va='bottom', fontsize=8)

            fig.tight_layout()

            # Hiển thị trong tkinter
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

        except Exception as e:
            tk.Label(parent, text=f"Lỗi hiển thị biểu đồ: {str(e)}",
                    font=('Arial', 10), bg='#f0f4f8', fg='#e74c3c').pack(pady=20)

    def display_project_chart(self, parent):
        """Hiển thị biểu đồ chi phí theo dự án."""
        try:
            if not self._charts_available(parent):
                return
            expenses = self.get_project_expense_summary()

            if not expenses:
                tk.Label(parent, text="Chưa có dữ liệu",
                        font=('Arial', 12), bg='#f0f4f8', fg='#999').pack(pady=50)
                return

            projects = [exp[0] or 'Không có dự án' for exp in expenses]
            amounts = [exp[1] for exp in expenses]

            fig = Figure(figsize=(10, 4), dpi=100, facecolor='#FFFFFF')
            ax = fig.add_subplot(111)

            # Biểu đồ bánh
            colors = ['#0F4C81', '#2E7D32', '#D68910', '#8E44AD', '#16A085', '#C0392B']
            ax.pie(amounts, labels=projects, autopct='%1.1f%%',
                  colors=colors[:len(projects)], startangle=90,
                  wedgeprops={'linewidth': 1, 'edgecolor': 'white'})
            ax.set_title('Phân bổ chi phí theo dự án', fontsize=12, fontweight='bold')

            fig.tight_layout()

            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)

        except Exception as e:
            tk.Label(parent, text=f"Lỗi: {str(e)}",
                    font=('Arial', 10), bg='#f0f4f8', fg='#e74c3c').pack(pady=20)

    def display_monthly_expense_chart(self, parent):
        """Hiển thị biểu đồ chi phí theo tháng."""
        try:
            if not self._charts_available(parent):
                return
            rows = self.get_monthly_expense_summary()
            if not rows:
                tk.Label(parent, text="Chưa có dữ liệu chi phí theo tháng",
                        font=('Arial', 12), bg='#f0f4f8', fg='#999').pack(pady=50)
                return
            months = [row[0] for row in rows]
            totals = [row[1] for row in rows]

            fig = Figure(figsize=(10, 4.1), dpi=100, facecolor='#FFFFFF')
            ax = fig.add_subplot(111)
            ax.set_facecolor('#FFFFFF')
            if len(rows) <= 1:
                ax.bar(months, totals, color='#2E7D32', width=0.42)
                ax.text(0.5, 0.90, 'Hiện chỉ có dữ liệu của 1 tháng',
                        transform=ax.transAxes, ha='center', va='center',
                        fontsize=9, color='#6B7280')
            else:
                ax.plot(months, totals, marker='o', color='#2E7D32', linewidth=2.2)
                ax.fill_between(months, totals, color='#A5D6A7', alpha=0.25)
            ax.set_title('Xu hướng chi phí theo tháng', fontsize=12, fontweight='bold')
            ax.set_ylabel('Số tiền (₫)')
            ax.grid(axis='y', color='#E8EEF5', linewidth=0.8)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.tick_params(axis='x', rotation=45, labelsize=9)
            for label in ax.get_xticklabels():
                label.set_ha('right')
            fig.subplots_adjust(left=0.08, right=0.98, top=0.86, bottom=0.30)
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
        except Exception as e:
            tk.Label(parent, text=f"Lỗi biểu đồ tháng: {str(e)}",
                    font=('Arial', 10), bg='#f0f4f8', fg='#e74c3c').pack(pady=20)

    def display_material_stock_chart(self, parent):
        """Hiển thị biểu đồ giá trị tồn kho vật tư."""
        try:
            if not self._charts_available(parent):
                return
            rows = self.get_material_stock_summary()
            if not rows:
                tk.Label(parent, text="Chưa có dữ liệu tồn kho vật tư",
                        font=('Arial', 12), bg='#f0f4f8', fg='#999').pack(pady=50)
                return
            names = [row[0] for row in rows]
            values = [row[3] or 0 for row in rows]

            fig = Figure(figsize=(10, 3.6), dpi=100, facecolor='#FFFFFF')
            ax = fig.add_subplot(111)
            ax.set_facecolor('#FFFFFF')
            ax.barh(names, values, color='#0F4C81')
            ax.set_title('Giá trị tồn kho vật tư', fontsize=12, fontweight='bold')
            ax.set_xlabel('Giá trị (₫)')
            ax.grid(axis='x', color='#E8EEF5', linewidth=0.8)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            fig.tight_layout()
            canvas = FigureCanvasTkAgg(fig, master=parent)
            canvas.draw()
            canvas.get_tk_widget().pack(fill='both', expand=True)
        except Exception as e:
            tk.Label(parent, text=f"Lỗi biểu đồ vật tư: {str(e)}",
                    font=('Arial', 10), bg='#f0f4f8', fg='#e74c3c').pack(pady=20)

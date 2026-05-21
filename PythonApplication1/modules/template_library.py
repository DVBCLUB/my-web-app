"""
MODULE TEMPLATE LIBRARY - Tách và tạo file biểu mẫu mặc định trong phần mềm.
"""

from pathlib import Path
import re
import shutil
from copy import copy

from database import get_connection
from modules.branding import ensure_logo_asset


COMPANY_INFO = {
    'company_name': 'CÔNG TY CỔ PHẦN XÂY DỰNG VÀ ĐẦU TƯ TRUNG HẢI',
    'company_tax_code': '0312019045',
    'company_representative': 'NGUYỄN ANH MINH',
}


def _safe_name(text):
    text = str(text or 'template').strip()
    text = re.sub(r'[\\/:*?"<>|]+', '_', text)
    text = re.sub(r'\s+', '_', text)
    return text[:90] or 'template'


def _add_company_header(ws):
    try:
        from openpyxl.styles import Font, Alignment
        start_col = 5
        if any(ws.cell(row=row, column=start_col).value for row in range(1, 4)):
            start_col = ws.max_column + 2
        rows = [
            ('Đơn vị', COMPANY_INFO['company_name']),
            ('MST', COMPANY_INFO['company_tax_code']),
            ('Đại diện', COMPANY_INFO['company_representative']),
        ]
        for index, (label, value) in enumerate(rows, 1):
            label_cell = ws.cell(row=index, column=start_col, value=label)
            value_cell = ws.cell(row=index, column=start_col + 1, value=value)
            label_cell.font = Font(bold=True, color='17324D')
            value_cell.font = Font(bold=index == 1, color='17324D')
            value_cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[ws.cell(row=1, column=start_col).column_letter].width = 14
        ws.column_dimensions[ws.cell(row=1, column=start_col + 1).column_letter].width = 48
    except Exception:
        pass


def extract_workbook_sheets_to_templates(template_dir='templates', output_dir='templates/forms'):
    """Tách từng sheet biểu mẫu thành file Excel riêng, giữ nguyên layout mẫu gốc."""
    try:
        import openpyxl
    except ImportError:
        return 0

    template_path = Path(template_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    desktop_sources = [
        (Path.home() / 'OneDrive/Desktop/1. Biễu mẫu QT06 Kế toán - Văn phòng.xlsx', 'Bieu_mau_QT06_Ke_Toan_Van_Phong.xlsx'),
        (Path.home() / 'OneDrive/Desktop/2. Biễu mẫu QT06 Kế toán - Dự án.xlsx', 'Bieu_mau_QT06_Ke_Toan_Du_An.xlsx'),
    ]
    for source, target_name in desktop_sources:
        if source.exists():
            shutil.copy2(source, template_path / target_name)

    source_files = [
        ('Bieu_mau_QT06_Ke_Toan_Van_Phong.xlsx', 'Văn phòng'),
        ('Bieu_mau_QT06_Ke_Toan_Du_An.xlsx', 'Dự án'),
        ('QT06_Ke_Toan_Van_Phong.xlsx', 'Văn phòng'),
        ('QT06_Ke_Toan_Du_An.xlsx', 'Dự án'),
    ]

    conn = get_connection()
    cursor = conn.cursor()
    count = 0

    for filename, scope in source_files:
        source = template_path / filename
        if not source.exists():
            continue
        source_wb = openpyxl.load_workbook(source)
        for sheet_name in source_wb.sheetnames:
            if sheet_name.lower().strip() in ('tong hop', 'văn bản quy định', 'van ban quy dinh'):
                continue
            file_name = f"{_safe_name(Path(filename).stem)}__{_safe_name(sheet_name)}.xlsx"
            target_path = output_path / file_name
            new_wb = openpyxl.load_workbook(source)
            for name in list(new_wb.sheetnames):
                if name != sheet_name:
                    new_wb.remove(new_wb[name])
            new_wb[sheet_name].sheet_state = 'visible'
            new_wb.active = 0
            new_wb.save(target_path)
            new_wb.close()

            form_code = sheet_name.replace(' ', '')
            cursor.execute('''
                UPDATE form_templates
                SET file_path = ?, storage_method = ?, active = 1
                WHERE source_workbook = ? AND sheet_name = ?
            ''', (str(target_path), str(output_path), filename, sheet_name))
            if cursor.rowcount == 0:
                try:
                    cursor.execute('''
                        INSERT INTO form_templates
                        (form_code, form_name, scope, source_workbook, sheet_name, file_path,
                         storage_method, active)
                        VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                    ''', (form_code, sheet_name, scope, filename, sheet_name, str(target_path), str(output_path)))
                except Exception:
                    pass
            count += 1
        source_wb.close()

    conn.commit()
    conn.close()
    return count


def create_construction_template_workbooks(output_dir='templates/forms'):
    """Tạo thêm các biểu mẫu xây dựng/công trường phổ biến dạng Excel có thể chỉnh sửa."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except ImportError:
        return 0

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    templates = [
        ('BM-CT-NKTC', 'Nhật ký thi công công trình',
         ['Ngày', 'Dự án/Công trình', 'Thời tiết', 'Nhân lực', 'Máy móc/Thiết bị', 'Công việc thực hiện', 'Khối lượng', 'Vướng mắc', 'Chữ ký nhà thầu', 'Chữ ký TVGS/Chủ đầu tư']),
        ('BM-CT-YCNT', 'Phiếu yêu cầu nghiệm thu',
         ['Số phiếu', 'Ngày đề nghị', 'Công trình', 'Hạng mục/Công việc', 'Vị trí', 'Thời gian nghiệm thu', 'Hồ sơ kèm theo', 'Đại diện nhà thầu', 'TVGS/Chủ đầu tư']),
        ('BM-CT-NTCV', 'Biên bản nghiệm thu công việc xây dựng',
         ['Số biên bản', 'Ngày nghiệm thu', 'Công trình', 'Công việc nghiệm thu', 'Vị trí', 'Căn cứ nghiệm thu', 'Kết quả kiểm tra', 'Kết luận', 'Nhà thầu', 'TVGS/Chủ đầu tư']),
        ('BM-CT-NTVL', 'Biên bản nghiệm thu vật liệu đầu vào',
         ['Số biên bản', 'Ngày', 'Tên vật liệu', 'Nhà cung cấp', 'Khối lượng', 'Chứng chỉ/CO/CQ', 'Kết quả kiểm tra', 'Kết luận', 'Thủ kho', 'TVGS/Chủ đầu tư']),
        ('BM-CT-NTKL', 'Biên bản nghiệm thu khối lượng hoàn thành',
         ['Số biên bản', 'Ngày', 'Hạng mục', 'Đơn vị', 'Khối lượng hợp đồng', 'Khối lượng kỳ này', 'Lũy kế', 'Đơn giá', 'Thành tiền', 'Ghi chú']),
        ('BM-CT-HTHM', 'Biên bản nghiệm thu hoàn thành hạng mục/công trình',
         ['Số biên bản', 'Ngày', 'Công trình/Hạng mục', 'Thành phần tham gia', 'Hồ sơ hoàn công', 'Đánh giá chất lượng', 'Tồn tại cần khắc phục', 'Kết luận', 'Các bên ký xác nhận']),
        ('BM-CT-BGMB', 'Biên bản bàn giao mặt bằng',
         ['Số biên bản', 'Ngày', 'Địa điểm', 'Bên giao', 'Bên nhận', 'Hiện trạng mặt bằng', 'Mốc giới/Tài liệu kèm theo', 'Ý kiến các bên', 'Chữ ký']),
        ('BM-CT-KHNT', 'Kế hoạch kiểm tra nghiệm thu công việc xây dựng',
         ['STT', 'Công việc/Hạng mục', 'Giai đoạn', 'Tài liệu áp dụng', 'Biểu mẫu sử dụng', 'Người phụ trách', 'Thời điểm kiểm tra', 'Ghi chú']),
        ('BM-HD-XL', 'Hợp đồng giao nhận thầu xây lắp công trình',
         ['Số hợp đồng', 'Ngày ký', 'Bên giao thầu', 'Bên nhận thầu', 'Tên công trình', 'Phạm vi công việc', 'Giá trị hợp đồng', 'Tiến độ', 'Thanh toán', 'Bảo hành', 'Đại diện ký']),
        ('BM-HD-TL', 'Biên bản thanh lý hợp đồng xây dựng',
         ['Số biên bản', 'Ngày', 'Số hợp đồng', 'Các bên tham gia', 'Khối lượng đã thực hiện', 'Giá trị quyết toán', 'Công nợ còn lại', 'Hồ sơ bàn giao', 'Kết luận thanh lý', 'Chữ ký']),
        ('BM-ATLD-KT', 'Phiếu kiểm tra an toàn lao động công trường',
         ['Ngày kiểm tra', 'Công trình', 'Khu vực', 'Nội dung kiểm tra', 'Kết quả', 'Rủi ro/Tồn tại', 'Biện pháp xử lý', 'Người phụ trách', 'Hạn hoàn thành', 'Trạng thái']),
        ('BM-GB-CT', 'Biên bản họp giao ban công trường',
         ['Số biên bản', 'Ngày họp', 'Thành phần', 'Nội dung đã thực hiện', 'Tồn tại', 'Kế hoạch tiếp theo', 'Người phụ trách', 'Thời hạn', 'Kết luận']),
    ]

    conn = get_connection()
    cursor = conn.cursor()
    count = 0
    header_fill = PatternFill('solid', fgColor='17324D')
    label_fill = PatternFill('solid', fgColor='E8EEF5')
    thin = Side(style='thin', color='AAB7C4')

    for code, name, fields in templates:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = code[:31]
        try:
            from openpyxl.drawing.image import Image
            logo_path = ensure_logo_asset()
            if logo_path:
                logo = Image(logo_path)
                logo.width = 145
                logo.height = 60
                ws.add_image(logo, 'A1')
        except Exception:
            pass
        _add_company_header(ws)
        ws['A1'] = name.upper()
        ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
        ws['A1'].fill = header_fill
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws['A3'] = 'Mã biểu mẫu'
        ws['B3'] = code
        ws['C3'] = 'Ngày lập'
        ws['D3'] = ''
        ws['A5'] = 'STT'
        ws['B5'] = 'Trường dữ liệu'
        ws['C5'] = 'Nội dung nhập'
        ws['D5'] = 'Ghi chú'
        for cell in ws[5]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for index, field in enumerate(fields, 1):
            row = 5 + index
            ws.cell(row=row, column=1, value=index)
            ws.cell(row=row, column=2, value=field)
            ws.cell(row=row, column=3, value='')
            ws.cell(row=row, column=4, value='')
        sign_row = 8 + len(fields)
        ws.cell(sign_row, 1, 'Người đề nghị')
        ws.cell(sign_row, 2, 'Trưởng phòng ban')
        ws.cell(sign_row, 3, 'Kế toán')
        ws.cell(sign_row, 4, 'Giám đốc/Chủ đầu tư')
        for col in range(1, 5):
            ws.cell(sign_row, col).font = Font(bold=True)
            ws.cell(sign_row, col).alignment = Alignment(horizontal='center')
            ws.cell(sign_row + 1, col, '(Ký, ghi rõ họ tên)')
            ws.cell(sign_row + 1, col).alignment = Alignment(horizontal='center')
            ws.cell(sign_row + 1, col).font = Font(italic=True, size=9)
            ws.cell(sign_row + 5, col, '')
        for row in ws.iter_rows(min_row=3, max_row=sign_row + 5, min_col=1, max_col=4):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        for col in range(1, 5):
            ws.cell(sign_row, col).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(sign_row + 1, col).alignment = Alignment(horizontal='center', vertical='center')
        for row_index in range(sign_row + 2, sign_row + 5):
            ws.row_dimensions[row_index].height = 26
        for cell in ws[3]:
            cell.fill = label_fill
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 34
        ws.column_dimensions['C'].width = 52
        ws.column_dimensions['D'].width = 28
        ws.freeze_panes = 'A6'
        target_path = output_path / f"{code}_{_safe_name(name)}.xlsx"
        wb.save(target_path)
        wb.close()

        try:
            cursor.execute('''
                INSERT INTO form_templates
                (form_code, form_name, scope, source_workbook, sheet_name, file_path,
                 used_when, required_signatures, storage_owner, storage_method, usage_notes, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (code, name, 'Công trường/Xây dựng', 'Bộ mẫu xây dựng bổ sung', code,
                  str(target_path), name, 'Các bên liên quan theo mẫu', 'Phòng Kế hoạch/Kế toán/Công trường',
                  str(output_path), 'Mẫu Excel mặc định có thể chỉnh sửa trường dữ liệu'))
        except Exception:
            cursor.execute('''
                UPDATE form_templates
                SET form_name = ?, file_path = ?, used_when = ?, storage_method = ?, active = 1
                WHERE form_code = ? AND source_workbook = ?
            ''', (name, str(target_path), name, str(output_path), code, 'Bộ mẫu xây dựng bổ sung'))

        for order, field in enumerate(fields, 1):
            field_key = _safe_name(field).lower()
            try:
                cursor.execute('''
                    INSERT INTO form_template_fields
                    (form_code, field_key, field_label, field_type, required, display_order)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (code, field_key, field, 'text', 0, order))
            except Exception:
                cursor.execute('''
                    UPDATE form_template_fields
                    SET field_label = ?, display_order = ?, active = 1
                    WHERE form_code = ? AND field_key = ?
                ''', (field, order, code, field_key))
        count += 1

    conn.commit()
    conn.close()
    return count


def add_logo_to_existing_templates(template_dir='templates/forms'):
    """Chèn logo vào các file biểu mẫu Excel hiện có nếu mở được."""
    try:
        import openpyxl
        from openpyxl.drawing.image import Image
    except ImportError:
        return 0

    logo_path = ensure_logo_asset()
    if not logo_path:
        return 0

    count = 0
    for path in Path(template_dir).glob('*.xlsx'):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            has_image = bool(getattr(ws, '_images', []))
            if not has_image:
                logo = Image(logo_path)
                logo.width = 145
                logo.height = 60
                ws.add_image(logo, 'A1')
                wb.save(path)
                count += 1
            wb.close()
        except Exception:
            continue
    return count


def add_company_info_to_existing_templates(template_dir='templates/forms'):
    """Khong chen them thong tin vao mau goc de tranh lam lech layout in."""
    return 0


def infer_default_field_mappings(template_dir='templates/forms'):
    """Tao mapping o dien du lieu tu cac nhan co san, khong sua file Excel."""
    try:
        import openpyxl
    except ImportError:
        return 0

    label_map = {
        'tên công ty': 'company_name',
        'đơn vị': 'company_name',
        'mã số thuế': 'company_tax_code',
        'mst': 'company_tax_code',
        'đại diện': 'company_representative',
        'người đại diện': 'company_representative',
        'số chứng từ': 'document_number',
        'số văn bản': 'document_number',
        'số phiếu': 'document_number',
        'ngày lập': 'document_date',
        'ngày đề nghị': 'document_date',
        'ngày tháng': 'document_date',
        'ngày thanh toán': 'document_date',
        'tên dự án': 'project_name',
        'công trình': 'project_name',
        'người đề nghị': 'requester',
        'nhà cung cấp': 'partner_name',
        'đối tác': 'partner_name',
        'người nhận': 'partner_name',
        'số tiền': 'amount',
        'giá trị': 'amount',
        'nội dung': 'content',
        'v/v': 'content',
        'lý do': 'content',
        'hồ sơ đính kèm': 'attachments',
        'phòng ban': 'department',
        'mục đích': 'purpose',
        'mục đích sử dụng': 'purpose',
        'danh sách': 'item_list',
        'người lập': 'prepared_by',
        'kế toán': 'accounting_staff',
        'trưởng phòng': 'department_head',
        'trưởng phòng ban': 'department_head',
        'người phê duyệt': 'approved_by',
    }
    ignored_labels = ('ngày ban hành', 'lần ban hành', 'mã số biểu mẫu', 'số tiền bằng chữ')

    def upsert_mapping(template, field_key, cell_address):
        cursor.execute('''
            INSERT INTO form_field_mappings
            (form_template_id, form_code, sheet_name, field_key, cell_address, row_mode, active, updated_at)
            VALUES (?, ?, ?, ?, ?, 'fixed', 1, CURRENT_TIMESTAMP)
            ON CONFLICT(form_code, sheet_name, field_key) DO UPDATE SET
                form_template_id = excluded.form_template_id,
                cell_address = excluded.cell_address,
                row_mode = excluded.row_mode,
                active = 1,
                updated_at = CURRENT_TIMESTAMP
        ''', (template['id'], template['form_code'], template['sheet_name'], field_key, cell_address))

    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, form_code, COALESCE(sheet_name, '') AS sheet_name, COALESCE(file_path, '') AS file_path
        FROM form_templates
        WHERE active = 1 AND COALESCE(file_path, '') <> ''
    ''')
    templates = cursor.fetchall()
    count = 0
    for template in templates:
        path = Path(template['file_path'])
        if not path.exists() or path.suffix.lower() != '.xlsx':
            continue
        try:
            wb = openpyxl.load_workbook(path, data_only=False)
            ws = wb.active
            table_header_rows = set()
            for row in ws.iter_rows():
                normalized = [
                    re.sub(r'\s+', ' ', str(cell.value or '').strip().lower())
                    for cell in row
                ]
                if 'stt' not in normalized:
                    continue
                row_number = row[0].row
                table_header_rows.add(row_number)
                total_amount_col = None
                amount_col = None
                for cell in row:
                    text = re.sub(r'\s+', ' ', str(cell.value or '').strip().lower())
                    if 'tổng cộng' in text or 'thành tiền' in text:
                        total_amount_col = cell.column
                    elif 'số tiền' in text:
                        amount_col = cell.column
                    elif 'nội dung' in text or 'danh sách' in text:
                        upsert_mapping(template, 'content', ws.cell(row=row_number + 1, column=cell.column).coordinate)
                        upsert_mapping(template, 'item_list', ws.cell(row=row_number + 1, column=cell.column).coordinate)
                    elif 'ngày' in text:
                        upsert_mapping(template, 'document_date', ws.cell(row=row_number + 1, column=cell.column).coordinate)
                    elif any(marker in text for marker in ('nhà thầu', 'nhà cung cấp', 'đối tác', 'công ty')):
                        upsert_mapping(template, 'partner_name', ws.cell(row=row_number + 1, column=cell.column).coordinate)
                    elif 'ghi chú' in text:
                        upsert_mapping(template, 'notes', ws.cell(row=row_number + 1, column=cell.column).coordinate)
                if total_amount_col or amount_col:
                    upsert_mapping(
                        template,
                        'amount',
                        ws.cell(row=row_number + 1, column=total_amount_col or amount_col).coordinate
                    )
            for row in ws.iter_rows():
                for cell in row:
                    if cell.row in table_header_rows:
                        continue
                    text = re.sub(r'\s+', ' ', str(cell.value or '').strip().lower())
                    if not text:
                        continue
                    if any(label in text for label in ignored_labels):
                        continue
                    field_key = None
                    for marker, key in label_map.items():
                        if marker in text:
                            field_key = key
                            break
                    if not field_key or cell.column >= ws.max_column:
                        continue
                    target = ws.cell(row=cell.row, column=cell.column + 1)
                    upsert_mapping(template, field_key, target.coordinate)
                    count += 1
            wb.close()
        except Exception:
            continue
    conn.commit()
    conn.close()
    return count

"""
MODULE TEMPLATE RENDERER - Sinh biểu mẫu để in từ mẫu Excel gốc.
"""

from datetime import datetime
from pathlib import Path
import re
import shutil

from database import get_connection
from modules.branding import ensure_logo_asset


class TemplateRenderer:
    """Tạo file biểu mẫu để in nhưng giữ nguyên layout Excel gốc."""

    def __init__(self):
        self.conn = get_connection()
        self.output_dir = Path('documents/generated_forms')
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def get_company_settings(self):
        defaults = {
            'company_name': 'CÔNG TY CỔ PHẦN XÂY DỰNG VÀ ĐẦU TƯ TRUNG HẢI',
            'company_tax_code': '0312019045',
            'company_representative': 'NGUYỄN ANH MINH',
            'company_short_name': 'TRUNG HẢI',
            'legal_representative': 'NGUYỄN ANH MINH',
            'director': 'NGUYỄN ANH MINH',
            'approved_by': 'NGUYỄN ANH MINH',
        }
        try:
            cursor = self.conn.cursor()
            cursor.execute('SELECT key, value FROM app_settings WHERE key LIKE "company_%"')
            for row in cursor.fetchall():
                defaults[row['key']] = row['value']
            representative = defaults.get('company_representative', '')
            defaults['legal_representative'] = representative
            defaults['director'] = representative
            defaults['approved_by'] = representative
        except Exception:
            pass
        return defaults

    def get_forms(self, keyword=None):
        cursor = self.conn.cursor()
        params = []
        query = '''
            SELECT id, form_code, form_name, COALESCE(scope, '') AS scope,
                   COALESCE(file_path, '') AS file_path,
                   COALESCE(source_workbook, '') AS source_workbook
            FROM form_templates
            WHERE active = 1
        '''
        if keyword:
            query += " AND (form_code LIKE ? OR form_name LIKE ? OR COALESCE(scope, '') LIKE ?)"
            search = f'%{keyword}%'
            params.extend([search, search, search])
        query += ' ORDER BY form_code, form_name'
        cursor.execute(query, params)
        return cursor.fetchall()

    def get_form(self, form_id):
        cursor = self.conn.cursor()
        cursor.execute('SELECT * FROM form_templates WHERE id = ?', (form_id,))
        return cursor.fetchone()

    def get_form_fields(self, form_id):
        form = self.get_form(form_id)
        if not form:
            return []
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT field_key, field_label, field_type, required,
                   COALESCE(default_value, ''), display_order
            FROM form_template_fields
            WHERE active = 1 AND (form_template_id = ? OR form_code = ?)
            ORDER BY display_order, id
        ''', (form_id, form['form_code']))
        return cursor.fetchall()

    def get_field_mappings(self, form_id):
        form = self.get_form(form_id)
        if not form:
            return []
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT m.id, m.form_template_id, m.form_code, COALESCE(m.sheet_name, '') AS sheet_name,
                   m.field_key, m.cell_address, COALESCE(m.row_mode, 'fixed') AS row_mode,
                   COALESCE(f.field_type, 'text') AS field_type,
                   COALESCE(f.default_value, '') AS default_value
            FROM form_field_mappings m
            LEFT JOIN form_template_fields f
              ON f.form_template_id = m.form_template_id AND f.field_key = m.field_key
            WHERE m.active = 1 AND (m.form_template_id = ? OR m.form_code = ?)
            ORDER BY m.id
        ''', (form_id, form['form_code']))
        return cursor.fetchall()

    def get_recommended_forms_for_expense(self, expense_id):
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT e.description, COALESCE(e.amount, 0) AS amount,
                   COALESCE(ec.name, '') AS category_name
            FROM expenses e
            LEFT JOIN expense_categories ec ON e.category_id = ec.id
            WHERE e.id = ?
        ''', (expense_id,))
        expense = cursor.fetchone()
        text = ''
        amount = 0
        if expense:
            text = f"{expense['description'] or ''} {expense['category_name'] or ''}".lower()
            amount = expense['amount'] or 0

        preferred_scope = 'Văn phòng' if amount > 10000000 else 'Dự án'

        keywords = [
            'Đề nghị thanh toán', 'Phiếu chi', 'Phiếu nhập', 'Phiếu xuất',
            'Biên bản giao', 'Đề xuất', 'Tạm ứng', 'Hoàn ứng',
            'Hóa đơn', 'Nghiệm thu', 'Bảng kê',
            'BM07', 'BM08', 'BM03', 'BM05', 'BM06',
        ]
        if any(marker in text for marker in ['vpp', 'văn phòng', 'van phong']):
            keywords.extend(['Đề nghị mua', 'Đề xuất', 'Phiếu nhập', 'Biên bản giao'])
        if any(marker in text for marker in ['vật tư', 'vat tu', 'vật liệu', 'vat lieu']):
            keywords.extend(['nghiệm thu vật liệu', 'Phiếu nhập', 'Biên bản giao'])
        if any(marker in text for marker in ['nhân công', 'nhan cong']):
            keywords.extend(['chấm công', 'thanh toán'])

        conditions = []
        params = []
        for keyword in keywords:
            conditions.append('(form_name LIKE ? OR form_code LIKE ? OR used_when LIKE ?)')
            search = f'%{keyword}%'
            params.extend([search, search, search])
        query = '''
            SELECT id, form_code, form_name, COALESCE(scope, '') AS scope,
                   COALESCE(file_path, '') AS file_path
            FROM form_templates
            WHERE active = 1
              AND COALESCE(file_path, '') LIKE '%.xlsx'
              AND source_workbook LIKE 'Bieu_mau_QT06%'
              AND COALESCE(scope, '') = ?
              AND (''' + ' OR '.join(conditions) + ''')
            ORDER BY
                CASE
                    WHEN form_code LIKE 'BM07.TT%' THEN 0
                    WHEN form_code LIKE 'BM07.1%' THEN 1
                    WHEN form_code LIKE 'BM07.2%' THEN 2
                    WHEN form_code LIKE 'BM07.3%' THEN 3
                    WHEN form_name LIKE '%thanh toán%' THEN 1
                    WHEN form_name LIKE '%Phiếu chi%' THEN 4
                    WHEN form_name LIKE '%Phiếu nhập%' THEN 5
                    ELSE 9
                END,
                form_code,
                form_name
            LIMIT 80
        '''
        params.insert(0, preferred_scope)
        cursor.execute(query, params)
        rows = cursor.fetchall()
        if rows:
            return rows
        cursor.execute('''
            SELECT id, form_code, form_name, COALESCE(scope, '') AS scope,
                   COALESCE(file_path, '') AS file_path
            FROM form_templates
            WHERE active = 1
              AND COALESCE(file_path, '') LIKE '%.xlsx'
              AND source_workbook LIKE 'Bieu_mau_QT06%'
            ORDER BY
                CASE WHEN COALESCE(scope, '') = ? THEN 0 ELSE 1 END,
                CASE
                    WHEN form_code LIKE 'BM07.TT%' THEN 0
                    WHEN form_code LIKE 'BM07.1%' THEN 1
                    WHEN form_code LIKE 'BM07.2%' THEN 2
                    WHEN form_code LIKE 'BM07.3%' THEN 3
                    ELSE 9
                END,
                form_code,
                form_name
            LIMIT 80
        ''', (preferred_scope,))
        return cursor.fetchall()

    def get_accounting_records(self, record_type, keyword=None, limit=200):
        cursor = self.conn.cursor()
        params = []
        search = f'%{keyword}%' if keyword else None
        if record_type == 'expense':
            query = '''
                SELECT e.id, e.expense_date, COALESCE(p.name, ''), COALESCE(ec.name, ''),
                       e.description, e.amount, e.status
                FROM expenses e
                LEFT JOIN projects p ON e.project_id = p.id
                LEFT JOIN expense_categories ec ON e.category_id = ec.id
                WHERE 1 = 1
            '''
            if keyword:
                query += ' AND (e.description LIKE ? OR p.name LIKE ? OR ec.name LIKE ? OR e.paid_by LIKE ?)'
                params.extend([search, search, search, search])
            query += ' ORDER BY e.expense_date DESC, e.id DESC LIMIT ?'
        elif record_type == 'document':
            query = '''
                SELECT d.id, d.doc_date, COALESCE(d.doc_number, ''), COALESCE(d.supplier_name, ''),
                       d.description, d.amount, d.status
                FROM documents d
                WHERE 1 = 1
            '''
            if keyword:
                query += ' AND (d.doc_number LIKE ? OR d.supplier_name LIKE ? OR d.description LIKE ?)'
                params.extend([search, search, search])
            query += ' ORDER BY d.doc_date DESC, d.id DESC LIMIT ?'
        elif record_type == 'project':
            query = '''
                SELECT p.id, COALESCE(p.code, ''), p.name, COALESCE(p.location, ''),
                       COALESCE(p.budget, 0), COALESCE(p.status, '')
                FROM projects p
                WHERE 1 = 1
            '''
            if keyword:
                query += ' AND (p.code LIKE ? OR p.name LIKE ? OR p.location LIKE ?)'
                params.extend([search, search, search])
            query += ' ORDER BY p.code, p.name LIMIT ?'
        else:
            query = '''
                SELECT m.id, m.code, m.name, COALESCE(m.unit, ''), COALESCE(m.quantity, 0),
                       COALESCE(m.unit_price, 0), COALESCE(m.category, ''), COALESCE(m.supplier, '')
                FROM materials m
                WHERE 1 = 1
            '''
            if keyword:
                query += ' AND (m.code LIKE ? OR m.name LIKE ? OR m.category LIKE ? OR m.supplier LIKE ?)'
                params.extend([search, search, search, search])
            query += ' ORDER BY m.code, m.name LIMIT ?'
        params.append(limit)
        cursor.execute(query, params)
        return cursor.fetchall()

    def build_context_from_record(self, record_type, record_id):
        cursor = self.conn.cursor()
        context = self.get_company_settings()
        context.update({
            'document_date': datetime.now().strftime('%d/%m/%Y'),
            'prepared_by': '',
            'approved_by': context.get('company_representative', ''),
        })
        if record_type == 'expense':
            cursor.execute('''
                SELECT e.*, COALESCE(p.name, '') AS project_name, COALESCE(p.code, '') AS project_code,
                       COALESCE(ec.name, '') AS category_name
                FROM expenses e
                LEFT JOIN projects p ON e.project_id = p.id
                LEFT JOIN expense_categories ec ON e.category_id = ec.id
                WHERE e.id = ?
            ''', (record_id,))
            row = cursor.fetchone()
            if row:
                context.update({
                    'document_number': f"CP-{row['id']:05d}",
                    'document_date': self._format_date(row['expense_date'] or context['document_date']),
                    'project_name': row['project_name'],
                    'content': row['description'] or '',
                    'expense_description': row['description'] or '',
                    'amount': row['amount'] or 0,
                    'partner_name': row['paid_by'] or '',
                    'requester': row['paid_by'] or '',
                    'category_name': row['category_name'],
                    'payment_method': row['payment_method'] or '',
                    'notes': row['notes'] or '',
                    'department': row['department'] or '',
                    'purpose': row['purpose'] or '',
                    'item_list': row['item_list'] or row['description'] or '',
                    'accounting_staff': row['accounting_staff'] or '',
                    'department_head': row['department_head'] or '',
                    'prepared_by': row['prepared_by'] or '',
                    'attachments': row['attachments'] or '',
                    'signature_requester': row['paid_by'] or row['prepared_by'] or '',
                    'signature_department': row['department_head'] or '',
                    'signature_accounting': row['accounting_staff'] or '',
                    'signature_director': context.get('company_representative', ''),
                })
        elif record_type == 'document':
            cursor.execute('''
                SELECT d.*, COALESCE(p.name, '') AS project_name, COALESCE(ec.name, '') AS category_name
                FROM documents d
                LEFT JOIN projects p ON d.project_id = p.id
                LEFT JOIN expense_categories ec ON d.category_id = ec.id
                WHERE d.id = ?
            ''', (record_id,))
            row = cursor.fetchone()
            if row:
                context.update({
                    'document_number': row['doc_number'] or f"CT-{row['id']:05d}",
                    'document_date': self._format_date(row['doc_date'] or context['document_date']),
                    'project_name': row['project_name'],
                    'partner_name': row['supplier_name'] or '',
                    'content': row['description'] or '',
                    'amount': row['amount'] or 0,
                    'category_name': row['category_name'],
                    'document_type': row['doc_type'],
                })
        elif record_type == 'project':
            cursor.execute('SELECT * FROM projects WHERE id = ?', (record_id,))
            row = cursor.fetchone()
            if row:
                context.update({
                    'document_number': row['code'] or f"DA-{row['id']:05d}",
                    'project_name': row['name'],
                    'content': row['name'],
                    'amount': row['budget'] or 0,
                    'location': row['location'] or '',
                    'status': row['status'] or '',
                })
        elif record_type == 'project_accounting':
            try:
                from modules.project_accounting import ProjectAccountingManager
                mgr = ProjectAccountingManager()
                payload = mgr.export_qt06_data(record_id)
                project = payload.get('project', {})
                dash = payload.get('dashboard', {})
                context.update({
                    'document_number': project.get('code', ''),
                    'project_name': project.get('name', ''),
                    'content': f"Báo cáo kế toán dự án {project.get('name', '')}",
                    'amount': dash.get('spent', 0),
                    'partner_name': project.get('owner_name', ''),
                })
            except Exception:
                pass
        elif record_type == 'material':
            cursor.execute('SELECT * FROM materials WHERE id = ?', (record_id,))
            row = cursor.fetchone()
            if row:
                context.update({
                    'document_number': row['code'] or f"VT-{row['id']:05d}",
                    'content': row['name'],
                    'partner_name': row['supplier'] or '',
                    'amount': (row['quantity'] or 0) * (row['unit_price'] or 0),
                    'material_name': row['name'],
                    'unit': row['unit'] or '',
                    'quantity': row['quantity'] or 0,
                    'unit_price': row['unit_price'] or 0,
                })
        return context

    def render_form(self, form_id, record_type='manual', record_id=None, manual_values=None):
        form = self.get_form(form_id)
        if not form:
            raise ValueError('Không tìm thấy biểu mẫu')
        context = self.get_company_settings()
        if record_type != 'manual' and record_id:
            context.update(self.build_context_from_record(record_type, record_id))
        context.update(manual_values or {})
        if context.get('document_date'):
            context['document_date'] = self._format_date(context['document_date'])

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        form_code = form['form_code'] or f"FORM{form_id}"
        output_path = self.output_dir / f"{form_code}_{timestamp}.xlsx"
        source_path = Path(form['file_path'] or '')

        if source_path.exists() and source_path.suffix.lower() == '.xlsx':
            shutil.copy2(source_path, output_path)
            self._fill_existing_workbook(output_path, form_id, context)
        else:
            self._create_workbook_from_fields(output_path, form, self.get_form_fields(form_id), context)
        return str(output_path)

    def _fill_existing_workbook(self, path, form_id, context):
        import openpyxl
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        mappings = self.get_field_mappings(form_id)
        filled = self._fill_mapped_cells(wb, ws, mappings, context)
        if not filled:
            self._fill_known_cells(ws, context)
        wb.save(path)
        wb.close()

    def _create_workbook_from_fields(self, path, form, fields, context):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (form['form_code'] or 'Bieu mau')[:31]
        self._add_logo(ws)
        self._add_company_block_for_generated_template(ws, context)
        ws['A5'] = (form['form_name'] or 'Biểu mẫu').upper()
        ws['A5'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A5'].fill = PatternFill('solid', fgColor='17324D')
        ws.merge_cells('A5:D5')
        ws.append([])
        ws.append(['STT', 'Trường dữ liệu', 'Kiểu dữ liệu', 'Nội dung'])
        thin = Side(style='thin', color='AAB7C4')
        for cell in ws[7]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='17324D')
            cell.alignment = Alignment(horizontal='center')
        for index, field in enumerate(fields, 1):
            key = field['field_key']
            field_type = field['field_type'] or 'text'
            default_value = field['default_value'] if 'default_value' in field.keys() else ''
            output_value = context.get(key, default_value)
            ws.append([index, field['field_label'], field_type, output_value])
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 60
        wb.save(path)
        wb.close()

    def _fill_mapped_cells(self, wb, default_ws, mappings, context):
        filled = 0
        for mapping in mappings:
            value = context.get(mapping['field_key'])
            if value in (None, ''):
                continue
            ws = wb[mapping['sheet_name']] if mapping['sheet_name'] and mapping['sheet_name'] in wb.sheetnames else default_ws
            cell_address = str(mapping['cell_address'] or '').strip()
            if not cell_address:
                continue
            try:
                self._write_cell(ws, cell_address, value)
                filled += 1
            except Exception:
                continue
        return filled

    def _add_logo(self, ws):
        logo_path = ensure_logo_asset()
        if not logo_path:
            return
        try:
            from openpyxl.drawing.image import Image
            logo = Image(logo_path)
            logo.width = 145
            logo.height = 60
            ws.add_image(logo, 'A1')
        except Exception:
            return

    def _add_company_block_for_generated_template(self, ws, context):
        try:
            from openpyxl.styles import Font, Alignment
            rows = [
                ('Đơn vị', context.get('company_name', '')),
                ('MST', context.get('company_tax_code', '')),
                ('Đại diện', context.get('company_representative', '')),
            ]
            for index, (label, value) in enumerate(rows, 1):
                ws.cell(row=index, column=2, value=label).font = Font(bold=True, color='17324D')
                cell = ws.cell(row=index, column=3, value=value)
                cell.font = Font(bold=index == 1, color='17324D')
                cell.alignment = Alignment(wrap_text=True)
        except Exception:
            return

    def _fill_known_cells(self, ws, context):
        """Điền thận trọng vào ô trống cạnh nhãn có sẵn, không thêm dòng/cột."""
        label_map = {
            'số phiếu': 'document_number',
            'số chứng từ': 'document_number',
            'số văn bản': 'document_number',
            'ngày lập': 'document_date',
            'ngày đề nghị': 'document_date',
            'ngày tháng': 'document_date',
            'ngày thanh toán': 'document_date',
            'công trình': 'project_name',
            'dự án': 'project_name',
            'nội dung': 'content',
            'lý do': 'content',
            'số tiền': 'amount',
            'giá trị': 'amount',
            'nhà cung cấp': 'partner_name',
            'người nhận': 'partner_name',
            'người đề nghị': 'requester',
            'người lập': 'prepared_by',
            'người duyệt': 'approved_by',
            'kế toán': 'accounting_staff',
            'trưởng phòng': 'department_head',
            'phòng ban': 'department',
            'mục đích': 'purpose',
            'danh sách': 'item_list',
            'tên công ty': 'company_name',
            'đơn vị': 'company_name',
            'mã số thuế': 'company_tax_code',
            'mst': 'company_tax_code',
            'đại diện': 'company_representative',
            'người đại diện': 'company_representative',
            'giám đốc': 'director',
        }
        for row in ws.iter_rows():
            for cell in row:
                text = self._normalize_text(cell.value)
                if not text:
                    continue
                if any(label in text for label in ('ngày ban hành', 'lần ban hành', 'mã số biểu mẫu', 'số tiền bằng chữ')):
                    continue
                matched_key = None
                for marker, key in label_map.items():
                    if marker in text and key in context:
                        matched_key = key
                        break
                if not matched_key:
                    continue
                target_col = cell.column + 1
                if target_col <= ws.max_column:
                    target = ws.cell(row=cell.row, column=target_col)
                    writable = self._writable_cell(ws, target.coordinate)
                    if not writable.value:
                        writable.value = context.get(matched_key, '')

    def _normalize_text(self, value):
        text = str(value or '').strip().lower()
        return re.sub(r'\s+', ' ', text)

    def _format_date(self, value):
        if not value:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%d/%m/%Y')
        text = str(value).strip()
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(text[:10], fmt).strftime('%d/%m/%Y')
            except ValueError:
                continue
        return text

    def _write_cell(self, ws, cell_address, value):
        cell = self._writable_cell(ws, cell_address)
        cell.value = value

    def _writable_cell(self, ws, cell_address):
        cell = ws[cell_address]
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
        return cell

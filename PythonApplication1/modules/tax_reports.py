"""VAT tax report automation for monthly or quarterly filing."""

import tkinter as tk
from tkinter import ttk
from pathlib import Path

from database import get_connection


class VATReportManager:
    """Build VAT input/output schedules and deduction-method declaration."""

    def __init__(self):
        self.conn = get_connection()

    def get_input_vat_schedule(self, start_date, end_date):
        cursor = self.conn.cursor()
        rows = []
        cursor.execute("""
            SELECT d.doc_date, d.doc_number, d.supplier_name,
                   COALESCE(s.tax_code, '') AS tax_code,
                   COALESCE(d.amount, 0) AS total_amount,
                   COALESCE(d.vat_rate, 10) AS vat_rate
            FROM documents d
            LEFT JOIN suppliers s ON s.id = d.supplier_id
            WHERE d.doc_date BETWEEN ? AND ?
              AND COALESCE(d.amount, 0) > 0
            ORDER BY d.doc_date, d.id
        """, (start_date, end_date))
        for row in cursor.fetchall():
            total = float(row["total_amount"] or 0)
            vat_rate = max(float(row["vat_rate"] or 10), 0) / 100
            taxable = total / (1 + vat_rate) if vat_rate else total
            vat = total - taxable
            rows.append({
                "invoice_date": row["doc_date"],
                "invoice_number": row["doc_number"],
                "partner_name": row["supplier_name"],
                "tax_code": row["tax_code"],
                "taxable_amount": taxable,
                "vat_amount": vat,
                "total_amount": total,
            })

        cursor.execute("""
            SELECT j.entry_date, j.description, j.amount
            FROM journal_entries j
            WHERE j.entry_date BETWEEN ? AND ?
              AND j.debit_account LIKE '133%'
        """, (start_date, end_date))
        for row in cursor.fetchall():
            vat = float(row["amount"] or 0)
            rows.append({
                "invoice_date": row["entry_date"],
                "invoice_number": "",
                "partner_name": row["description"] or "But toan VAT dau vao",
                "tax_code": "",
                "taxable_amount": 0,
                "vat_amount": vat,
                "total_amount": vat,
            })
        return rows

    def get_output_vat_schedule(self, start_date, end_date):
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT r.revenue_date, COALESCE(c.contract_no, '') AS contract_no,
                   COALESCE(c.partner_name, '') AS partner_name,
                   r.amount, r.vat_amount, r.description
            FROM project_revenues r
            LEFT JOIN project_contracts c ON c.id = r.contract_id
            WHERE r.revenue_date BETWEEN ? AND ?
            ORDER BY r.revenue_date, r.id
        """, (start_date, end_date))
        rows = []
        for row in cursor.fetchall():
            taxable = float(row["amount"] or 0)
            vat = float(row["vat_amount"] or 0)
            rows.append({
                "invoice_date": row["revenue_date"],
                "invoice_number": row["contract_no"],
                "partner_name": row["partner_name"],
                "tax_code": "",
                "taxable_amount": taxable,
                "vat_amount": vat,
                "total_amount": taxable + vat,
                "description": row["description"] or "",
            })
        return rows

    def get_vat_declaration(self, start_date, end_date):
        input_rows = self.get_input_vat_schedule(start_date, end_date)
        output_rows = self.get_output_vat_schedule(start_date, end_date)
        input_vat = sum(row["vat_amount"] for row in input_rows)
        output_vat = sum(row["vat_amount"] for row in output_rows)
        return {
            "period": f"{start_date} - {end_date}",
            "input_rows": input_rows,
            "output_rows": output_rows,
            "input_taxable": sum(row["taxable_amount"] for row in input_rows),
            "input_vat": input_vat,
            "output_taxable": sum(row["taxable_amount"] for row in output_rows),
            "output_vat": output_vat,
            "vat_payable": max(output_vat - input_vat, 0),
            "vat_credit": max(input_vat - output_vat, 0),
        }

    def get_vat_report(self, start_date, end_date):
        """Alias than thien hon cho cac module khac."""
        return self.get_vat_declaration(start_date, end_date)

    def export_vat_01(self, start_date, end_date, output_path):
        """Xuat bang ke VAT dau vao/dau ra va tom tat to khai 01/GTGT."""
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError("Can cai openpyxl de xuat bao cao thue") from exc
        data = self.get_vat_declaration(start_date, end_date)
        wb = Workbook()
        ws = wb.active
        ws.title = "To khai 01"
        ws.append(["Ky", data["period"]])
        ws.append(["VAT dau ra", data["output_vat"]])
        ws.append(["VAT dau vao", data["input_vat"]])
        ws.append(["Thue phai nop", data["vat_payable"]])
        ws.append(["Con duoc khau tru", data["vat_credit"]])
        self._append_schedule_sheet(wb, "01-2 GTGT dau vao", data["input_rows"])
        self._append_schedule_sheet(wb, "Bang ke dau ra", data["output_rows"])
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        cursor = self.conn.cursor()
        period = f"{start_date}_{end_date}"
        cursor.execute("""
            INSERT INTO tax_declarations
            (tax_type, period, start_date, end_date, output_tax, input_tax,
             payable_amount, status, file_path)
            VALUES ('VAT', ?, ?, ?, ?, ?, ?, 'exported', ?)
            ON CONFLICT(tax_type, period) DO UPDATE SET
                output_tax = excluded.output_tax,
                input_tax = excluded.input_tax,
                payable_amount = excluded.payable_amount,
                status = 'exported',
                file_path = excluded.file_path
        """, (period, start_date, end_date, data["output_vat"], data["input_vat"],
              data["vat_payable"], str(output)))
        self.conn.commit()
        return str(output)

    def get_pit_annual_report(self, year):
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT e.id AS employee_id, e.full_name,
                   SUM(COALESCE(pl.gross_amount, 0)) AS gross_amount,
                   SUM(COALESCE(pl.pit_amount, 0)) AS pit_amount,
                   SUM(COALESCE(pl.net_amount, 0)) AS net_amount
            FROM payroll_lines pl
            JOIN payroll_periods pp ON pp.id = pl.payroll_period_id
            LEFT JOIN employees e ON e.id = pl.employee_id
            WHERE substr(pp.period, 1, 4) = ?
            GROUP BY e.id, e.full_name
            ORDER BY e.full_name
        """, (str(year),))
        return [dict(row) for row in cursor.fetchall()]

    def export_bhxh_list(self, payroll_period, output_path):
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise RuntimeError("Can cai openpyxl de xuat danh sach BHXH") from exc
        cursor = self.conn.cursor()
        cursor.execute("""
            SELECT e.full_name, pp.period, pl.gross_amount,
                   pl.bhxh_employee, pl.bhyt_employee, pl.bhtn_employee
            FROM payroll_lines pl
            JOIN payroll_periods pp ON pp.id = pl.payroll_period_id
            LEFT JOIN employees e ON e.id = pl.employee_id
            WHERE pp.period = ?
            ORDER BY e.full_name
        """, (payroll_period,))
        wb = Workbook()
        ws = wb.active
        ws.title = "BHXH"
        ws.append(["Nhan vien", "Ky", "Luong dong BH", "BHXH", "BHYT", "BHTN"])
        for row in cursor.fetchall():
            ws.append([
                row["full_name"], row["period"], row["gross_amount"],
                row["bhxh_employee"], row["bhyt_employee"], row["bhtn_employee"],
            ])
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return str(output)

    def display_vat_declaration(self, parent, start_date=None, end_date=None):
        start_date = start_date or self._current_month_start()
        end_date = end_date or self._current_month_end()
        data = self.get_vat_declaration(start_date, end_date)
        summary = tk.Frame(parent, bg="#FFFFFF")
        summary.pack(fill="x", padx=10, pady=8)
        text = (
            f"Ky: {data['period']}    VAT dau ra: {data['output_vat']:,.0f}    "
            f"VAT dau vao: {data['input_vat']:,.0f}    Phai nop: {data['vat_payable']:,.0f}    "
            f"Con duoc khau tru: {data['vat_credit']:,.0f}"
        )
        tk.Label(summary, text=text, bg="#FFFFFF", fg="#17324D",
                 font=("Segoe UI", 10, "bold")).pack(anchor="w")

        notebook = ttk.Notebook(parent)
        notebook.pack(fill="both", expand=True, padx=10, pady=8)
        input_frame = tk.Frame(notebook, bg="#FFFFFF")
        output_frame = tk.Frame(notebook, bg="#FFFFFF")
        notebook.add(input_frame, text="01-2/GTGT dau vao")
        notebook.add(output_frame, text="Bang ke dau ra")
        self._render_schedule(input_frame, data["input_rows"])
        self._render_schedule(output_frame, data["output_rows"])

    def _render_schedule(self, parent, rows):
        cols = ("Ngay HD", "So HD/Hop dong", "Doi tac", "MST", "Gia tri truoc thue", "VAT", "Tong")
        tree = ttk.Treeview(parent, columns=cols, show="headings", height=14)
        for col, width in zip(cols, (90, 120, 220, 110, 130, 110, 130)):
            tree.heading(col, text=col)
            tree.column(col, width=width)
        for row in rows:
            tree.insert("", "end", values=(
                row["invoice_date"], row["invoice_number"], row["partner_name"], row["tax_code"],
                f"{row['taxable_amount']:,.0f}", f"{row['vat_amount']:,.0f}",
                f"{row['total_amount']:,.0f}",
            ))
        tree.pack(fill="both", expand=True, padx=8, pady=8)

    def _current_month_start(self):
        from datetime import date
        return date.today().replace(day=1).isoformat()

    def _current_month_end(self):
        from datetime import date, timedelta
        today = date.today()
        if today.month == 12:
            nxt = today.replace(year=today.year + 1, month=1, day=1)
        else:
            nxt = today.replace(month=today.month + 1, day=1)
        return (nxt - timedelta(days=1)).isoformat()

    def _append_schedule_sheet(self, wb, title, rows):
        ws = wb.create_sheet(title)
        ws.append(["Ngay HD", "So HD/Hop dong", "Doi tac", "MST", "Gia tri truoc thue", "VAT", "Tong"])
        for row in rows:
            ws.append([
                row["invoice_date"], row["invoice_number"], row["partner_name"],
                row["tax_code"], row["taxable_amount"], row["vat_amount"], row["total_amount"],
            ])

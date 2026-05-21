"""
MODULE UTILS - Các hàm tiện ích
"""

import os
from datetime import datetime
import re
from importlib.metadata import PackageNotFoundError, version

_pd = None

def _get_pd():
    global _pd
    if _pd is None:
        import pandas as pd
        _pd = pd
    return _pd


def _parse_version_tuple(version_str):
    parts = re.findall(r'\d+', version_str)
    return tuple(int(x) for x in parts[:3]) if parts else (0, 0, 0)


def _ensure_openpyxl_version(min_version='3.1.5'):
    try:
        current_version = version('openpyxl')
    except PackageNotFoundError as exc:
        raise ImportError('Cần cài openpyxl >= 3.1.5 để đọc Excel.') from exc

    if _parse_version_tuple(current_version) < _parse_version_tuple(min_version):
        raise ImportError(
            f"Pandas yêu cầu openpyxl phiên bản >= {min_version}. Hiện tại cài {current_version}."
        )


def format_currency(amount):
    """Định dạng số tiền kiểu Việt Nam."""
    pd = _get_pd()
    try:
        if pd.isna(amount) or amount is None:
            return '0.000'
        num = parse_number(amount)
        return f"{num:,.3f}"
    except (ValueError, TypeError):
        return str(amount)


def parse_number(value):
    """Đọc số kiểu Excel/VN/US: 1,234.56 hoặc 1.234,56 đều được."""
    pd = _get_pd()
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(' ', '')
    if not text:
        return 0.0
    comma = text.rfind(',')
    dot = text.rfind('.')
    if comma > -1 and dot > -1:
        decimal_sep = ',' if comma > dot else '.'
        thousand_sep = '.' if decimal_sep == ',' else ','
        text = text.replace(thousand_sep, '').replace(decimal_sep, '.')
    elif comma > -1:
        parts = text.split(',')
        text = ''.join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else text.replace(',', '.')
    elif dot > -1:
        parts = text.split('.')
        text = ''.join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else text
    return float(text)


def format_money(amount):
    """Số tiền hiển thị 3 số thập phân, có phân tách hàng nghìn."""
    return f"{parse_number(amount):,.3f}"


def format_quantity(quantity):
    """Số lượng hiển thị 4 số thập phân, có phân tách hàng nghìn."""
    return f"{parse_number(quantity):,.4f}"


def format_date(date_value, output_format='%d/%m/%Y'):
    """Định dạng ngày tháng."""
    pd = _get_pd()
    if pd.isna(date_value) or date_value is None or str(date_value).strip() == '':
        return ''

    try:
        if isinstance(date_value, (pd.Timestamp, datetime)):
            return date_value.strftime(output_format)

        dt = pd.to_datetime(date_value, dayfirst=True, errors='coerce')
        if pd.notna(dt):
            return dt.strftime(output_format)
    except Exception:
        pass

    return str(date_value)


def number_to_text_vn(number):
    """Chuyển số thành chữ tiếng Việt (số tiền)."""
    if number is None or str(number).strip() == '':
        return 'Không đồng'

    try:
        num = int(float(str(number).replace(',', '').replace(' ', '')))
    except (ValueError, TypeError):
        return 'Không đồng'

    if num == 0:
        return 'Không đồng chẵn'

    if num < 0:
        return 'Âm ' + number_to_text_vn(-num)

    ones = ['', 'một', 'hai', 'ba', 'bốn', 'năm', 'sáu', 'bảy', 'tám', 'chín']
    tens = ['', 'mười', 'hai mươi', 'ba mươi', 'bốn mươi', 'năm mươi', 
            'sáu mươi', 'bảy mươi', 'tám mươi', 'chín mươi']

    def read_three_digits(n):
        """Đọc 3 chữ số."""
        if n == 0:
            return ''

        hundred = n // 100
        remainder = n % 100
        ten = remainder // 10
        one = remainder % 10

        result = ''

        if hundred > 0:
            result += ones[hundred] + ' trăm'
            if ten == 0 and one > 0:
                result += ' lẻ'

        if ten == 1:
            result += ' mười'
            if one > 0:
                result += ' ' + ones[one]
        elif ten > 1:
            result += ' ' + tens[ten]
            if one == 1:
                result += ' mốt'
            elif one > 0:
                result += ' ' + ones[one]
        elif one > 0 and hundred == 0:
            result += ones[one]

        return result.strip()

    parts = []

    billion = num // 1_000_000_000
    million = (num % 1_000_000_000) // 1_000_000
    thousand = (num % 1_000_000) // 1_000
    remainder = num % 1_000

    if billion > 0:
        parts.append(read_three_digits(billion) + ' tỷ')
    if million > 0:
        parts.append(read_three_digits(million) + ' triệu')
    if thousand > 0:
        parts.append(read_three_digits(thousand) + ' nghìn')
    if remainder > 0:
        parts.append(read_three_digits(remainder))

    result = ' '.join(parts).strip()

    if result:
        result = result[0].upper() + result[1:] + ' đồng chẵn'

    return result


def validate_email(email):
    """Kiểm tra email hợp lệ."""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_phone(phone):
    """Kiểm tra số điện thoại hợp lệ."""
    phone = re.sub(r'\D', '', phone)
    return len(phone) >= 10 and len(phone) <= 11


def slugify(text):
    """Chuyển text thành slug (cho tên file)."""
    text = text.lower()
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    text = re.sub(r'[\s]+', '-', text)
    return text.strip('-')


def safe_filename(filename):
    """Tạo tên file an toàn."""
    filename = os.path.basename(filename)
    # Xóa ký tự không hợp lệ
    invalid_chars = r'[<>:"/\\|?*]'
    filename = re.sub(invalid_chars, '', filename)
    return filename


class ExcelImporter:
    """Nhập dữ liệu từ file Excel."""

    @staticmethod
    def read_excel(file_path, sheet_name=None, header=0):
        """Đọc file Excel."""
        try:
            pd = _get_pd()
            _ensure_openpyxl_version()
            if sheet_name is None:
                # Tự động tìm sheet có dữ liệu
                xl = pd.ExcelFile(file_path)
                sheet_name = xl.sheet_names[0]

            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header)
            # Xóa dòng hoàn toàn trống
            df = df.dropna(how='all')
            return df, sheet_name
        except Exception as e:
            raise Exception(f"Lỗi đọc Excel: {str(e)}")

    @staticmethod
    def import_expenses_from_excel(file_path, expense_manager):
        """Nhập chi phí từ Excel."""
        try:
            df, _ = ExcelImporter.read_excel(file_path)

            results = {'success': 0, 'failed': 0, 'errors': []}

            for idx, row in df.iterrows():
                try:
                    expense_date = format_date(row.get('Ngày', datetime.now().date()), '%Y-%m-%d')
                    project_id = row.get('Dự án ID')
                    category_id = row.get('Loại chi phí ID')
                    description = str(row.get('Mô tả', '')).strip()
                    amount = parse_number(row.get('Số tiền', 0))
                    paid_by = str(row.get('Người chi', '')).strip()
                    payment_method = str(row.get('Hình thức', 'Tiền mặt')).strip()
                    notes = str(row.get('Ghi chú', '')).strip()

                    if not description or amount <= 0:
                        results['failed'] += 1
                        results['errors'].append(f"Dòng {idx+2}: Thiếu mô tả hoặc số tiền không hợp lệ")
                        continue

                    expense_manager.add_expense(
                        expense_date, project_id, category_id, description,
                        amount, paid_by, payment_method, notes, 1,
                        {
                            'department': str(row.get('Phòng ban', '')).strip(),
                            'purpose': str(row.get('Mục đích sử dụng', '')).strip(),
                            'item_list': str(row.get('Danh sách/Nội dung lên mẫu', '')).strip(),
                            'accounting_staff': str(row.get('Kế toán ký', '')).strip(),
                            'department_head': str(row.get('Trưởng phòng ký', '')).strip(),
                            'prepared_by': str(row.get('Người lập', '')).strip(),
                            'attachments': str(row.get('Hồ sơ đính kèm', '')).strip(),
                        }
                    )
                    results['success'] += 1

                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"Dòng {idx+2}: {str(e)}")

            return results

        except Exception as e:
            raise Exception(f"Lỗi nhập dữ liệu: {str(e)}")

    @staticmethod
    def import_materials_from_excel(file_path, material_manager):
        """Nhập danh mục vật tư từ Excel."""
        df, _ = ExcelImporter.read_excel(file_path)
        results = {'success': 0, 'failed': 0, 'errors': []}
        for idx, row in df.iterrows():
            try:
                code = str(row.get('Mã vật tư', row.get('Mã', ''))).strip()
                name = str(row.get('Tên vật tư', row.get('Tên', ''))).strip()
                unit = str(row.get('Đơn vị', '')).strip()
                unit_price = parse_number(row.get('Đơn giá', 0))
                category = str(row.get('Danh mục', '')).strip()
                supplier = str(row.get('Nhà cung cấp', '')).strip()
                quantity = parse_number(row.get('Tồn kho', 0))
                if not code or not name:
                    raise ValueError('Thiếu mã hoặc tên vật tư')
                material_id = material_manager.add_material(code, name, unit, unit_price, category, supplier)
                if quantity > 0:
                    material_manager.add_inventory_transaction(material_id, 'import', quantity, None, 'Nhập tồn đầu từ Excel', 1)
                results['success'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Dòng {idx+2}: {str(e)}")
        return results

    @staticmethod
    def import_accounts_from_excel(file_path, account_manager):
        """Nhập tài khoản kế toán/tiểu mục từ Excel."""
        df, _ = ExcelImporter.read_excel(file_path)
        results = {'success': 0, 'failed': 0, 'errors': []}
        for idx, row in df.iterrows():
            try:
                code = str(row.get('Số hiệu', row.get('Mã tài khoản', ''))).strip()
                name = str(row.get('Tên tài khoản', '')).strip()
                acc_type = str(row.get('Loại', row.get('Loại tài khoản', 'Tùy chỉnh'))).strip()
                parent = str(row.get('Tài khoản cha', '')).strip() or None
                desc = str(row.get('Mô tả', '')).strip()
                if not code or not name:
                    raise ValueError('Thiếu số hiệu hoặc tên tài khoản')
                account_manager.add_account(code, name, acc_type, parent, desc)
                results['success'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Dòng {idx+2}: {str(e)}")
        return results

    @staticmethod
    def import_documents_from_excel(file_path, document_manager):
        """Nhập chứng từ từ Excel."""
        df, _ = ExcelImporter.read_excel(file_path)
        results = {'success': 0, 'failed': 0, 'errors': []}
        for idx, row in df.iterrows():
            try:
                doc_type = str(row.get('Loại', row.get('Loại chứng từ', 'Chứng từ khác'))).strip()
                doc_number = str(row.get('Số CT', row.get('Số chứng từ', ''))).strip()
                doc_date = format_date(row.get('Ngày', row.get('Ngày chứng từ', datetime.now().date())), '%Y-%m-%d')
                supplier = str(row.get('Nhà cung cấp', '')).strip()
                description = str(row.get('Nội dung', row.get('Mô tả', ''))).strip()
                amount = parse_number(row.get('Số tiền', 0))
                expense_id = row.get('Chi phí ID')
                expense_id = int(expense_id) if str(expense_id).strip() not in ('', 'nan', 'None') else None
                if not doc_type or amount < 0:
                    raise ValueError('Thiếu loại chứng từ hoặc số tiền không hợp lệ')
                document_manager.add_document(
                    doc_type, doc_number, doc_date, supplier, description,
                    amount, None, None, None, 1, expense_id, 'draft'
                )
                results['success'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append(f"Dòng {idx+2}: {str(e)}")
        return results


class ExcelExporter:
    """Xuất dữ liệu ra Excel."""

    MONEY_FORMAT = '#,##0.000'
    QUANTITY_FORMAT = '#,##0.0000'

    @staticmethod
    def export_expenses(expenses, filename='expenses_export.xlsx'):
        """Xuất danh sách chi phí."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chi phí"

            # Header
            headers = ['ID', 'Ngày', 'Dự án', 'Loại chi phí', 'Mô tả', 'Số tiền', 'Người chi', 'Trạng thái']
            ws.append(headers)

            # Định dạng header
            header_fill = PatternFill(start_color='1a56a5', end_color='1a56a5', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Dữ liệu
            for expense in expenses:
                ws.append(expense)
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=6).number_format = ExcelExporter.MONEY_FORMAT

            # Độ rộng cột
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12

            wb.save(filename)
            return filename

        except Exception as e:
            raise Exception(f"Lỗi xuất Excel: {str(e)}")

    @staticmethod
    def export_audit_log(rows, filename='audit_log_export.xlsx'):
        """Export audit trail rows to Excel."""
        import openpyxl
        from openpyxl.styles import Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit log"
        ws.append(['Thời gian', 'User', 'Hành động', 'Đối tượng', 'ID', 'Chi tiết'])
        ExcelExporter._style_header(ws)

        for row in rows:
            ws.append([
                row['created_at'],
                row['actor_id'],
                row['action'],
                row['entity_type'],
                row['entity_id'],
                row['detail'],
            ])

        for col, width in {'A': 20, 'B': 10, 'C': 18, 'D': 24, 'E': 10, 'F': 90}.items():
            ws.column_dimensions[col].width = width
        for row_cells in ws.iter_rows(min_row=2):
            row_cells[5].alignment = Alignment(wrap_text=True, vertical='top')
        ws.freeze_panes = 'A2'
        wb.save(filename)
        return filename

    @staticmethod
    def _style_header(ws):
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill(start_color='1a56a5', end_color='1a56a5', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

    @staticmethod
    def export_expense_import_template(filename):
        """Tạo mẫu Excel chuẩn nhiều sheet để nhập chi phí hàng loạt."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Chi phi"
        headers = [
            'Ngày', 'Dự án ID', 'Loại chi phí ID', 'Mô tả', 'Số tiền',
            'Người chi', 'Hình thức', 'Ghi chú', 'Phòng ban',
            'Mục đích sử dụng', 'Danh sách/Nội dung lên mẫu',
            'Kế toán ký', 'Trưởng phòng ký', 'Người lập', 'Hồ sơ đính kèm'
        ]
        ws.append(headers)
        ws.append([
            datetime.now().strftime('%d/%m/%Y'), '', '', 'Thuê xe cẩu chở vật tư về dự án',
            5940000.000, 'Nguyễn Văn A', 'Tiền mặt', '', 'Phòng VT-TB',
            'Phục vụ công việc tại dự án', 'Thuê xe cẩu chở vật tư',
            'Kế toán', 'Trưởng phòng', 'Người lập', 'Hợp đồng, hóa đơn, biên bản giao hàng'
        ])
        ExcelExporter._style_header(ws)
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=2, column=5).number_format = ExcelExporter.MONEY_FORMAT
        for col, width in {
            'A': 14, 'B': 12, 'C': 14, 'D': 36, 'E': 18, 'F': 18, 'G': 16, 'H': 24,
            'I': 18, 'J': 28, 'K': 34, 'L': 18, 'M': 18, 'N': 18, 'O': 34
        }.items():
            ws.column_dimensions[col].width = width
        ws.freeze_panes = 'A2'

        doc_ws = wb.create_sheet("Chung tu kem theo")
        doc_ws.append(['Dòng chi phí', 'Loại chứng từ', 'Số chứng từ', 'Ngày chứng từ', 'Nhà cung cấp', 'Nội dung', 'Số tiền', 'Ghi chú'])
        doc_ws.append([1, 'Hóa đơn', '', datetime.now().strftime('%d/%m/%Y'), '', '', 0.000, ''])
        ExcelExporter._style_header(doc_ws)
        doc_ws.cell(row=2, column=7).number_format = ExcelExporter.MONEY_FORMAT
        for col, width in {'A': 12, 'B': 20, 'C': 18, 'D': 16, 'E': 28, 'F': 36, 'G': 18, 'H': 28}.items():
            doc_ws.column_dimensions[col].width = width
        doc_ws.freeze_panes = 'A2'

        mat_ws = wb.create_sheet("Vat tu hang hoa")
        mat_ws.append(['Mã vật tư', 'Tên vật tư', 'Đơn vị', 'Số lượng', 'Đơn giá', 'Nhóm/Chủng loại', 'Nhà cung cấp', 'Ghi chú'])
        mat_ws.append(['VT001', 'Vật tư mẫu', 'cái', 1.0000, 0.000, 'Vật tư xây dựng', '', ''])
        ExcelExporter._style_header(mat_ws)
        mat_ws.cell(row=2, column=4).number_format = ExcelExporter.QUANTITY_FORMAT
        mat_ws.cell(row=2, column=5).number_format = ExcelExporter.MONEY_FORMAT
        for col, width in {'A': 14, 'B': 34, 'C': 12, 'D': 16, 'E': 18, 'F': 22, 'G': 26, 'H': 28}.items():
            mat_ws.column_dimensions[col].width = width
        mat_ws.freeze_panes = 'A2'

        guide_ws = wb.create_sheet("Huong dan")
        guide_rows = [
            ['Sheet', 'Cách dùng'],
            ['Chi phi', 'Điền mỗi dòng là một khoản chi phí. Dự án ID và Loại chi phí ID lấy ở các sheet danh mục.'],
            ['Chung tu kem theo', 'Nhập chứng từ liên quan, cột Dòng chi phí trỏ về số dòng trên sheet Chi phi.'],
            ['Vat tu hang hoa', 'Dùng khi cần nhập danh sách vật tư/hàng hóa liên quan chi phí.'],
            ['Định dạng số', 'Số tiền hiển thị 3 số thập phân; số lượng hiển thị 4 số thập phân theo chuẩn Excel máy đang dùng.'],
        ]
        for row in guide_rows:
            guide_ws.append(row)
        ExcelExporter._style_header(guide_ws)
        guide_ws.column_dimensions['A'].width = 24
        guide_ws.column_dimensions['B'].width = 110
        guide_ws['A1'].fill = PatternFill(start_color='17324D', end_color='17324D', fill_type='solid')
        guide_ws['B1'].fill = PatternFill(start_color='17324D', end_color='17324D', fill_type='solid')

        try:
            from database import get_connection
            conn = get_connection()
            cursor = conn.cursor()
            ref_projects = wb.create_sheet("DM Du an")
            ref_projects.append(['ID', 'Mã dự án', 'Tên dự án'])
            cursor.execute("SELECT id, code, name FROM projects ORDER BY name")
            for row in cursor.fetchall():
                ref_projects.append([row['id'], row['code'], row['name']])
            ExcelExporter._style_header(ref_projects)
            ref_projects.column_dimensions['A'].width = 10
            ref_projects.column_dimensions['B'].width = 18
            ref_projects.column_dimensions['C'].width = 42

            ref_cats = wb.create_sheet("DM Loai chi phi")
            ref_cats.append(['ID', 'Mã', 'Tên loại chi phí'])
            cursor.execute("SELECT id, code, name FROM expense_categories ORDER BY name")
            for row in cursor.fetchall():
                ref_cats.append([row['id'], row['code'], row['name']])
            ExcelExporter._style_header(ref_cats)
            ref_cats.column_dimensions['A'].width = 10
            ref_cats.column_dimensions['B'].width = 18
            ref_cats.column_dimensions['C'].width = 42
            conn.close()
        except Exception:
            pass

        wb.save(filename)
        return filename

    @staticmethod
    def export_materials(materials, filename='materials_export.xlsx'):
        """Xuất danh mục vật tư ra Excel."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vật tư"
            headers = ['ID', 'Mã vật tư', 'Tên vật tư', 'Đơn vị', 'Tồn kho', 'Đơn giá', 'Danh mục', 'Trạng thái']
            ws.append(headers)
            ExcelExporter._style_header(ws)
            for row in materials:
                ws.append(list(row))
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=5).number_format = ExcelExporter.QUANTITY_FORMAT
                ws.cell(row=row, column=6).number_format = ExcelExporter.MONEY_FORMAT
            for col, width in {'A': 8, 'B': 14, 'C': 30, 'D': 12, 'E': 12, 'F': 16, 'G': 24, 'H': 14}.items():
                ws.column_dimensions[col].width = width
            wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Lỗi xuất vật tư Excel: {str(e)}")

    @staticmethod
    def export_accounts(accounts, filename='accounts_export.xlsx'):
        """Xuất hệ thống tài khoản kế toán ra Excel."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tài khoản"
            headers = ['Số hiệu', 'Tên tài khoản', 'Loại', 'Cấp', 'Tài khoản cha', 'Căn cứ', 'Mô tả']
            ws.append(headers)
            ExcelExporter._style_header(ws)
            for row in accounts:
                ws.append(list(row))
            for col, width in {'A': 12, 'B': 36, 'C': 18, 'D': 8, 'E': 14, 'F': 34, 'G': 45}.items():
                ws.column_dimensions[col].width = width
            wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Lỗi xuất tài khoản Excel: {str(e)}")

    @staticmethod
    def export_ar_ap_items(items, filename='cong_no_export.xlsx'):
        """Export receivable/payable aging rows."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cong no"
            headers = [
                'ID', 'Loai', 'Doi tac', 'Du an', 'Han thanh toan',
                'So tien', 'Da thanh toan', 'Con lai', 'Trang thai', 'Nguon', 'Ghi chu'
            ]
            ws.append(headers)
            ExcelExporter._style_header(ws)
            for row in items:
                ws.append([
                    row['id'],
                    'Phai thu' if row['partner_type'] == 'customer' else 'Phai tra',
                    row['partner_name'], row['project_code'], row['due_date'],
                    float(row['amount'] or 0), float(row['paid_amount'] or 0),
                    float(row['outstanding'] or 0), row['status'],
                    f"{row['source_type']} {row['source_id']}".strip(),
                    row['notes'],
                ])
            for r in range(2, ws.max_row + 1):
                for col in (6, 7, 8):
                    ws.cell(row=r, column=col).number_format = ExcelExporter.MONEY_FORMAT
            for col, width in {
                'A': 8, 'B': 12, 'C': 30, 'D': 12, 'E': 16,
                'F': 18, 'G': 18, 'H': 18, 'I': 12, 'J': 16, 'K': 40,
            }.items():
                ws.column_dimensions[col].width = width
            wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Loi xuat cong no Excel: {str(e)}")

    @staticmethod
    def export_documents(documents, filename='documents_export.xlsx'):
        """Xuất chứng từ ra Excel."""
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Chứng từ"
            headers = ['ID', 'Loại', 'Số CT', 'Ngày', 'Nhà cung cấp', 'Nội dung', 'Số tiền', 'Dự án', 'Trạng thái', 'Chi phí ID']
            ws.append(headers)
            ExcelExporter._style_header(ws)
            for row in documents:
                ws.append(list(row))
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=7).number_format = ExcelExporter.MONEY_FORMAT
            for col in range(1, 11):
                ws.column_dimensions[chr(64 + col)].width = 18
            wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Lỗi xuất chứng từ Excel: {str(e)}")
